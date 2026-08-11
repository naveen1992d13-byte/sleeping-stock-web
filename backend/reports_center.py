"""NMTS unified Reports Center: current-month Excel reports, analytics, and archived reports."""
from __future__ import annotations

import os, re, uuid
from collections import defaultdict
from datetime import datetime, timezone, date, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pymongo import ReturnDocument

router = APIRouter(prefix="/reports-center", tags=["Reports Center"])
REPORT_STORAGE = Path(os.getenv("REPORT_STORAGE_DIR", Path(__file__).parent / "report_files"))
REPORT_STORAGE.mkdir(parents=True, exist_ok=True)
_report_security = HTTPBearer()
_AUTH_DEP = None

async def _current_user(credentials: HTTPAuthorizationCredentials = Depends(_report_security)):
    if _AUTH_DEP is None:
        raise HTTPException(500, "Reports Center authentication is not initialized")
    return await _AUTH_DEP(credentials)

CATALOGUE = [
    ("inventory", "01", "Inventory Report"), ("aging", "02", "Aging Report"),
    ("stock-value", "03", "Stock Value Report"), ("order-summary", "04", "Order Summary"),
    ("order-details", "05", "Order Details"), ("request-summary", "06", "Request Summary"),
    ("request-details", "07", "Request Details"), ("branch-summary", "08", "Branch Summary"),
    ("dealer-summary", "09", "Dealer Summary"), ("brand-summary", "10", "Brand Summary"),
    ("stock-movement", "11", "Stock Movement Report"), ("missed-opportunity", "12", "Missed Opportunity Report"),
]
CATALOGUE_MAP = {k: (n, t) for k, n, t in CATALOGUE}
BUCKETS = ["0–30 Days", "31–90 Days", "91–180 Days", "181–360 Days", "Above 360 Days"]
TERMINAL_MISSED = {"rejected", "cancelled", "closed", "not accepted", "expired"}


def init_reports_center(db, get_current_user, UserResponse):
    """Bind existing application dependencies without duplicating auth/database logic."""
    globals()["db"] = db
    globals()["get_current_user"] = get_current_user
    globals()["UserResponse"] = UserResponse
    globals()["_AUTH_DEP"] = get_current_user


def _num(v, default=0.0):
    try: return float(v if v not in (None, "") else default)
    except (TypeError, ValueError): return float(default)


def _text(*values):
    for v in values:
        if v is not None and str(v).strip(): return str(v).strip()
    return ""


def _dt(v):
    if isinstance(v, datetime): return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    if isinstance(v, date): return datetime(v.year, v.month, v.day, tzinfo=timezone.utc)
    s = str(v or "").strip()
    if not s: return None
    try: return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(timezone.utc)
    except Exception:
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try: return datetime.strptime(s[:10], f).replace(tzinfo=timezone.utc)
            except Exception: pass
    return None


def _bucket(v):
    if v in (None, "", "-"): return ""
    d = _num(v, -1)
    if d < 0: return ""
    if d <= 30: return BUCKETS[0]
    if d <= 90: return BUCKETS[1]
    if d <= 180: return BUCKETS[2]
    if d <= 360: return BUCKETS[3]
    return BUCKETS[4]


def _aging_source_date(row, aging_type):
    """Return the original source date used for aging without mixing purchase and sales."""
    if aging_type == "purchase":
        value = _text(
            row.get("last_receipt_date"), row.get("last_purchase_date"),
            row.get("receipt_date"), row.get("purchase_date")
        )
    else:
        value = _text(
            row.get("last_sales_date"), row.get("last_sale_date"),
            row.get("sales_date"), row.get("sale_date")
        )
    parsed = _dt(value)
    return parsed.date() if parsed else value


def _aging_days(row, aging_type):
    """Use stored aging first; safely derive it from its own source date for older records."""
    if aging_type == "purchase":
        candidates = (
            "purchase_aging_days", "purchase_ageing_days", "purchase_aging", "purchase_ageing",
            "purchase_aging_days_at_upload", "purchase_aging_days_at_request", "aging_days"
        )
    else:
        candidates = (
            "sales_aging_days", "sales_ageing_days", "sales_aging", "sales_ageing",
            "sales_aging_days_at_upload", "sales_aging_days_at_request", "sale_aging", "sale_ageing"
        )
    for key in candidates:
        value = row.get(key)
        if value not in (None, "", "-"):
            # Existing data may contain values such as "125 Days". Extract the numeric portion safely.
            match = re.search(r"-?\d+(?:\.\d+)?", str(value))
            if match:
                try:
                    return max(0, int(float(match.group(0))))
                except (TypeError, ValueError):
                    pass
    source = _aging_source_date(row, aging_type)
    parsed = _dt(source)
    if parsed:
        return max(0, (datetime.now(timezone.utc).date() - parsed.date()).days)
    return ""


def _period(from_date: str, to_date: str, current_only=True):
    try: start = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except Exception: raise HTTPException(400, "Invalid From Date")
    try: end = datetime.strptime(to_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
    except Exception: raise HTTPException(400, "Invalid To Date")
    today = datetime.now(timezone.utc).date()
    first = today.replace(day=1)
    if current_only and (start.date() < first or end.date() - timedelta(days=1) > today):
        raise HTTPException(400, "Reports tab permits only the current month up to today")
    if start >= end: raise HTTPException(400, "From Date must be on or before To Date")
    return start, end


def _scope(user, brand=None, dealer=None, branch=None):
    role = (user.role or "").lower()
    brand, dealer, branch = [_text(x) for x in (brand, dealer, branch)]
    is_all = lambda x: not x or x.lower().startswith("all ")
    if role == "master":
        return {k: v for k, v in (("brand", brand), ("dealer", dealer), ("branch", branch)) if not is_all(v)}
    if role == "admin":
        if brand and not is_all(brand) and brand.casefold() != _text(user.brand).casefold(): raise HTTPException(403, "Brand outside assigned scope")
        if dealer and not is_all(dealer) and dealer.casefold() != _text(user.group).casefold(): raise HTTPException(403, "Dealer outside assigned scope")
        return {"brand": _text(user.brand), "dealer": _text(user.group), **({"branch": branch} if not is_all(branch) else {})}
    if brand and not is_all(brand) and brand.casefold() != _text(user.brand).casefold(): raise HTTPException(403, "Brand outside assigned scope")
    if dealer and not is_all(dealer) and dealer.casefold() != _text(user.group).casefold(): raise HTTPException(403, "Dealer outside assigned scope")
    if branch and not is_all(branch) and branch.casefold() != _text(user.location).casefold(): raise HTTPException(403, "Branch outside assigned scope")
    return {"brand": _text(user.brand), "dealer": _text(user.group), "branch": _text(user.location)}


def _query(scope, mapping=("brand_name", "dealer_name", "branch")):
    return {field: scope[key] for key, field in zip(("brand", "dealer", "branch"), mapping) if scope.get(key)}


def _date_clause(fields, start, end):
    clauses=[]
    for f in fields:
        clauses += [{f: {"$gte": start, "$lt": end}}, {f: {"$gte": start.isoformat(), "$lt": end.isoformat()}}]
    return {"$or": clauses}


def _and(*parts):
    p=[x for x in parts if x]
    return {} if not p else p[0] if len(p)==1 else {"$and": p}


async def _inventory(scope, current=True):
    q=_query(scope)
    if current: q=_and(q, {"$or":[{"is_active_today":True},{"publish_status":"Published"}]})
    rows=await db.products.find(q,{"_id":0}).sort("part_number",1).to_list(length=300000)
    for r in rows:
        r["qty"]=_num(r.get("available_qty_number",r.get("quantity",0)))
        # NMTS stores unit value under different legacy/current field names. Use one canonical
        # Unit Value in the report and do not expose a duplicate MAV column.
        r["unit"]=_num(_text(
            r.get("unit_value_number"), r.get("unit_value"), r.get("mav_value"),
            r.get("price"), r.get("cost")
        ))
        r["total"]=_num(r.get("total_value_number",r.get("total_value",r["qty"]*r["unit"])))
        # Support all known upload field aliases while keeping one FMS / ABC report column.
        r["report_fms_abc"]=_text(
            r.get("fms_abc"), r.get("fms_and_abc"), r.get("fms_abc_value"),
            r.get("fmsabc"), r.get("fms"), r.get("abc")
        )
        r["purchase_aging_date"]=_aging_source_date(r, "purchase")
        r["sales_aging_date"]=_aging_source_date(r, "sales")
        r["pa"]=_aging_days(r, "purchase")
        r["sa"]=_aging_days(r, "sales")
    return rows


async def _orders(scope,start,end):
    q=_and(_query(scope),_date_clause(["created_at"],start,end))
    return await db.order_headers.find(q,{"_id":0}).sort("created_at",-1).to_list(length=200000)


async def _order_items(order_ids):
    if not order_ids:return []
    return await db.order_items.find({"order_id":{"$in":list(order_ids)}},{"_id":0}).to_list(length=500000)


async def _requests(scope,start,end):
    dateq=_date_clause(["requested_at","created_at"],start,end)
    side=[]
    if scope.get("brand"): side.append({"$or":[{"requesting_brand":scope["brand"]},{"supplying_brand":scope["brand"]}]})
    if scope.get("dealer"): side.append({"$or":[{"requesting_dealer":scope["dealer"]},{"supplying_dealer":scope["dealer"]}]})
    if scope.get("branch"): side.append({"$or":[{"requesting_branch":scope["branch"]},{"supplying_branch":scope["branch"]}]})
    return await db.order_requests.find(_and(dateq,*side),{"_id":0}).sort("requested_at",-1).to_list(length=500000)


def _missed(reqs, inventory):
    inv={(_text(x.get("brand_name")).casefold(),_text(x.get("dealer_name")).casefold(),_text(x.get("branch")).casefold(),_text(x.get("part_number")).casefold()):x for x in inventory if x.get("qty",0)>0}
    out=[]; seen=set()
    for r in reqs:
        key=_text(r.get("id"),r.get("request_number")+"|"+_text(r.get("part_number")))
        if key in seen: continue
        seen.add(key)
        requested=_num(r.get("requested_qty",r.get("quantity",r.get("required_qty",0))))
        supplied=max(_num(r.get("completed_qty",0)),_num(r.get("received_qty",0)),_num(r.get("dispatched_qty",0)),_num(r.get("accepted_qty",r.get("approved_qty",0))))
        unfulfilled=max(0,requested-supplied)
        status=_text(r.get("status")).lower()
        if unfulfilled<=0 or not (status in TERMINAL_MISSED or supplied<requested): continue
        ik=(_text(r.get("supplying_brand"),r.get("requesting_brand")).casefold(),_text(r.get("supplying_dealer")).casefold(),_text(r.get("supplying_branch")).casefold(),_text(r.get("part_number")).casefold())
        stock=inv.get(ik)
        if not stock: continue
        eligible=min(unfulfilled,stock["qty"])
        if eligible<=0: continue
        unit=_num(r.get("part_value",r.get("unit_value",stock["unit"])))
        out.append({**r,"unfulfilled":unfulfilled,"current_available":stock["qty"],"eligible":eligible,"unit":unit,"missed_value":eligible*unit,"stock":stock})
    return out


def _rows(report, inventory, orders, order_items, reqs):
    if report=="inventory":
        headers=["Part Number","Part Name","Available Quantity","Unit Value","Total Value","Part Category","FMS / ABC","Location / LOC","Brand","Dealer","Branch","Uploaded Date","Uploaded User","Active Status","Last Receipt Date","Purchase Aging","Last Sales Date","Sales Aging"]
        data=[[
            r.get("part_number"),
            _text(r.get("item_name"),r.get("part_name"),r.get("description")),
            r["qty"],r["unit"],r["total"],
            _text(r.get("part_category"), r.get("category"), r.get("parts_type")),
            r.get("report_fms_abc", ""),
            _text(r.get("loc"),r.get("location")),
            r.get("brand_name"),r.get("dealer_name"),r.get("branch"),
            _text(r.get("uploaded_at"),r.get("created_at"),r.get("upload_date")),
            _text(r.get("uploaded_user_name"),r.get("uploaded_by_name"),r.get("uploaded_by")),
            "Active" if r.get("is_active_today") else r.get("publish_status",""),
            r["purchase_aging_date"],r["pa"],r["sales_aging_date"],r["sa"]
        ] for r in inventory]
    elif report=="aging":
        headers=["Part Number","Part Name","Purchase Aging","Sales Aging","Purchase Aging Bucket","Sales Aging Bucket","Available Quantity","Unit Value","Total Value","Location","Brand","Dealer","Branch"]
        data=[[r.get("part_number"),r.get("item_name"),r["pa"],r["sa"],_bucket(r["pa"]),_bucket(r["sa"]),r["qty"],r["unit"],r["total"],_text(r.get("loc"),r.get("location")),r.get("brand_name"),r.get("dealer_name"),r.get("branch")] for r in inventory]
    elif report=="stock-value":
        headers=["Brand","Dealer","Branch","Part Number","Part Name","Available Quantity","Unit Value","Total Stock Value","Part Category","Purchase Aging","Sales Aging","Location"]
        data=[[r.get("brand_name"),r.get("dealer_name"),r.get("branch"),r.get("part_number"),r.get("item_name"),r["qty"],r["unit"],r["total"],r.get("part_category"),r["pa"],r["sa"],_text(r.get("loc"),r.get("location"))] for r in inventory]
    elif report=="order-summary":
        headers=["Order Number","Created Date","Created By","User Name","Brand","Dealer","Branch","Total Line Items","Requested Quantity","Available Quantity","Not Available Quantity","Reserved Quantity","Order Value","Current Order Status","Request Generated","Related Request Count"]
        grouped=defaultdict(list)
        for r in reqs: grouped[r.get("order_id")].append(r)
        data=[[o.get("order_number"),o.get("created_at"),o.get("created_by"),o.get("created_user_name"),o.get("brand_name"),o.get("dealer_name"),o.get("branch"),o.get("item_count"),o.get("total_required_qty"),o.get("total_available_qty"),o.get("total_not_available_qty"),o.get("total_reserved_qty"),o.get("total_order_value"),o.get("status"),"Yes" if grouped.get(o.get("id")) else "No",len({x.get("request_number") for x in grouped.get(o.get("id"),[]) if x.get("request_number")})] for o in orders]
    elif report=="order-details":
        om={o.get("id"):o for o in orders}; data=[]
        headers=["Order Number","Order Date","Created By","Source Brand","Source Dealer","Source Branch","Part Number","Part Name / Description","Requested Quantity","Available Quantity","Not Available Quantity","Suggested Branch","Reserved Quantity","Purchase Aging","Sales Aging","Location","Unit Value","Total Value","Item Status","Related Request Number"]
        for i in order_items:
            o=om.get(i.get("order_id"),{}); rq=_num(i.get("required_qty")); av=_num(i.get("available_qty",i.get("allocated_qty"))); unit=_num(i.get("unit_value")); alloc=i.get("allocations") or []
            data.append([o.get("order_number",i.get("order_number")),o.get("created_at"),o.get("created_user_name",o.get("created_by")),o.get("brand_name"),o.get("dealer_name"),o.get("branch"),i.get("part_number"),i.get("description"),rq,av,max(0,rq-av),_text(i.get("suggested_branch"),alloc[0].get("branch") if alloc else ""),_num(i.get("reserved_qty",i.get("allocated_qty"))),i.get("purchase_aging_days",i.get("purchase_aging")),i.get("sales_aging_days",i.get("sales_aging")),_text(i.get("loc"),i.get("location")),unit,rq*unit,i.get("status"),i.get("request_number")])
    elif report=="request-summary":
        headers=["Request Number","Reference / Order Number","Request Date","Requested By","Requested From","Requested To","Source Brand","Source Dealer","Source Branch","Destination Brand","Destination Dealer","Destination Branch","Total Line Items","Total Requested Quantity","Total Approved Quantity","Total Value","Current Status","Approved Date","Dispatched Date","Received Date","Completed Date","Rejected Date","Rejection Reason"]
        groups=defaultdict(list)
        for r in reqs: groups[_text(r.get("request_number"),r.get("id"))].append(r)
        data=[]
        for num,items in groups.items():
            r=items[0]; data.append([num,r.get("order_number"),_text(r.get("requested_at"),r.get("created_at")),r.get("requested_user_name",r.get("requested_by")),r.get("supplying_branch"),r.get("requesting_branch"),r.get("supplying_brand"),r.get("supplying_dealer"),r.get("supplying_branch"),r.get("requesting_brand"),r.get("requesting_dealer"),r.get("requesting_branch"),len(items),sum(_num(x.get("requested_qty",x.get("quantity"))) for x in items),sum(_num(x.get("accepted_qty",x.get("approved_qty"))) for x in items),sum(_num(x.get("part_value",x.get("unit_value")))*_num(x.get("requested_qty",x.get("quantity"))) for x in items),r.get("status"),r.get("approved_at"),r.get("dispatched_at"),r.get("received_at"),r.get("completed_at"),r.get("rejected_at"),_text(r.get("rejection_reason"),r.get("remarks"))])
    elif report=="request-details":
        headers=["Request Number","Reference / Order Number","Request Date","Requested By","Requested From","Requested To","Part Number","Part Name","Request Quantity","Approval Quantity","Purchase Aging","Sales Aging","Location / LOC","Part Value","Total Requested Value","Remarks","Item Status","Overall Request Status","Rejection Reason","Dispatch / Receive / Complete State"]
        data=[[r.get("request_number"),r.get("order_number"),_text(r.get("requested_at"),r.get("created_at")),r.get("requested_user_name",r.get("requested_by")),r.get("supplying_branch"),r.get("requesting_branch"),r.get("part_number"),r.get("part_name",r.get("description")),_num(r.get("requested_qty",r.get("quantity"))),_num(r.get("accepted_qty",r.get("approved_qty"))),r.get("purchase_aging_days",r.get("purchase_aging")),r.get("sales_aging_days",r.get("sales_aging")),_text(r.get("loc_at_request"),r.get("loc"),r.get("location")),_num(r.get("part_value",r.get("unit_value"))),_num(r.get("part_value",r.get("unit_value")))*_num(r.get("requested_qty",r.get("quantity"))),r.get("remarks"),r.get("item_status",r.get("status")),r.get("status"),r.get("rejection_reason")," / ".join(x for x in (r.get("dispatch_status"),r.get("receive_status"),r.get("complete_status")) if x)] for r in reqs]
    elif report=="missed-opportunity":
        headers=["Request Number","Reference / Order Number","Request Date","Requested By","Requested From","Requested To","Source Brand","Source Dealer","Source Branch","Destination Branch","Part Number","Part Name","Requested Quantity","Approved / Supplied Quantity","Unfulfilled Quantity","Current Available Quantity","Eligible Missed Quantity","Unit Value","Missed Opportunity Value","Request Status","Rejection / Failure Reason","Purchase Aging","Sales Aging","Location"]
        data=[]
        for m in _missed(reqs,inventory):
            r=m; s=m["stock"]; requested=_num(r.get("requested_qty",r.get("quantity"))); supplied=requested-m["unfulfilled"]
            data.append([r.get("request_number"),r.get("order_number"),_text(r.get("requested_at"),r.get("created_at")),r.get("requested_user_name",r.get("requested_by")),r.get("supplying_branch"),r.get("requesting_branch"),r.get("supplying_brand"),r.get("supplying_dealer"),r.get("supplying_branch"),r.get("requesting_branch"),r.get("part_number"),r.get("part_name",r.get("description")),requested,supplied,m["unfulfilled"],m["current_available"],m["eligible"],m["unit"],m["missed_value"],r.get("status"),_text(r.get("rejection_reason"),r.get("remarks")),s.get("pa"),s.get("sa"),_text(s.get("loc"),s.get("location"))])
    else:
        level={"branch-summary":"branch","dealer-summary":"dealer_name","brand-summary":"brand_name"}.get(report)
        if report=="stock-movement":
            headers=["Aging Bucket","Opening Line Items","Opening Quantity","Opening Value","Added Line Items","Added Quantity","Added Value","Reduced Line Items","Reduced Quantity","Reduced Value","Closing Line Items","Closing Quantity","Closing Value","Balance"]
            data=[]
            for b in BUCKETS:
                rr=[x for x in inventory if _bucket(x["pa"])==b]; q=sum(x["qty"] for x in rr); v=sum(x["total"] for x in rr)
                data.append([b,0,0,0,len(rr),q,v,0,0,0,len(rr),q,v,v])
            return headers,data
        headers=["Brand","Dealer","Branch / Scope","Current Inventory Line Items","Current Available Quantity","Current Inventory Value","0–30 Day Value","31–90 Day Value","91–180 Day Value","181–360 Day Value","Above 360 Day Value","Orders Created","Order Line Items","Requests Sent","Requests Received","Requested Quantity","Approved Quantity","Rejected Request Count","Completed Request Count","Completed Transfer Value","Current Missed Opportunity Value"]
        groups=defaultdict(list)
        for x in inventory: groups[_text(x.get(level))].append(x)
        missed=_missed(reqs,inventory)
        data=[]
        for key,items in groups.items():
            first=items[0]; branches={x.get("branch") for x in items}; scope_req=[r for r in reqs if key in (_text(r.get("requesting_branch")),_text(r.get("supplying_branch")),_text(r.get("requesting_dealer")),_text(r.get("supplying_dealer")),_text(r.get("requesting_brand")),_text(r.get("supplying_brand")))]
            scope_orders=[o for o in orders if key in (_text(o.get("branch")),_text(o.get("dealer_name")),_text(o.get("brand_name")))]
            vals=[sum(x["total"] for x in items if _bucket(x["pa"])==b) for b in BUCKETS]
            data.append([first.get("brand_name"),first.get("dealer_name"),key if level!="branch" else first.get("branch"),len(items),sum(x["qty"] for x in items),sum(x["total"] for x in items),*vals,len(scope_orders),sum(_num(o.get("item_count")) for o in scope_orders),len({r.get("request_number") for r in scope_req if _text(r.get("requesting_branch")) in branches}),len({r.get("request_number") for r in scope_req if _text(r.get("supplying_branch")) in branches}),sum(_num(r.get("requested_qty",r.get("quantity"))) for r in scope_req),sum(_num(r.get("accepted_qty",r.get("approved_qty"))) for r in scope_req),len({r.get("request_number") for r in scope_req if _text(r.get("status")).lower()=="rejected"}),len({r.get("request_number") for r in scope_req if _text(r.get("status")).lower()=="completed"}),sum(_num(r.get("part_value",r.get("unit_value")))*_num(r.get("accepted_qty",r.get("approved_qty"))) for r in scope_req if _text(r.get("status")).lower()=="completed"),sum(m["missed_value"] for m in missed if key in (_text(m.get("supplying_branch")),_text(m.get("supplying_dealer")),_text(m.get("supplying_brand"))))])
    return headers,data


def _workbook(title,number,headers,data,meta,summary=None):
    wb=openpyxl.Workbook(write_only=False); ws=wb.active; ws.title=re.sub(r'[:\\/?*\[\]]','',title)[:31]
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(1,len(headers))); ws.cell(1,1,f"NMTS – {number} {title}").font=Font(size=16,bold=True,color="FFFFFF"); ws.cell(1,1).fill=PatternFill("solid",fgColor="059669"); ws.cell(1,1).alignment=Alignment(horizontal="center")
    meta_rows=[("Selected Period",f"{meta['from']} to {meta['to']}"),("Generated At",meta['generated']),("Generated By",meta['user']),("Brand",meta.get('brand','All Brands')),("Dealer",meta.get('dealer','All Dealers')),("Branch",meta.get('branch','All Branches'))]
    row=3
    for k,v in meta_rows: ws.cell(row,1,k).font=Font(bold=True); ws.cell(row,2,v); row+=1
    hr=row+1
    for c,h in enumerate(headers,1):
        cell=ws.cell(hr,c,h); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="374151"); cell.alignment=Alignment(horizontal="center",wrap_text=True)
    money_terms=("Value","Unit Value","MAV","Balance")
    for ri,values in enumerate(data,hr+1):
        for ci,v in enumerate(values,1):
            cell=ws.cell(ri,ci,v)
            if isinstance(v,(int,float)):
                cell.number_format='#,##0.00' if any(t in str(headers[ci-1]) for t in money_terms) else '#,##0.00'
    if data:
        total_row=hr+len(data)+1; ws.cell(total_row,1,"TOTAL").font=Font(bold=True)
        for ci,h in enumerate(headers,1):
            if any(x in h for x in ("Quantity","Value","Line Items","Count","Orders","Requests")):
                letter=get_column_letter(ci); ws.cell(total_row,ci,f"=SUM({letter}{hr+1}:{letter}{total_row-1})").font=Font(bold=True); ws.cell(total_row,ci).number_format='#,##0.00'
    ws.freeze_panes=f"A{hr+1}"; ws.auto_filter.ref=f"A{hr}:{get_column_letter(len(headers))}{hr+len(data)}"
    for ci,h in enumerate(headers,1): ws.column_dimensions[get_column_letter(ci)].width=min(40,max(12,len(str(h))+3))
    ws.sheet_view.showGridLines=False
    return wb

@router.get("/catalogue")
async def catalogue(current_user=Depends(_current_user)):
    return [{"key":k,"number":n,"name":t} for k,n,t in CATALOGUE]

@router.get("/scope-options")
async def scope_options(brand:Optional[str]=None,dealer:Optional[str]=None,current_user=Depends(_current_user)):
    """Reports-specific scope endpoint kept for API compatibility.

    It mirrors the existing NMTS master-data field variants and performs
    case-insensitive matching, preventing a selected Brand from incorrectly
    clearing valid Dealers created under older schemas.
    """
    user=current_user; scope=_scope(user,brand,dealer,None)
    active={"$or":[{"status":"active"},{"status":{"$exists":False}},{"status":None},{"status":""}]}
    bq=active if user.role=="master" else _and(active,{"name":{"$regex":f"^{re.escape(_text(user.brand))}$","$options":"i"}})
    brands=await db.brands.find(bq,{"_id":0,"name":1}).sort("name",1).to_list(1000)

    selected=scope.get("brand")
    dq=active
    if selected:
        rx={"$regex":f"^{re.escape(selected)}$","$options":"i"}
        dq=_and(dq,{"$or":[{"brand":rx},{"brand_name":rx},{"brandName":rx}]})
    if user.role!="master":
        dq=_and(active,{"name":{"$regex":f"^{re.escape(_text(user.group))}$","$options":"i"}})
    dealers=await db.dealers.find(dq,{"_id":0,"name":1,"brand":1,"brand_name":1,"brandName":1}).sort("name",1).to_list(1000)

    selected_d=scope.get("dealer")
    brq=active
    clauses=[]
    if selected:
        rx={"$regex":f"^{re.escape(selected)}$","$options":"i"}
        clauses.append({"$or":[{"brand":rx},{"brand_name":rx},{"brandName":rx},{"brand":{"$exists":False}},{"brand_name":{"$exists":False}}]})
    if selected_d:
        rx={"$regex":f"^{re.escape(selected_d)}$","$options":"i"}
        clauses.append({"$or":[{"dealer":rx},{"dealer_name":rx},{"dealerName":rx}]})
    if clauses: brq=_and(brq,*clauses)
    if user.role=="user": brq=_and(active,{"name":{"$regex":f"^{re.escape(_text(user.location))}$","$options":"i"}})
    branches=await db.branches.find(brq,{"_id":0,"name":1,"dealer":1,"dealer_name":1,"dealerName":1,"brand":1,"brand_name":1,"brandName":1}).sort("name",1).to_list(5000)
    return {"brands":[x.get("name") for x in brands if x.get("name")],"dealers":[x.get("name") for x in dealers if x.get("name")],"branches":[x.get("name") for x in branches if x.get("name")]}

@router.get("/export/{report}")
async def export_report(report:str,from_date:str,to_date:str,brand:Optional[str]=None,dealer:Optional[str]=None,branch:Optional[str]=None,current_user=Depends(_current_user)):
    try:
        import excel_permissions
    except ImportError:
        from . import excel_permissions
    excel_permissions.require_excel_export(current_user)
    if report not in CATALOGUE_MAP: raise HTTPException(404,"Unknown report")
    start,end=_period(from_date,to_date,True); scope=_scope(current_user,brand,dealer,branch)
    inventory=await _inventory(scope); orders=await _orders(scope,start,end); items=await _order_items({o.get("id") for o in orders}); reqs=await _requests(scope,start,end)
    headers,data=_rows(report,inventory,orders,items,reqs); number,title=CATALOGUE_MAP[report]
    wb=_workbook(title,number,headers,data,{"from":from_date,"to":to_date,"generated":datetime.now(timezone.utc).isoformat(),"user":current_user.username,"brand":scope.get("brand","All Brands"),"dealer":scope.get("dealer","All Dealers"),"branch":scope.get("branch","All Branches")})
    out=BytesIO(); wb.save(out); out.seek(0); filename=f"NMTS_{re.sub('[^A-Za-z0-9]+','_',title).strip('_')}_{from_date}_{to_date}.xlsx"
    return StreamingResponse(out,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{filename}"'})

@router.get("/analytics")
async def analytics(from_date:str,to_date:str,brand:Optional[str]=None,dealer:Optional[str]=None,branch:Optional[str]=None,current_user=Depends(_current_user)):
    start,end=_period(from_date,to_date,True); scope=_scope(current_user,brand,dealer,branch)
    inv=await _inventory(scope); orders=await _orders(scope,start,end); reqs=await _requests(scope,start,end); missed=_missed(reqs,inv)
    aging=[]
    for b in BUCKETS:
        rows=[x for x in inv if _bucket(x["pa"])==b]; aging.append({"bucket":b,"quantity":sum(x["qty"] for x in rows),"value":sum(x["total"] for x in rows),"line_items":len(rows)})
    branches=[]
    for name in sorted({_text(x.get("branch")) for x in inv if x.get("branch")}):
        rows=[x for x in inv if _text(x.get("branch"))==name]; rq=[r for r in reqs if name in (_text(r.get("requesting_branch")),_text(r.get("supplying_branch")))]
        branches.append({"name":name,"stock_value":sum(x["total"] for x in rows),"above_360":sum(x["total"] for x in rows if _bucket(x["pa"])==BUCKETS[-1]),"orders":sum(1 for o in orders if _text(o.get("branch"))==name),"requests_sent":len({r.get("request_number") for r in rq if _text(r.get("requesting_branch"))==name}),"requests_received":len({r.get("request_number") for r in rq if _text(r.get("supplying_branch"))==name}),"completed_value":sum(_num(r.get("part_value",r.get("unit_value")))*_num(r.get("accepted_qty",r.get("approved_qty"))) for r in rq if _text(r.get("status")).lower()=="completed"),"missed_value":sum(m["missed_value"] for m in missed if _text(m.get("supplying_branch"))==name)})
    daily=defaultdict(lambda:{"orders":0,"requests":set(),"completed":set(),"transfer_value":0,"uploads":0})
    for o in orders:
        d=(_dt(o.get("created_at")) or start).date().isoformat(); daily[d]["orders"]+=1
    for r in reqs:
        d=(_dt(r.get("requested_at") or r.get("created_at")) or start).date().isoformat(); daily[d]["requests"].add(r.get("request_number"));
        if _text(r.get("status")).lower()=="completed": daily[d]["completed"].add(r.get("request_number")); daily[d]["transfer_value"]+=_num(r.get("part_value",r.get("unit_value")))*_num(r.get("accepted_qty",r.get("approved_qty")))
    result_daily=[{"date":d,"orders":v["orders"],"requests":len(v["requests"]),"completed":len(v["completed"]),"transfer_value":v["transfer_value"]} for d,v in sorted(daily.items())]
    pa=[_num(x["pa"]) for x in inv if x["pa"] not in (None,"")]; sa=[_num(x["sa"]) for x in inv if x["sa"] not in (None,"")]
    statuses=defaultdict(set)
    for r in reqs: statuses[_text(r.get("status"),"Unknown")].add(r.get("request_number"))
    return {"scope":scope,"stock":{"line_items":len(inv),"quantity":sum(x["qty"] for x in inv),"value":sum(x["total"] for x in inv),"avg_purchase_aging":sum(pa)/len(pa) if pa else 0,"avg_sales_aging":sum(sa)/len(sa) if sa else 0},"aging":aging,"movement":{"added_line_items":len(inv),"added_quantity":sum(x["qty"] for x in inv),"added_value":sum(x["total"] for x in inv),"reduced_line_items":0,"reduced_quantity":0,"reduced_value":0},"orders":{"total":len(orders),"line_items":sum(_num(o.get("item_count")) for o in orders),"requested_qty":sum(_num(o.get("total_required_qty")) for o in orders),"available_qty":sum(_num(o.get("total_available_qty")) for o in orders),"not_available_qty":sum(_num(o.get("total_not_available_qty")) for o in orders),"reserved_qty":sum(_num(o.get("total_reserved_qty")) for o in orders),"value":sum(_num(o.get("total_order_value")) for o in orders)},"requests":{"total":len({r.get("request_number") for r in reqs}),"requested_qty":sum(_num(r.get("requested_qty",r.get("quantity"))) for r in reqs),"approved_qty":sum(_num(r.get("accepted_qty",r.get("approved_qty"))) for r in reqs),"statuses":{k:len(v) for k,v in statuses.items()},"value":sum(_num(r.get("part_value",r.get("unit_value")))*_num(r.get("requested_qty",r.get("quantity"))) for r in reqs),"completed_value":sum(_num(r.get("part_value",r.get("unit_value")))*_num(r.get("accepted_qty",r.get("approved_qty"))) for r in reqs if _text(r.get("status")).lower()=="completed")},"missed":{"requests":len({m.get("request_number") for m in missed}),"line_items":len(missed),"quantity":sum(m["eligible"] for m in missed),"value":sum(m["missed_value"] for m in missed)},"branches":branches,"daily":result_daily}

async def _next_old_number():
    key=datetime.now(timezone.utc).strftime("%y%m%d"); doc=await db.counters.find_one_and_update({"_id":f"old_report_{key}"},{"$inc":{"seq":1}},upsert=True,return_document=ReturnDocument.AFTER); return f"OR{key}{int(doc.get('seq',1)):04d}"

@router.post("/old-requests")
async def create_old_request(payload:Dict[str,Any],current_user=Depends(_current_user)):
    start,end=_period(payload.get("from_date"),payload.get("to_date"),False); scope=_scope(current_user,payload.get("brand"),payload.get("dealer"),payload.get("branch"))
    if start.date()>=datetime.now(timezone.utc).date().replace(day=1): raise HTTPException(400,"Use Reports tab for current-month reports")
    now=datetime.now(timezone.utc); doc={"id":str(uuid.uuid4()),"request_number":await _next_old_number(),"owner_id":current_user.id,"owner_name":current_user.username,"owner_role":current_user.role,"report_type":payload.get("report_type"),"from_date":payload.get("from_date"),"to_date":payload.get("to_date"),"scope":scope,"aging_type":payload.get("aging_type","Both"),"required_format":payload.get("required_format","Excel"),"reason":payload.get("reason",""),"status":"Requested","created_at":now,"updated_at":now,"download_count":0}
    await db.old_report_requests.insert_one(doc.copy()); doc.pop("_id",None); return doc

@router.get("/old-requests")
async def list_old_requests(page:int=1,page_size:int=20,current_user=Depends(_current_user)):
    q={} if current_user.role=="master" else ({"owner_id":current_user.id} if current_user.role=="user" else {"$or":[{"owner_id":current_user.id},{"scope.brand":current_user.brand,"scope.dealer":current_user.group}]})
    total=await db.old_report_requests.count_documents(q); rows=await db.old_report_requests.find(q,{"_id":0,"storage_path":0}).sort("created_at",-1).skip(max(0,page-1)*page_size).limit(min(page_size,100)).to_list(length=100)
    return {"records":rows,"total":total,"page":page,"page_size":page_size}

@router.patch("/old-requests/{request_number}")
async def update_old_request(request_number:str,payload:Dict[str,Any],current_user=Depends(_current_user)):
    if current_user.role!="master": raise HTTPException(403,"Master Admin only")
    allowed={"Under Review","Processing","Ready for Download","Rejected","Cancelled","Expired"}; status=payload.get("status")
    if status and status not in allowed: raise HTTPException(400,"Invalid status")
    if status=="Rejected" and not _text(payload.get("admin_remarks"),payload.get("rejection_reason")): raise HTTPException(400,"Rejection reason is mandatory")
    update={k:v for k,v in payload.items() if k in {"status","admin_remarks","expiry_date"}}; update["updated_at"]=datetime.now(timezone.utc)
    result=await db.old_report_requests.find_one_and_update({"request_number":request_number},{"$set":update},return_document=ReturnDocument.AFTER,projection={"_id":0,"storage_path":0})
    if not result: raise HTTPException(404,"Request not found")
    return result

@router.post("/old-requests/{request_number}/upload")
async def upload_old_report(request_number:str,file:UploadFile=File(...),expiry_date:Optional[str]=Form(None),admin_remarks:Optional[str]=Form(None),current_user=Depends(_current_user)):
    if current_user.role!="master": raise HTTPException(403,"Master Admin only")
    ext=Path(file.filename or "").suffix.lower()
    if ext not in {".xlsx",".xls",".zip"}: raise HTTPException(400,"Only Excel or ZIP files are allowed")
    req=await db.old_report_requests.find_one({"request_number":request_number})
    if not req: raise HTTPException(404,"Request not found")
    content=await file.read(); safe=f"{request_number}_{uuid.uuid4().hex[:8]}{ext}"
    try:
        import file_objects
    except ImportError:
        from . import file_objects
    stored=await file_objects.store_bytes(module="old-reports",relative_key=f"{request_number}/{safe}",data=content,original_filename=Path(file.filename).name,content_type=file.content_type or "application/octet-stream")
    path=REPORT_STORAGE/safe; path.write_bytes(content)  # keep local copy for legacy readers during transition
    expiry=_dt(expiry_date) if expiry_date else datetime.now(timezone.utc)+timedelta(days=30)
    update={"status":"Ready for Download","file_name":Path(file.filename).name,"storage_path":str(path),"storage_provider":stored.get("storage_provider"),"storage_key":stored.get("storage_key"),"content_type":stored.get("content_type"),"sha256":stored.get("sha256"),"archived_at":stored.get("archived_at"),"file_size":len(content),"uploaded_by":current_user.id,"uploaded_by_name":current_user.username,"uploaded_at":datetime.now(timezone.utc),"expiry_date":expiry,"admin_remarks":admin_remarks or "","updated_at":datetime.now(timezone.utc)}
    await db.old_report_requests.update_one({"request_number":request_number},{"$set":update}); return {"message":"File uploaded","request_number":request_number}

@router.get("/old-requests/{request_number}/download")
async def download_old_report(request_number:str,current_user=Depends(_current_user)):
    try:
        import excel_permissions, file_objects
    except ImportError:
        from . import excel_permissions, file_objects
    excel_permissions.require_excel_export(current_user)
    req=await db.old_report_requests.find_one({"request_number":request_number})
    if not req: raise HTTPException(404,"Request not found")
    _scope(current_user,req.get("scope",{}).get("brand"),req.get("scope",{}).get("dealer"),req.get("scope",{}).get("branch"))
    if current_user.role!="master" and req.get("owner_id")!=current_user.id and current_user.role=="user": raise HTTPException(403,"Not authorized")
    if req.get("status")!="Ready for Download": raise HTTPException(409,"Report is not ready")
    expiry=_dt(req.get("expiry_date"));
    if expiry and expiry<datetime.now(timezone.utc): await db.old_report_requests.update_one({"id":req["id"]},{"$set":{"status":"Expired"}}); raise HTTPException(410,"Report has expired")
    if not file_objects.meta_has_readable_bytes(req):
        raise HTTPException(404,"Stored report file not found")
    now=datetime.now(timezone.utc); await db.old_report_requests.update_one({"id":req["id"]},{"$inc":{"download_count":1},"$set":{"last_downloaded_at":now}}); await db.old_report_downloads.insert_one({"id":str(uuid.uuid4()),"request_number":request_number,"file_name":req.get("file_name"),"downloaded_by":current_user.id,"downloaded_by_name":current_user.username,"downloaded_at":now})
    return file_objects.streaming_response_from_meta(req, filename=req.get("file_name") or "report.bin")

async def ensure_indexes():
    await db.old_report_requests.create_index("request_number",unique=True)
    await db.old_report_requests.create_index([("owner_id",1),("created_at",-1)])
    await db.old_report_requests.create_index([("scope.brand",1),("scope.dealer",1),("scope.branch",1),("status",1)])
    await db.old_report_downloads.create_index([("request_number",1),("downloaded_at",-1)])
