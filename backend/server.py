from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form, WebSocket, WebSocketDisconnect
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument
from pymongo.errors import DuplicateKeyError, OperationFailure
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, field_validator
from typing import List, Optional
import uuid
import hashlib
import asyncio
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from passlib.context import CryptContext
from jose import JWTError, jwt
import openpyxl
from io import BytesIO
import socketio
import re
try:
    from . import notifications
except ImportError:
    import notifications
try:
    from . import order_desk_workflow as odw
except ImportError:
    import order_desk_workflow as odw
try:
    from . import s3_storage
    from . import file_objects
    from . import archive_manifest
    from . import history_archive
    from . import hybrid_history
    from . import archive_scheduler
    from . import excel_permissions
except ImportError:
    import s3_storage
    import file_objects
    import archive_manifest
    import history_archive
    import hybrid_history
    import archive_scheduler
    import excel_permissions

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.getenv("SECRET_KEY", "nmts-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440  # 24 hours

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Socket.IO for real-time notifications
sio = socketio.AsyncServer(async_mode='asgi', cors_allowed_origins='*')
socket_app = socketio.ASGIApp(sio, app)

# ==================== MODELS ====================

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str = ""  # Custom user ID in pattern SSYYMMDD##
    username: str
    email: str
    password: str
    role: str  # master, admin, user
    phone: str = ""
    state: str = ""
    brand: str = ""
    group: str = ""
    location: str = ""
    status: str = "active"  # active, inactive
    permissions: dict = {}  # Menu permissions
    last_login: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    user_id: str = ""
    username: str
    email: str
    password: str
    role: str
    phone: str = ""
    state: str = ""
    brand: str = ""
    group: str = ""
    location: str = ""
    joining_date: str = ""
    permissions: dict = {}

class UserUpdate(BaseModel):
    email: Optional[str] = None
    phone: Optional[str] = None
    username: Optional[str] = None
    brand: Optional[str] = None
    group: Optional[str] = None
    location: Optional[str] = None
    permissions: Optional[dict] = None

class PasswordChange(BaseModel):
    old_password: str
    new_password: str

class ActivityLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    action: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))



def normalize_permissions(value):
    """Return permissions as a clean List[str] regardless of DB storage format."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, dict):
        return [str(key).strip() for key, allowed in value.items() if allowed and str(key).strip()]
    if isinstance(value, str):
        return [value.strip()] if value.strip() else []
    return []


def normalize_email(value: str) -> str:
    return (value or "").strip().lower()


class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str = ""
    username: str
    email: str
    role: str
    phone: str = ""
    state: str = ""
    brand: str = ""
    group: str = ""
    location: str = ""
    status: str = "active"
    permissions: List[str] = []
    last_login: Optional[datetime] = None
    created_at: datetime

    @field_validator("permissions", mode="before")
    @classmethod
    def clean_permissions(cls, value):
        return normalize_permissions(value)

# Brand and Group models
class Brand(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Group(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BrandCreate(BaseModel):
    name: str

class GroupCreate(BaseModel):
    name: str

class LoginRequest(BaseModel):
    # Field name kept as "email" for backward compatibility with the existing
    # API contract and test suite; the value may be either a User ID or an
    # Email address — see login() below, which tries User ID first.
    email: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    item_name: str
    part_number: str
    quantity: float  # Changed to float to support decimal quantities
    price: float
    category: str = ""
    upload_method: str = "manual"  # manual or excel
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    item_name: str
    part_number: str
    quantity: float  # Changed to float
    price: float
    category: str = ""

class ProductUpdate(BaseModel):
    item_name: Optional[str] = None
    part_number: Optional[str] = None
    quantity: Optional[float] = None  # Changed to float
    price: Optional[float] = None
    category: Optional[str] = None

class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    quantity: int
    total_price: float
    status: str = "pending"  # pending, approved, rejected
    assigned_to: str  # admin id
    created_by: str  # admin id
    remarks: str = ""
    is_seen: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OrderCreate(BaseModel):
    product_id: str
    quantity: int
    assigned_to: str

class OrderAction(BaseModel):
    status: str  # approved or rejected
    remarks: Optional[str] = ""

class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_id: str
    recipient_id: str
    message: str
    status: str = "pending"
    is_read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UploadLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    file_name: str
    uploaded_by: str
    status: str
    rows_processed: int
    rows_imported: int
    errors: List[str] = []
    upload_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise credentials_exception
    return UserResponse(**user)

# ==================== SOCKET.IO ====================

connected_users = {}  # {user_id: sid}

@sio.event
async def connect(sid, environ):
    logging.info(f"Client connected: {sid}")

@sio.event
async def register(sid, data):
    user_id = data.get("user_id")
    if user_id:
        connected_users[user_id] = sid
        logging.info(f"User {user_id} registered with socket {sid}")

# ==================== UPLOAD HELPERS ====================

def normalize_header(value: str) -> str:
    return str(value or "").strip().lower().replace("_", " ").replace("-", " ")


def parse_excel_date(value):
    """Return a date object from common Excel/date text formats, or None."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def format_date_for_display(value):
    parsed = parse_excel_date(value)
    return parsed.strftime("%d-%m-%Y") if parsed else ""


def calculate_ageing_days(value):
    parsed = parse_excel_date(value)
    if not parsed:
        return None
    today = datetime.now(timezone.utc).date()
    return max((today - parsed).days, 0)


async def resolve_brand_code_for_upload(brand_value: str) -> str:
    """Resolve brand code from Brand Master; fallback to first two cleaned letters."""
    value = (brand_value or "").strip()
    if not value:
        return "XX"

    brand = await db.brands.find_one({
        "$or": [
            {"name": {"$regex": f"^{re.escape(value)}$", "$options": "i"}},
            {"code": {"$regex": f"^{re.escape(value)}$", "$options": "i"}},
        ]
    }, {"_id": 0})

    code = (brand or {}).get("code") or value
    clean_code = re.sub(r"[^A-Za-z0-9]", "", str(code)).upper()[:2]
    return clean_code or "XX"


async def generate_upload_no(upload_type: str, brand_code: str):
    """Generate upload number without hyphen: PUHY260705001 / OUHY260705001."""
    today_key = datetime.now(timezone.utc).strftime("%y%m%d")
    type_code = "PU" if upload_type == "product" else "OU"
    clean_brand_code = re.sub(r"[^A-Za-z0-9]", "", str(brand_code or "XX")).upper()[:2] or "XX"
    counter_id = f"upload_{type_code}_{clean_brand_code}_{today_key}"
    counter = await db.counters.find_one_and_update(
        {"_id": counter_id},
        {"$inc": {"seq": 1}, "$setOnInsert": {"date_key": today_key, "type": type_code, "brand_code": clean_brand_code}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    seq = int(counter.get("seq", 1))
    return f"{type_code}{clean_brand_code}{today_key}{seq:03d}"


async def get_user_upload_context(current_user: UserResponse):
    """Derive upload context from logged-in user and Brand Master."""
    user_code = current_user.user_id or current_user.id
    brand_name = current_user.brand or ""
    dealer_name = current_user.group or ""
    branch_name = current_user.location or ""
    brand_code = await resolve_brand_code_for_upload(brand_name)
    return {
        "user_code": user_code,
        "uploaded_user_id": current_user.id,
        "uploaded_user_name": current_user.username,
        "dealer_code": dealer_name or user_code,
        "dealer_name": dealer_name,
        "brand_code": brand_code,
        "brand_name": brand_name,
        "brand": brand_name,
        "branch": branch_name,
        "location": branch_name,
    }

@sio.event
async def disconnect(sid):
    # Remove from connected_users
    for user_id, socket_id in list(connected_users.items()):
        if socket_id == sid:
            del connected_users[user_id]
            logging.info(f"User {user_id} disconnected")
            break

async def send_notification_to_user(user_id: str, notification_data: dict):
    if user_id in connected_users:
        sid = connected_users[user_id]
        await sio.emit("new_notification", notification_data, room=sid)

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/login", response_model=Token)
async def login(login_data: LoginRequest):
    # Single field accepts User ID or Email. Trim spaces first.
    # Business rule: search user_id first, then email case-insensitive.
    identifier = (login_data.email or "").strip()
    if not identifier:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect User ID/Email or password"
        )

    user = await db.users.find_one({"user_id": identifier}, {"_id": 0})

    if not user:
        escaped_uid = re.escape(identifier)
        user = await db.users.find_one(
            {"user_id": {"$regex": f"^{escaped_uid}$", "$options": "i"}},
            {"_id": 0},
        )

    if not user:
        escaped_email = re.escape(normalize_email(identifier))
        user = await db.users.find_one(
            {"email": {"$regex": f"^{escaped_email}$", "$options": "i"}},
            {"_id": 0}
        )

    stored_hash = user.get("password") if user else None
    if not user or not stored_hash or not verify_password(login_data.password, stored_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect User ID/Email or password"
        )

    if str(user.get("status") or "").strip().lower() == "inactive":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account is inactive. Please contact administrator."
        )

    now = datetime.now(timezone.utc).isoformat()
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"last_login": now, "lastLogin": now}}
    )
    user["last_login"] = now
    user["permissions"] = normalize_permissions(user.get("permissions"))

    activity_log = ActivityLog(user_id=user["id"], action="login")
    log_doc = activity_log.model_dump()
    log_doc['created_at'] = log_doc['created_at'].isoformat()
    await db.activity_logs.insert_one(log_doc)

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["id"]}, expires_delta=access_token_expires
    )

    user_response = UserResponse(**user)
    return Token(access_token=access_token, token_type="bearer", user=user_response)

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: UserResponse = Depends(get_current_user)):
    return current_user

# ==================== USER ROUTES ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    users = await db.users.find({}, {"_id": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]

@api_router.post("/users", response_model=UserResponse)
async def create_user(user_data: UserCreate, current_user: UserResponse = Depends(get_current_user)):
    # Master can create admins and users, Admin can only create users
    if current_user.role == "master":
        if user_data.role not in ["master", "admin", "user"]:
            raise HTTPException(status_code=400, detail="Invalid role")
    elif current_user.role == "admin":
        if user_data.role != "user":
            raise HTTPException(status_code=403, detail="Admins can only create users")
    else:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user_data.email = normalize_email(user_data.email)
    user_data.user_id = (user_data.user_id or "").strip()

    state_record = await _resolve_master_record("states", user_data.state, "State")
    brand_record = await _resolve_master_record("brands", user_data.brand, "Brand")

    if not user_data.user_id:
        user_data.user_id = await _build_next_user_id(state_record["code"], brand_record["code"])

    # Store display names in user records, but generate IDs only from master codes.
    user_data.state = state_record["name"]
    user_data.brand = brand_record["name"]

    # Check if user exists
    escaped_email = re.escape(user_data.email)
    existing = await db.users.find_one({"email": {"$regex": f"^{escaped_email}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="User already exists")
    
    # Check if user_id already exists
    if user_data.user_id:
        existing_user_id = await db.users.find_one({"user_id": user_data.user_id})
        if existing_user_id:
            raise HTTPException(status_code=400, detail="User ID already exists")
    
    user = User(
        user_id=user_data.user_id,
        username=user_data.username,
        email=user_data.email,
        password=hash_password(user_data.password),
        role=user_data.role,
        phone=user_data.phone,
        state=user_data.state,
        brand=user_data.brand,
        group=user_data.group,
        location=user_data.location,
        permissions=user_data.permissions,
        status="active"
    )
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    return UserResponse(**user.model_dump())

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role != "master":
        raise HTTPException(status_code=403, detail="Only master can delete users")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# Update user permissions
class PermissionUpdate(BaseModel):
    permissions: dict

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, data: PermissionUpdate, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"permissions": data.permissions}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Permissions updated successfully"}

@api_router.get("/users/{user_id}/permissions")
async def get_user_permissions(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"permissions": user.get("permissions", {})}

# ==================== PRODUCT ROUTES ====================

@api_router.get("/products")
async def get_products(current_user: UserResponse = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    # Return raw data to preserve all fields including location, branch, mav_value, etc.
    return products

@api_router.get("/products/low-stock")
async def get_low_stock(threshold: int = 10, current_user: UserResponse = Depends(get_current_user)):
    products = await db.products.find({"quantity": {"$lte": threshold}}, {"_id": 0}).to_list(1000)
    return products

@api_router.get("/products/sleeping-stock")
async def get_sleeping_stock(days: int = 90, current_user: UserResponse = Depends(get_current_user)):
    # Products with no orders in last X days
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=days)
    
    # Get all products
    all_products = await db.products.find({}, {"_id": 0}).to_list(1000)
    
    sleeping = []
    for product in all_products:
        # Check if product has any recent orders
        recent_order = await db.orders.find_one({
            "product_id": product["id"],
            "created_at": {"$gte": cutoff_date.isoformat()}
        })
        if not recent_order:
            sleeping.append(product)
    
    return sleeping

@api_router.post("/products", response_model=Product)
async def create_product(product_data: ProductCreate, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if part number exists
    existing = await db.products.find_one({"part_number": product_data.part_number})
    if existing:
        raise HTTPException(status_code=400, detail="Part number already exists")
    
    product = Product(**product_data.model_dump())
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.products.insert_one(doc)
    
    return product

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product_data: ProductUpdate, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {k: v for k, v in product_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.products.update_one({"id": product_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    return Product(**product)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product deleted successfully"}

# ==================== UPLOAD ROUTES ====================

async def process_product_upload(file: UploadFile, current_user: UserResponse):
    if current_user.role not in ["master", "admin", "user"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
    upload_dt = datetime.now(timezone.utc)
    upload_date_str = upload_dt.strftime('%d-%m-%Y')
    upload_time_str = upload_dt.strftime('%H:%M:%S')
    user_ctx = await get_user_upload_context(current_user)
    upload_no = await generate_upload_no("product", user_ctx["brand_code"])

    content = await file.read()
    file_size = len(content)
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    header_lower = [normalize_header(h) for h in headers]
    col_mapping = {}
    
    column_definitions = {
        'part_number': ['part no', 'part number', 'partno'],
        'item_name': ['part name', 'item name', 'partname', 'name'],
        'quantity': ['available qty', 'available quantity', 'qty', 'quantity'],
        'location': ['bin location', 'location', 'loc'],
        'cost': ['cost', 'purchase cost', 'landing cost'],
        'mav_value': ['mav value', 'mavvalue', 'mrp', 'price', 'value'],
        'parts_type': ['part types', 'parts type', 'part type', 'type'],
        'fms_abc': ['fms abc', 'fms & abc', 'fms', 'abc'],
        'last_receipt_date': ['last receipt date', 'last purchase date', 'receipt date', 'purchase date'],
        'last_sales_date': ['last sales date', 'last sale date', 'sales date', 'sale date'],
    }
    
    for field, possible_names in column_definitions.items():
        for i, header in enumerate(header_lower):
            if header in possible_names:
                col_mapping[field] = i
                break
    
    if 'part_number' not in col_mapping:
        raise HTTPException(status_code=400, detail="Required column 'Part No' not found")
    if 'item_name' not in col_mapping:
        raise HTTPException(status_code=400, detail="Required column 'Part Name' not found")
    
    errors = []
    rows_processed = 0
    rows_imported = 0
    preview_data = []
    
    def safe_float(value, default=0.0):
        try:
            if value in (None, ''):
                return default
            return float(value)
        except Exception:
            return default
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        rows_processed += 1
        
        try:
            def get_cell_value(field, default=None):
                if field not in col_mapping:
                    return default
                col_idx = col_mapping[field]
                if col_idx < len(row):
                    val = row[col_idx]
                    return val if val is not None else default
                return default
            
            part_number = str(get_cell_value('part_number', '') or '').strip()
            item_name = str(get_cell_value('item_name', '') or '').strip()
            if not part_number or not item_name:
                errors.append(f"Row {idx}: Missing Part No or Part Name")
                continue
            
            quantity = safe_float(get_cell_value('quantity', 0), 0.0)
            location = str(get_cell_value('location', '') or '').strip()
            cost = safe_float(get_cell_value('cost', 0), 0.0)
            mav_value = safe_float(get_cell_value('mav_value', cost), cost)
            parts_type = str(get_cell_value('parts_type', '') or '').strip()
            fms_abc = str(get_cell_value('fms_abc', '') or '').strip()
            last_receipt_raw = get_cell_value('last_receipt_date', '')
            last_sales_raw = get_cell_value('last_sales_date', '')
            last_receipt_date = format_date_for_display(last_receipt_raw)
            last_sales_date = format_date_for_display(last_sales_raw)
            purchase_ageing_days = calculate_ageing_days(last_receipt_raw)
            sales_ageing_days = calculate_ageing_days(last_sales_raw)
            
            product_data = {
                "item_name": item_name,
                "part_number": part_number,
                "quantity": quantity,
                "price": mav_value,
                "cost": cost,
                "location": location,
                "parts_type": parts_type,
                "fms_abc": fms_abc,
                "mav_value": mav_value,
                "last_receipt_date": last_receipt_date,
                "last_sales_date": last_sales_date,
                "purchase_ageing_days": purchase_ageing_days,
                "sales_ageing_days": sales_ageing_days,
                "upload_no": upload_no,
                "upload_date": upload_date_str,
                "upload_time": upload_time_str,
                "upload_method": "excel",
                "updated_at": upload_dt.isoformat(),
                **user_ctx,
            }
            
            # Dealer-wise unique stock: same part can exist for different dealers/users.
            existing = await db.products.find_one({
                "part_number": part_number,
                "dealer_code": user_ctx["dealer_code"],
                "brand_code": user_ctx["brand_code"],
                "branch": user_ctx["branch"],
            })
            if existing:
                await db.products.update_one({"id": existing["id"]}, {"$set": product_data})
                status_text = "updated"
            else:
                product_data["id"] = str(uuid.uuid4())
                product_data["created_at"] = upload_dt.isoformat()
                product_data["category"] = parts_type
                await db.products.insert_one(product_data)
                status_text = "success"
            
            rows_imported += 1
            preview_data.append({**product_data, "status": status_text})
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    upload_record = {
        "id": str(uuid.uuid4()),
        "upload_no": upload_no,
        "upload_type": "product",
        "file_name": file.filename,
        "file_size": file_size,
        "uploaded_by": current_user.id,
        "uploaded_user_id": current_user.id,
        "uploaded_user_name": current_user.username,
        "user_code": user_ctx["user_code"],
        "dealer_code": user_ctx["dealer_code"],
        "dealer_name": user_ctx["dealer_name"],
        "brand_code": user_ctx["brand_code"],
        "brand_name": user_ctx["brand_name"],
        "branch": user_ctx["branch"],
        "brand": user_ctx["brand_name"],
        "location": user_ctx["branch"],
        "item_count": rows_imported,
        "rows_processed": rows_processed,
        "rows_imported": rows_imported,
        "status": "Imported" if rows_imported and not errors else ("Partial" if rows_imported else "Failed"),
        "publish_status": "Waiting" if rows_imported else "Failed",
        "failed_count": len(errors),
        "errors": errors,
        "upload_date": upload_date_str,
        "upload_time": upload_time_str,
        "created_at": upload_dt.isoformat(),
    }
    await db.uploads.insert_one(upload_record)
    
    return {
        "message": "Upload completed",
        "upload_no": upload_no,
        "upload_date": upload_date_str,
        "upload_time": upload_time_str,
        "rows_processed": rows_processed,
        "rows_imported": rows_imported,
        "errors": errors,
        "preview": preview_data[:100],
    }


@api_router.post("/upload/excel")
async def upload_excel(file: UploadFile = File(...), current_user: UserResponse = Depends(get_current_user)):
    return await process_product_upload(file, current_user)


@api_router.post("/upload")
async def upload_product_center(
    file: UploadFile = File(...),
    upload_type: str = Form("product"),
    current_user: UserResponse = Depends(get_current_user),
):
    if upload_type != "product":
        raise HTTPException(status_code=400, detail="Only product upload is supported on this endpoint")
    return await process_product_upload(file, current_user)

@api_router.get("/uploads")
async def get_uploads(
    type: str = None,
    brand: str = None,
    dealer: str = None,
    branch: str = None,
    current_user: UserResponse = Depends(get_current_user)
):
    if current_user.role not in ["master", "admin", "user"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    query = {}
    if type:
        query["upload_type"] = type

    def is_all(value: str) -> bool:
        return not value or str(value).startswith("All ") or value == "N/A"

    if current_user.role == "master":
        if not is_all(brand):
            query["brand_name"] = brand
        if not is_all(dealer):
            query["dealer_name"] = dealer
        if not is_all(branch):
            query["branch"] = branch
    elif current_user.role == "admin":
        query["brand_name"] = current_user.brand
        query["dealer_name"] = current_user.group
        if not is_all(branch):
            query["branch"] = branch
    else:
        query["brand_name"] = current_user.brand
        query["dealer_name"] = current_user.group
        query["branch"] = current_user.location

    # Sort needs an index / disk spill on large upload collections (avoids 32MB sort RAM 500).
    # Exclude raw_file_bytes from the list payload — binary Excel blobs break JSON encoding
    # and are not needed for Upload Center list/summary views.
    cursor = db.uploads.find(query, {"_id": 0, "raw_file_bytes": 0}).sort("created_at", -1)
    try:
        cursor = cursor.allow_disk_use(True)
    except Exception:
        pass
    uploads = await cursor.to_list(500)
    return uploads


@api_router.put("/uploads/{upload_id}/publish")
async def publish_upload(upload_id: str, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Only Master/Admin can publish uploads")

    upload = await db.uploads.find_one({"id": upload_id}, {"_id": 0})
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")

    if current_user.role == "admin":
        if upload.get("brand_name") != current_user.brand or upload.get("dealer_name") != current_user.group:
            raise HTTPException(status_code=403, detail="Not allowed to publish this upload")

    now = datetime.now(timezone.utc).isoformat()
    await db.uploads.update_one(
        {"id": upload_id},
        {"$set": {
            "publish_status": "Published",
            "published_at": now,
            "published_by": current_user.id,
            "published_user_name": current_user.username,
        }}
    )
    return {"message": "Upload published successfully"}


@api_router.delete("/uploads/reset")
async def reset_uploads(type: str = None, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role != "master":
        raise HTTPException(status_code=403, detail="Only Master Admin can reset data")

    query = {}
    if type:
        query["upload_type"] = type
    result = await db.uploads.delete_many(query)
    if type == "product":
        await db.products.delete_many({})
    if type == "order":
        await db.order_stocks.delete_many({})
    return {"message": f"Reset complete. Deleted {result.deleted_count} upload records."}

@api_router.get("/upload/sample-template")
async def download_sample_template(current_user: UserResponse = Depends(get_current_user)):
    """Generate and return a sample Excel template for product upload"""
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Add headers matching the sample format
    headers = ['Part No', 'Part Name', 'Available Qty', 'Bin Location', 'Cost', 'MRP', 'Part Types', 'FMS & ABC', 'Last Receipt Date', 'Last Sales Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Add sample data
    sample_data = [
        ['IKNL01150ESQQH', 'CABLE TIE', 7.00, '11A030109I', 1.67, 2.10, 'C', 'F', '06/11/2024', '10/01/2026'],
        ['MB30116525', 'BOLT-REAMER', 8.00, '11A010205A', 77.14, 95.00, 'X', 'S', '30/11/2025', '15/02/2026'],
        ['MF40117154', 'BUSHING', 2.00, '11A030204I', 153.67, 180.00, 'X', 'N', '15/09/2025', ''],
    ]
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    column_widths = [18, 30, 14, 15, 12, 12, 12, 12, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_upload_template.xlsx"}
    )

@api_router.get("/upload/export-products")
async def export_uploaded_products(current_user: UserResponse = Depends(get_current_user)):
    """Export all uploaded products to Excel - Master Admin / Admin only"""
    from fastapi.responses import StreamingResponse
    excel_permissions.require_excel_export(current_user)
    
    # Fetch all products
    products = await db.products.find({}, {"_id": 0}).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uploaded Products"
    
    # Add headers - matching the display table columns
    headers = ['Upload No', 'Upload Date', 'Upload Time', 'User Code', 'Dealer Code', 'Dealer Name', 'Brand Code', 'Brand Name', 'Branch', 'Part No', 'Part Name', 'Available Qty', 'Bin Location', 'Cost', 'MRP', 'Part Types', 'FMS & ABC', 'Last Receipt Date', 'Purchase Aging Days', 'Last Sales Date', 'Sales Aging Days', 'Uploaded User']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Add data
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.get('upload_no', ''))
        ws.cell(row=row_num, column=2, value=product.get('upload_date', ''))
        ws.cell(row=row_num, column=3, value=product.get('upload_time', ''))
        ws.cell(row=row_num, column=4, value=product.get('user_code', ''))
        ws.cell(row=row_num, column=5, value=product.get('dealer_code', ''))
        ws.cell(row=row_num, column=6, value=product.get('dealer_name', ''))
        ws.cell(row=row_num, column=7, value=product.get('brand_code', ''))
        ws.cell(row=row_num, column=8, value=product.get('brand_name', ''))
        ws.cell(row=row_num, column=9, value=product.get('branch', ''))
        ws.cell(row=row_num, column=10, value=product.get('part_number', ''))
        ws.cell(row=row_num, column=11, value=product.get('item_name', ''))
        ws.cell(row=row_num, column=12, value=product.get('quantity', 0))
        ws.cell(row=row_num, column=13, value=product.get('location', product.get('loc', '')))
        ws.cell(row=row_num, column=14, value=product.get('cost', 0))
        ws.cell(row=row_num, column=15, value=product.get('mav_value', product.get('price', 0)))
        ws.cell(row=row_num, column=16, value=product.get('parts_type', product.get('category', '')))
        ws.cell(row=row_num, column=17, value=product.get('fms_abc', ''))
        ws.cell(row=row_num, column=18, value=product.get('last_receipt_date', ''))
        ws.cell(row=row_num, column=19, value=product.get('purchase_ageing_days', ''))
        ws.cell(row=row_num, column=20, value=product.get('last_sales_date', ''))
        ws.cell(row=row_num, column=21, value=product.get('sales_ageing_days', ''))
        ws.cell(row=row_num, column=22, value=product.get('uploaded_user_name', product.get('upload_users', '')))
    
    # Adjust column widths
    column_widths = [14, 12, 10, 16, 16, 22, 14, 16, 16, 18, 30, 14, 15, 12, 12, 12, 12, 18, 18, 18, 16, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=uploaded_products_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

@api_router.delete("/upload/reset-products")
async def reset_all_products(current_user: UserResponse = Depends(get_current_user)):
    """Delete all uploaded products - Master Admin only"""
    if current_user.role != "master":
        raise HTTPException(status_code=403, detail="Only Master Admin can reset data")
    
    result = await db.products.delete_many({})
    await db.uploads.delete_many({})
    
    return {"message": f"Reset complete. Deleted {result.deleted_count} products."}

# ==================== ORDER ROUTES ====================

@api_router.get("/orders", response_model=List[Order])
async def get_orders(current_user: UserResponse = Depends(get_current_user)):
    if current_user.role == "user":
        # Users see only their own orders
        orders = await db.orders.find({"created_by": current_user.id}, {"_id": 0}).to_list(1000)
    elif current_user.role in ["admin", "master"]:
        # Admins see orders assigned to them or created by them
        orders = await db.orders.find({
            "$or": [
                {"assigned_to": current_user.id},
                {"created_by": current_user.id}
            ]
        }, {"_id": 0}).to_list(1000)
    else:
        orders = []
    
    return [Order(**o) for o in orders]

@api_router.post("/orders", response_model=Order)
async def create_order(order_data: OrderCreate, current_user: UserResponse = Depends(get_current_user)):
    # Check if product exists
    product = await db.products.find_one({"id": order_data.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check if assigned user is admin
    assigned_user = await db.users.find_one({"id": order_data.assigned_to}, {"_id": 0})
    if not assigned_user or assigned_user["role"] not in ["admin", "master"]:
        raise HTTPException(status_code=400, detail="Can only assign to admin users")
    
    total_price = product["price"] * order_data.quantity
    
    order = Order(
        product_id=order_data.product_id,
        quantity=order_data.quantity,
        total_price=total_price,
        assigned_to=order_data.assigned_to,
        created_by=current_user.id
    )
    
    doc = order.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.orders.insert_one(doc)
    
    # Create notification
    notification = Notification(
        order_id=order.id,
        recipient_id=order_data.assigned_to,
        message=f"New order request from {current_user.username}",
        status="pending"
    )
    notif_doc = notification.model_dump()
    notif_doc['created_at'] = notif_doc['created_at'].isoformat()
    await db.notifications.insert_one(notif_doc)
    
    # Send real-time notification
    await send_notification_to_user(order_data.assigned_to, {
        "id": notification.id,
        "message": notification.message,
        "order_id": order.id
    })
    
    return order

@api_router.put("/orders/{order_id}/action")
async def action_order(order_id: str, action: OrderAction, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order["assigned_to"] != current_user.id and current_user.role != "master":
        raise HTTPException(status_code=403, detail="Not authorized to action this order")
    
    if action.status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Invalid action")
    
    if action.status == "rejected" and not action.remarks:
        raise HTTPException(status_code=400, detail="Remarks required for rejection")
    
    # Update order
    update_data = {
        "status": action.status,
        "remarks": action.remarks,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # If approved, reduce stock
    if action.status == "approved":
        product = await db.products.find_one({"id": order["product_id"]}, {"_id": 0})
        if product["quantity"] < order["quantity"]:
            raise HTTPException(status_code=400, detail="Insufficient stock")
        
        await db.products.update_one(
            {"id": order["product_id"]},
            {"$inc": {"quantity": -order["quantity"]}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    # Update notification
    await db.notifications.update_one(
        {"order_id": order_id},
        {"$set": {"status": action.status}}
    )
    
    # Notify creator
    creator_notification = Notification(
        order_id=order_id,
        recipient_id=order["created_by"],
        message=f"Your order has been {action.status}",
        status=action.status
    )
    notif_doc = creator_notification.model_dump()
    notif_doc['created_at'] = notif_doc['created_at'].isoformat()
    await db.notifications.insert_one(notif_doc)
    
    await send_notification_to_user(order["created_by"], {
        "id": creator_notification.id,
        "message": creator_notification.message,
        "order_id": order_id
    })
    
    return {"message": f"Order {action.status} successfully"}

# ==================== NOTIFICATION ROUTES ====================

@api_router.get("/notifications")
async def get_notifications(filter_status: Optional[str] = None, current_user: UserResponse = Depends(get_current_user)):
    query = {"recipient_id": current_user.id}
    if filter_status:
        query["status"] = filter_status
    
    notifications = await db.notifications.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return notifications

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: UserResponse = Depends(get_current_user)):
    result = await db.notifications.update_one(
        {"id": notification_id, "recipient_id": current_user.id},
        {"$set": {"is_read": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: UserResponse = Depends(get_current_user)):
    await db.notifications.update_many(
        {"recipient_id": current_user.id},
        {"$set": {"is_read": True}}
    )
    return {"message": "All notifications marked as read"}

@api_router.get("/notifications/unread-count")
async def get_unread_count(current_user: UserResponse = Depends(get_current_user)):
    count = await db.notifications.count_documents({"recipient_id": current_user.id, "is_read": False})
    return {"count": count}

@api_router.delete("/notifications/clear-all")
async def clear_all_notifications(current_user: UserResponse = Depends(get_current_user)):
    result = await db.notifications.delete_many({"recipient_id": current_user.id})
    return {"message": f"Cleared {result.deleted_count} notifications"}

# ==================== ORDER UPLOAD ROUTES ====================

@api_router.get("/orders/merged-data")
async def get_merged_order_data(current_user: UserResponse = Depends(get_current_user)):
    """Get all stock data merged by part number with user columns"""
    
    # Get all unique users who have uploaded stock data
    order_stocks = await db.order_stocks.find({}, {"_id": 0}).to_list(100000)
    
    # Build user list
    user_ids = list(set(s.get('user_id') for s in order_stocks if s.get('user_id')))
    users = []
    for uid in user_ids:
        user = await db.users.find_one({"id": uid}, {"_id": 0, "id": 1, "username": 1})
        if user:
            users.append({"id": user["id"], "name": user["username"]})
    
    # Merge data by part number
    merged = {}
    for stock in order_stocks:
        part_no = stock.get('part_number', '')
        if not part_no:
            continue
        
        if part_no not in merged:
            merged[part_no] = {
                "part_number": part_no,
                "part_name": stock.get('part_name', ''),
                "user_data": {}
            }
        
        user_id = stock.get('user_id')
        if user_id:
            merged[part_no]["user_data"][user_id] = {
                "available_qty": stock.get('available_qty', 0),
                "receipt_ageing": stock.get('receipt_ageing', ''),
                "request_qty": stock.get('request_qty', 0)
            }
    
    products = list(merged.values())
    
    return {
        "products": products,
        "users": users
    }

@api_router.post("/orders/upload-stock")
async def upload_order_stock(file: UploadFile = File(...), current_user: UserResponse = Depends(get_current_user)):
    """Upload stock data for order management - each user's data is kept separate"""
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
    upload_dt = datetime.now(timezone.utc)
    upload_date_str = upload_dt.strftime('%d-%m-%Y')
    upload_time_str = upload_dt.strftime('%H:%M:%S')
    user_ctx = await get_user_upload_context(current_user)
    upload_no = await generate_upload_no("order", user_ctx["brand_code"])

    content = await file.read()
    file_size = len(content)
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    
    # Read headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append('')
    
    header_lower = [h.lower() for h in headers]
    
    # Column mapping
    col_mapping = {}
    column_definitions = {
        'part_number': ['part no', 'part_no', 'partno', 'part number'],
        'part_name': ['part name', 'part_name', 'partname', 'name'],
        'available_qty': ['available qty', 'available_qty', 'qty', 'quantity'],
        'receipt_ageing': ['receipt ageing', 'receipt_ageing', 'ageing', 'aging']
    }
    
    for field, possible_names in column_definitions.items():
        for i, header in enumerate(header_lower):
            if header in possible_names:
                col_mapping[field] = i
                break
    
    if 'part_number' not in col_mapping:
        raise HTTPException(status_code=400, detail="Required column 'Part No' not found")
    
    rows_imported = 0
    user_name = current_user.username or 'User'
    
    # Delete existing data for this user (to allow re-upload)
    await db.order_stocks.delete_many({"user_id": current_user.id})
    
    # Process rows
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        def get_cell_value(field, default=None):
            if field not in col_mapping:
                return default
            col_idx = col_mapping[field]
            if col_idx < len(row):
                val = row[col_idx]
                return val if val is not None else default
            return default
        
        part_number = str(get_cell_value('part_number', '') or '').strip()
        if not part_number:
            continue
        
        part_name = str(get_cell_value('part_name', '') or '').strip()
        
        qty_raw = get_cell_value('available_qty', 0)
        try:
            available_qty = float(qty_raw) if qty_raw else 0
        except:
            available_qty = 0
        
        receipt_ageing = str(get_cell_value('receipt_ageing', '') or '').strip()
        
        # Insert stock record
        stock_doc = {
            "id": str(uuid.uuid4()),
            "part_number": part_number,
            "part_name": part_name,
            "available_qty": available_qty,
            "receipt_ageing": receipt_ageing,
            "request_qty": 0,
            "user_id": current_user.id,
            "user_name": user_name,
            "upload_no": upload_no,
            "upload_date": upload_date_str,
            "upload_time": upload_time_str,
            **user_ctx,
            "created_at": upload_dt.isoformat()
        }
        await db.order_stocks.insert_one(stock_doc)
        rows_imported += 1
    
    upload_record = {
        "id": str(uuid.uuid4()),
        "upload_no": upload_no,
        "upload_type": "order",
        "file_name": file.filename,
        "file_size": file_size,
        "uploaded_by": current_user.id,
        "uploaded_user_id": current_user.id,
        "uploaded_user_name": current_user.username,
        "user_code": user_ctx["user_code"],
        "dealer_code": user_ctx["dealer_code"],
        "dealer_name": user_ctx["dealer_name"],
        "brand_code": user_ctx["brand_code"],
        "brand_name": user_ctx["brand_name"],
        "brand": user_ctx["brand_name"],
        "branch": user_ctx["branch"],
        "location": user_ctx["branch"],
        "item_count": rows_imported,
        "rows_processed": rows_imported,
        "rows_imported": rows_imported,
        "failed_count": 0,
        "status": "Imported" if rows_imported else "Failed",
        "publish_status": "Waiting" if rows_imported else "Failed",
        "errors": [],
        "upload_date": upload_date_str,
        "upload_time": upload_time_str,
        "created_at": upload_dt.isoformat(),
    }
    await db.uploads.insert_one(upload_record)

    return {
        "message": "Upload completed",
        "upload_no": upload_no,
        "upload_date": upload_date_str,
        "upload_time": upload_time_str,
        "rows_imported": rows_imported,
        "user_name": user_name
    }



@api_router.post("/orders/upload")
async def upload_order_center(file: UploadFile = File(...), current_user: UserResponse = Depends(get_current_user)):
    return await upload_order_stock(file, current_user)

@api_router.post("/orders/send-requests")
async def send_order_requests(request_data: dict, current_user: UserResponse = Depends(get_current_user)):
    """Send order requests for specific parts from a user"""
    
    user_id = request_data.get("user_id")
    requests = request_data.get("requests", [])
    
    if not requests:
        raise HTTPException(status_code=400, detail="No requests provided")
    
    updated = 0
    for req in requests:
        part_number = req.get("part_number")
        requested_qty = req.get("requested_qty", 0)
        
        if part_number and requested_qty > 0:
            # Update the request quantity in the stock record
            result = await db.order_stocks.update_one(
                {"part_number": part_number, "user_id": user_id},
                {"$set": {"request_qty": requested_qty, "request_date": datetime.now(timezone.utc).isoformat()}}
            )
            if result.modified_count > 0:
                updated += 1
    
    # Emit real-time update via Socket.IO
    await sio.emit('order_request_update', {
        "user_id": user_id,
        "updated_count": updated
    })
    
    return {"message": f"Updated {updated} requests"}

@api_router.delete("/orders/reset-all")
async def reset_all_order_data(current_user: UserResponse = Depends(get_current_user)):
    """Delete all order stock data - Master Admin only"""
    if current_user.role != "master":
        raise HTTPException(status_code=403, detail="Only Master Admin can reset data")
    
    result = await db.order_stocks.delete_many({})
    return {"message": f"Reset complete. Deleted {result.deleted_count} records."}

@api_router.get("/orders/export")
async def export_order_data(current_user: UserResponse = Depends(get_current_user)):
    """Export all order data to Excel"""
    from fastapi.responses import StreamingResponse
    excel_permissions.require_excel_export(current_user)
    
    # Get merged data
    order_stocks = await db.order_stocks.find({}, {"_id": 0}).to_list(100000)
    
    # Get unique users
    user_ids = list(set(s.get('user_id') for s in order_stocks if s.get('user_id')))
    users = []
    for uid in user_ids:
        user = await db.users.find_one({"id": uid}, {"_id": 0, "id": 1, "username": 1})
        if user:
            users.append({"id": user["id"], "name": user["username"]})
    
    # Merge data by part number
    merged = {}
    for stock in order_stocks:
        part_no = stock.get('part_number', '')
        if not part_no:
            continue
        
        if part_no not in merged:
            merged[part_no] = {
                "part_number": part_no,
                "part_name": stock.get('part_name', ''),
                "user_data": {}
            }
        
        user_id = stock.get('user_id')
        if user_id:
            merged[part_no]["user_data"][user_id] = {
                "available_qty": stock.get('available_qty', 0),
                "receipt_ageing": stock.get('receipt_ageing', ''),
                "request_qty": stock.get('request_qty', 0)
            }
    
    products = list(merged.values())
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Data"
    
    # Build headers
    headers = ['Part No', 'Order Qty', 'Part Name']
    for u in users:
        headers.extend([f"{u['name']} - Available Qty", f"{u['name']} - Receipt Ageing", f"{u['name']} - Request Qty"])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Add data
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product['part_number'])
        
        # Calculate order qty (sum of all available qty)
        order_qty = sum(ud.get('available_qty', 0) for ud in product.get('user_data', {}).values())
        ws.cell(row=row_num, column=2, value=order_qty)
        
        ws.cell(row=row_num, column=3, value=product['part_name'])
        
        col = 4
        for u in users:
            ud = product.get('user_data', {}).get(u['id'], {})
            ws.cell(row=row_num, column=col, value=ud.get('available_qty', 0))
            ws.cell(row=row_num, column=col+1, value=ud.get('receipt_ageing', ''))
            ws.cell(row=row_num, column=col+2, value=ud.get('request_qty', 0))
            col += 3
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=order_data_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

@api_router.get("/orders/sample-template")
async def download_order_template(current_user: UserResponse = Depends(get_current_user)):
    """Generate sample Excel template for order stock upload"""
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Data"
    
    headers = ['Part No', 'Part Name', 'Available Qty', 'Receipt Ageing']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    sample_data = [
        ['IKNL01150ESQQH', 'CABLE TIE', 7.00, "361 Day's"],
        ['MB30116525', 'BOLT-REAMER', 8.00, "180 Day's"],
        ['MF40117154', 'BUSHING', 2.00, "270 Day's"],
    ]
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    column_widths = [18, 25, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=order_upload_template.xlsx"}
    )

# ==================== DASHBOARD ROUTES ====================

@api_router.get("/dashboard/metrics")
async def get_dashboard_metrics(current_user: UserResponse = Depends(get_current_user)):
    total_products = await db.products.count_documents({})
    low_stock_count = await db.products.count_documents({"quantity": {"$lte": 10}})
    pending_orders = await db.orders.count_documents({"status": "pending"})
    
    # Non-moving stock (no orders in last 90 days)
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=90)
    all_products = await db.products.find({}, {"_id": 0, "id": 1}).to_list(1000)
    sleeping_count = 0
    for product in all_products:
        recent_order = await db.orders.find_one({
            "product_id": product["id"],
            "created_at": {"$gte": cutoff_date.isoformat()}
        })
        if not recent_order:
            sleeping_count += 1
    
    return {
        "total_products": total_products,
        "low_stock_count": low_stock_count,
        "pending_orders": pending_orders,
        "sleeping_stock_count": sleeping_count
    }

# ==================== PROFILE ROUTES ====================

@api_router.get("/profile/{user_id}", response_model=UserResponse)
async def get_user_profile(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    # Users can only view their own profile unless they're admin/master
    if current_user.role not in ["master", "admin"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized to view this profile")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return UserResponse(**user)

@api_router.put("/profile/{user_id}", response_model=UserResponse)
async def update_user_profile(
    user_id: str,
    update_data: UserUpdate,
    current_user: UserResponse = Depends(get_current_user)
):
    # Users can only edit their own profile unless they're admin/master
    if current_user.role not in ["master", "admin"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this profile")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Log activity
    activity_log = ActivityLog(user_id=current_user.id, action=f"updated profile for user {user_id}")
    log_doc = activity_log.model_dump()
    log_doc['created_at'] = log_doc['created_at'].isoformat()
    await db.activity_logs.insert_one(log_doc)
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    return UserResponse(**user)

@api_router.put("/profile/{user_id}/password")
async def change_password(
    user_id: str,
    password_data: PasswordChange,
    current_user: UserResponse = Depends(get_current_user)
):
    # Users can only change their own password
    if current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify old password
    if not verify_password(password_data.old_password, user["password"]):
        raise HTTPException(status_code=400, detail="Incorrect old password")
    
    # Update password
    new_hashed_password = hash_password(password_data.new_password)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password": new_hashed_password}}
    )
    
    # Log activity
    activity_log = ActivityLog(user_id=user_id, action="changed password")
    log_doc = activity_log.model_dump()
    log_doc['created_at'] = log_doc['created_at'].isoformat()
    await db.activity_logs.insert_one(log_doc)
    
    return {"message": "Password changed successfully"}


class ResetPasswordRequest(BaseModel):
    password: str


@api_router.put("/users/{user_id}/reset-password")
async def reset_user_password(
    user_id: str,
    data: ResetPasswordRequest,
    current_user: UserResponse = Depends(get_current_user)
):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    target_user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    if current_user.role == "admin":
        if target_user.get("role") != "user":
            raise HTTPException(status_code=403, detail="Admin can reset only user password")
        if target_user.get("brand") != current_user.brand or target_user.get("group") != current_user.group:
            raise HTTPException(status_code=403, detail="Not authorized for this user")

    if not data.password or len(data.password.strip()) < 4:
        raise HTTPException(status_code=400, detail="Password must be at least 4 characters")

    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "password": hash_password(data.password.strip()),
            "passwordResetAt": datetime.now(timezone.utc).isoformat(),
            "passwordResetBy": current_user.id
        }}
    )

    await db.activity_logs.insert_one({
        "id": str(uuid.uuid4()),
        "action": "PASSWORD_RESET",
        "user_id": current_user.id,
        "target_user_id": user_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    return {"message": "Password reset successfully"}

@api_router.put("/profile/{user_id}/status")
async def toggle_user_status(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    # Only admin and master can change status
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Cannot deactivate yourself
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot change your own status")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_status = "inactive" if user.get("status") == "active" else "active"
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"status": new_status}}
    )
    
    # Log activity
    activity_log = ActivityLog(
        user_id=current_user.id,
        action=f"changed user {user_id} status to {new_status}"
    )
    log_doc = activity_log.model_dump()
    log_doc['created_at'] = log_doc['created_at'].isoformat()
    await db.activity_logs.insert_one(log_doc)
    
    return {"message": f"User status changed to {new_status}", "status": new_status}

@api_router.get("/profile/{user_id}/analytics")
async def get_user_analytics(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    # Users can only view their own analytics unless they're admin/master
    if current_user.role not in ["master", "admin"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Count total orders
    total_orders = await db.orders.count_documents({"created_by": user_id})
    
    # Calculate total purchase amount
    orders = await db.orders.find({"created_by": user_id}, {"_id": 0}).to_list(1000)
    total_amount = sum(order.get("total_price", 0) for order in orders)
    
    # Count by status
    pending_orders = await db.orders.count_documents({"created_by": user_id, "status": "pending"})
    approved_orders = await db.orders.count_documents({"created_by": user_id, "status": "approved"})
    rejected_orders = await db.orders.count_documents({"created_by": user_id, "status": "rejected"})
    
    return {
        "total_orders": total_orders,
        "total_amount": total_amount,
        "pending_orders": pending_orders,
        "approved_orders": approved_orders,
        "rejected_orders": rejected_orders
    }

@api_router.get("/profile/{user_id}/orders")
async def get_user_orders(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    # Users can only view their own orders unless they're admin/master
    if current_user.role not in ["master", "admin"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    orders = await db.orders.find({"created_by": user_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return orders

@api_router.get("/profile/{user_id}/activity-logs")
async def get_user_activity_logs(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    # Users can view their own logs, admins can view anyone's
    if current_user.role not in ["master", "admin"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    logs = await db.activity_logs.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    return logs

# ==================== BRANDS & GROUPS ====================

@api_router.get("/brands-groups")
async def get_brands_and_groups(current_user: UserResponse = Depends(get_current_user)):
    """Get all brands and groups"""
    brands = await db.brands.find({}, {"_id": 0}).to_list(100)
    groups = await db.groups.find({}, {"_id": 0}).to_list(100)
    return {
        "brands": [b["name"] for b in brands],
        "groups": [g["name"] for g in groups]
    }

@api_router.post("/brands")
async def create_brand(brand: BrandCreate, current_user: UserResponse = Depends(get_current_user)):
    """Create a new brand"""
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if brand already exists
    existing = await db.brands.find_one({"name": brand.name})
    if existing:
        raise HTTPException(status_code=400, detail="Brand already exists")
    
    new_brand = Brand(name=brand.name)
    doc = new_brand.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.brands.insert_one(doc)
    return {"message": "Brand created successfully", "name": brand.name}

@api_router.delete("/brands/{brand_name}")
async def delete_brand(brand_name: str, current_user: UserResponse = Depends(get_current_user)):
    """Delete a brand"""
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.brands.delete_one({"name": brand_name})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    return {"message": "Brand deleted successfully"}

@api_router.post("/groups")
async def create_group(group: GroupCreate, current_user: UserResponse = Depends(get_current_user)):
    """Create a new group"""
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if group already exists
    existing = await db.groups.find_one({"name": group.name})
    if existing:
        raise HTTPException(status_code=400, detail="Group already exists")
    
    new_group = Group(name=group.name)
    doc = new_group.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.groups.insert_one(doc)
    return {"message": "Group created successfully", "name": group.name}

@api_router.delete("/groups/{group_name}")
async def delete_group(group_name: str, current_user: UserResponse = Depends(get_current_user)):
    """Delete a group"""
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.groups.delete_one({"name": group_name})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Group not found")
    return {"message": "Group deleted successfully"}

# ==================== INIT ROUTE ====================

@api_router.post("/init")
async def initialize_system():
    result = await ensure_master_user_exists()
    return result

async def ensure_master_user_exists():
    """Create the default master admin if no master user exists yet.
    Called automatically on backend startup, and also exposed via POST /api/init
    for manual/idempotent use. Both paths share this single implementation so
    there is exactly one place that defines the default admin credentials.
    """
    master = await db.users.find_one({"role": "master"})
    if not master:
        master_user = User(
            username="Master Admin",
            email="admin@sleepingstock.in",
            password=hash_password("admin123"),
            role="master"
        )
        doc = master_user.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        return {"message": "Master user created", "email": "admin@sleepingstock.in", "password": "admin123"}
    return {"message": "System already initialized"}

# ==================== UPLOAD TRACKING ====================

class UploadRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    upload_no: str
    upload_type: str  # 'product' or 'order'
    uploaded_by: str = ""
    brand: str = ""
    group: str = ""
    location: str = ""
    item_count: int = 0
    total_value: float = 0.0
    status: str = "Pending"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Duplicate /uploads/reset route removed. The canonical Master-only route is defined earlier.

# ================= USER HUB / MASTER SETTINGS APIs =================

class MasterStateCreate(BaseModel):
    code: str
    name: str


class MasterBrandCreate(BaseModel):
    code: str
    name: str


class MasterDealerCreate(BaseModel):
    name: str
    brand: Optional[str] = None


class MasterBranchCreate(BaseModel):
    dealer: str
    name: str
    brand: Optional[str] = None


class UserHubCreate(BaseModel):
    userId: str
    name: str
    mobile: str
    email: str
    role: str
    state: str = ""
    brand: str
    dealer: str
    branch: str
    password: str
    confirmPassword: Optional[str] = None
    status: str = "active"
    permissions: List[str] = []


class UserHubUpdate(BaseModel):
    name: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    state: Optional[str] = None
    brand: Optional[str] = None
    dealer: Optional[str] = None
    branch: Optional[str] = None
    status: Optional[str] = None
    permissions: Optional[List[str]] = None


class TemplateMetaCreate(BaseModel):
    brand: str
    templateType: str


def _now_iso():
    return datetime.now(timezone.utc).isoformat()


def _permission_list(value):
    return normalize_permissions(value)


async def _ensure_master(current_user: UserResponse):
    if current_user.role != "master":
        raise HTTPException(status_code=403, detail="Only Master Admin can perform this action")


async def _ensure_master_or_admin(current_user: UserResponse):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")


def _clean_master_code(value: str) -> str:
    """Clean a master code without guessing or slicing it."""
    return (value or "").strip().upper()


def _clean_master_name(value: str) -> str:
    return (value or "").strip()


def _exact_ci(value: str) -> dict:
    return {"$regex": f"^{re.escape((value or '').strip())}$", "$options": "i"}


DEFAULT_MASTER_RECORDS = {
    "states": [
        {"code": "TN", "name": "Tamil Nadu", "status": "active"},
        {"code": "KL", "name": "Kerala", "status": "active"},
    ],
    "brands": [
        {"code": "HY", "name": "Hyundai", "status": "active"},
    ],
}


def _master_record_matches(record: dict, raw_value: str) -> bool:
    raw = (raw_value or "").strip()
    if not raw:
        return False
    code = str(record.get("code") or "").strip()
    name = str(record.get("name") or "").strip()
    return raw.upper() == code.upper() or raw.lower() == name.lower()


async def _resolve_master_record(collection_name: str, value: str, label: str) -> dict:
    """Resolve selected master value to the exact saved master record.

    This never guesses from the first two letters. Kerala must resolve to the
    State Master code KL, not KE, KERALA, or default TN.
    """
    raw_value = (value or "").strip()
    if not raw_value:
        raise HTTPException(status_code=400, detail=f"{label} is required")

    code_value = raw_value.upper()
    collection = getattr(db, collection_name)
    record = await collection.find_one({
        "$or": [
            {"code": code_value},
            {"code": {"$regex": f"^{re.escape(code_value)}$", "$options": "i"}},
            {"name": raw_value},
            {"name": {"$regex": f"^{re.escape(raw_value)}$", "$options": "i"}},
        ]
    }, {"_id": 0})

    # First-time setup may show default masters in the UI before they are saved
    # in MongoDB. Resolve those exact defaults too, without guessing.
    if not record:
        for default_record in DEFAULT_MASTER_RECORDS.get(collection_name, []):
            if _master_record_matches(default_record, raw_value):
                record = default_record
                break

    if not record or not (record.get("code") or "").strip():
        raise HTTPException(status_code=400, detail=f"Invalid {label} selected. Please check {label} master code.")

    return {
        **record,
        "code": _clean_master_code(record.get("code")),
        "name": _clean_master_name(record.get("name")) or raw_value,
    }


async def _resolve_master_code(collection_name: str, value: str, label: str) -> str:
    record = await _resolve_master_record(collection_name, value, label)
    return record["code"]


@api_router.get("/masters/states")
async def get_master_states(
    current_user: UserResponse = Depends(get_current_user)
):
    states = await db.states.find(
        {},
        {"_id": 0}
    ).sort("name", 1).to_list(1000)

    if states:
        return states

    for default_state in DEFAULT_MASTER_RECORDS["states"]:
        state_code = _clean_master_code(default_state.get("code"))

        await db.states.update_one(
            {"code": state_code},
            {
                "$setOnInsert": {
                    **default_state,
                    "code": state_code
                }
            },
            upsert=True
        )

    states = await db.states.find(
        {},
        {"_id": 0}
    ).sort("name", 1).to_list(1000)

    return states

@api_router.post("/masters/states")
async def add_master_state(data: MasterStateCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    code = _clean_master_code(data.code)
    name = data.name.strip()

    if not code or not name:
        raise HTTPException(status_code=400, detail="State code and name required")

    if await db.states.find_one({"code": code}):
        raise HTTPException(status_code=400, detail="State code already exists")

    if await db.states.find_one({"name": name}):
        raise HTTPException(status_code=400, detail="State name already exists")

    await db.states.insert_one({
        "id": str(uuid.uuid4()),
        "code": code,
        "name": name,
        "status": "active",
        "createdAt": _now_iso(),
        "createdBy": current_user.id
    })

    return {"message": "State added successfully"}


@api_router.put("/masters/states/{code}")
async def update_master_state(code: str, data: MasterStateCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    old_code = _clean_master_code(code)
    new_code = _clean_master_code(data.code)
    name = data.name.strip()

    if not new_code or not name:
        raise HTTPException(status_code=400, detail="State code and name required")

    existing = await db.states.find_one({"code": new_code})
    if existing and existing.get("code") != old_code:
        raise HTTPException(status_code=400, detail="State code already exists")

    result = await db.states.update_one(
        {"code": old_code},
        {"$set": {
            "code": new_code,
            "name": name,
            "status": "active",
            "updatedAt": _now_iso(),
            "updatedBy": current_user.id
        }}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="State not found")

    await db.users.update_many({"state": old_code}, {"$set": {"state": name}})
    return {"message": "State updated successfully"}


@api_router.delete("/masters/states/{code}")
async def delete_master_state(code: str, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    state = await db.states.find_one({"code": _clean_master_code(code)})
    if not state:
        raise HTTPException(status_code=404, detail="State not found")

    used = await db.users.find_one({"state": {"$in": [state.get("name"), state.get("code")]}})
    if used:
        raise HTTPException(status_code=400, detail="State is already used by users")

    await db.states.delete_one({"code": state.get("code")})
    return {"message": "State deleted successfully"}


@api_router.get("/masters/brands")
async def get_master_brands(current_user: UserResponse = Depends(get_current_user)):
    brands = await db.brands.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return brands or DEFAULT_MASTER_RECORDS["brands"]


@api_router.post("/masters/brands")
async def add_master_brand(data: MasterBrandCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    code = _clean_master_code(data.code)
    name = data.name.strip()

    if not code or not name:
        raise HTTPException(status_code=400, detail="Brand code and name required")

    if await db.brands.find_one({"code": code}):
        raise HTTPException(status_code=400, detail="Brand code already exists")

    if await db.brands.find_one({"name": name}):
        raise HTTPException(status_code=400, detail="Brand name already exists")

    await db.brands.insert_one({
        "id": str(uuid.uuid4()),
        "code": code,
        "name": name,
        "status": "active",
        "createdAt": _now_iso(),
        "createdBy": current_user.id
    })

    return {"message": "Brand added successfully"}


@api_router.put("/masters/brands/{code}")
async def update_master_brand(code: str, data: MasterBrandCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    old_code = _clean_master_code(code)
    new_code = _clean_master_code(data.code)
    name = data.name.strip()

    if not new_code or not name:
        raise HTTPException(status_code=400, detail="Brand code and name required")

    existing = await db.brands.find_one({"code": new_code})
    if existing and existing.get("code") != old_code:
        raise HTTPException(status_code=400, detail="Brand code already exists")

    result = await db.brands.update_one(
        {"code": old_code},
        {"$set": {
            "code": new_code,
            "name": name,
            "status": "active",
            "updatedAt": _now_iso(),
            "updatedBy": current_user.id
        }}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")

    await db.users.update_many({"brand": old_code}, {"$set": {"brand": name}})
    return {"message": "Brand updated successfully"}


@api_router.delete("/masters/brands/{code}")
async def delete_master_brand(code: str, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    brand = await db.brands.find_one({"code": _clean_master_code(code)})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    used = await db.users.find_one({"brand": brand.get("name")})
    if used:
        raise HTTPException(status_code=400, detail="Brand is already used by users")

    await db.brands.delete_one({"code": brand.get("code")})
    return {"message": "Brand deleted successfully"}


@api_router.get("/masters/dealers")
async def get_master_dealers(current_user: UserResponse = Depends(get_current_user)):
    query = {}
    if current_user.role in ["admin", "user"]:
        query = {"$or": [{"brand": {"$exists": False}}, {"brand": current_user.brand}, {"brand_name": current_user.brand}]}
    return await db.dealers.find(query, {"_id": 0}).sort("name", 1).to_list(1000)


@api_router.post("/masters/dealers")
async def add_master_dealer(data: MasterDealerCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    name = data.name.strip()
    brand = (data.brand or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Dealer name required")

    if await db.dealers.find_one({"name": _exact_ci(name), **({"brand": brand} if brand else {})}):
        raise HTTPException(status_code=400, detail="Dealer already exists")

    await db.dealers.insert_one({
        "id": str(uuid.uuid4()),
        "name": name,
        "brand": brand,
        "brand_name": brand,
        "status": "active",
        "createdAt": _now_iso(),
        "createdBy": current_user.id
    })

    return {"message": "Dealer added successfully"}


@api_router.put("/masters/dealers/{name}")
async def update_master_dealer(name: str, data: MasterDealerCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    old_name = name.strip()
    new_name = data.name.strip()
    brand = (data.brand or "").strip()

    if not new_name:
        raise HTTPException(status_code=400, detail="Dealer name required")

    existing = await db.dealers.find_one({"name": _exact_ci(new_name), **({"brand": brand} if brand else {})})
    if existing and existing.get("name") != old_name:
        raise HTTPException(status_code=400, detail="Dealer already exists")

    result = await db.dealers.update_one(
        {"name": old_name},
        {"$set": {
            "name": new_name,
            "brand": brand,
            "brand_name": brand,
            "status": "active",
            "updatedAt": _now_iso(),
            "updatedBy": current_user.id
        }}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dealer not found")

    await db.branches.update_many({"dealer": old_name}, {"$set": {"dealer": new_name}})
    await db.users.update_many({"group": old_name}, {"$set": {"group": new_name, "dealer": new_name}})
    return {"message": "Dealer updated successfully"}


@api_router.delete("/masters/dealers/{name}")
async def delete_master_dealer(name: str, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    dealer_name = name.strip()
    used_user = await db.users.find_one({"$or": [{"group": dealer_name}, {"dealer": dealer_name}]})
    used_branch = await db.branches.find_one({"dealer": dealer_name})

    if used_user or used_branch:
        raise HTTPException(status_code=400, detail="Dealer is already used")

    result = await db.dealers.delete_one({"name": dealer_name})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dealer not found")

    return {"message": "Dealer deleted successfully"}


@api_router.get("/masters/branches")
async def get_master_branches(current_user: UserResponse = Depends(get_current_user)):
    query = {}
    if current_user.role == "admin":
        query = {"dealer": current_user.group}
    elif current_user.role == "user":
        query = {"dealer": current_user.group, "name": current_user.location}

    return await db.branches.find(query, {"_id": 0}).sort("name", 1).to_list(1000)


@api_router.post("/masters/branches")
async def add_master_branch(data: MasterBranchCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    dealer = data.dealer.strip()
    name = data.name.strip()
    brand = (data.brand or "").strip()

    if not dealer or not name:
        raise HTTPException(status_code=400, detail="Dealer and branch name required")

    dealer_doc = await db.dealers.find_one({"name": _exact_ci(dealer)}, {"_id": 0})
    if not dealer_doc:
        raise HTTPException(status_code=400, detail="Invalid dealer selected")
    if not brand:
        brand = dealer_doc.get("brand") or dealer_doc.get("brand_name") or ""

    if await db.branches.find_one({"dealer": dealer, "name": _exact_ci(name), **({"brand": brand} if brand else {})}):
        raise HTTPException(status_code=400, detail="Branch already exists")

    await db.branches.insert_one({
        "id": str(uuid.uuid4()),
        "brand": brand,
        "brand_name": brand,
        "dealer": dealer,
        "dealer_name": dealer,
        "name": name,
        "status": "active",
        "createdAt": _now_iso(),
        "createdBy": current_user.id
    })

    return {"message": "Branch added successfully"}


@api_router.put("/masters/branches/{name}")
async def update_master_branch(name: str, data: MasterBranchCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    old_name = name.strip()
    dealer = data.dealer.strip()
    new_name = data.name.strip()
    brand = (data.brand or "").strip()

    if not dealer or not new_name:
        raise HTTPException(status_code=400, detail="Dealer and branch name required")

    dealer_doc = await db.dealers.find_one({"name": _exact_ci(dealer)}, {"_id": 0})
    if not dealer_doc:
        raise HTTPException(status_code=400, detail="Invalid dealer selected")
    if not brand:
        brand = dealer_doc.get("brand") or dealer_doc.get("brand_name") or ""

    existing = await db.branches.find_one({"dealer": dealer, "name": _exact_ci(new_name), **({"brand": brand} if brand else {})})
    if existing and existing.get("name") != old_name:
        raise HTTPException(status_code=400, detail="Branch already exists")

    result = await db.branches.update_one(
        {"name": old_name},
        {"$set": {
            "brand": brand,
            "brand_name": brand,
            "dealer": dealer,
            "dealer_name": dealer,
            "name": new_name,
            "status": "active",
            "updatedAt": _now_iso(),
            "updatedBy": current_user.id
        }}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")

    await db.users.update_many({"location": old_name}, {"$set": {"location": new_name, "branch": new_name}})
    return {"message": "Branch updated successfully"}


@api_router.delete("/masters/branches/{name}")
async def delete_master_branch(name: str, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    branch_name = name.strip()
    used = await db.users.find_one({"$or": [{"location": branch_name}, {"branch": branch_name}]})
    if used:
        raise HTTPException(status_code=400, detail="Branch is already used by users")

    result = await db.branches.delete_one({"name": branch_name})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")

    return {"message": "Branch deleted successfully"}


async def _build_next_user_id(state_code: str, brand_code: str) -> str:
    today = datetime.now(timezone.utc)
    date_code = today.strftime("%y%m%d")
    clean_state_code = _clean_master_code(state_code)
    clean_brand_code = _clean_master_code(brand_code)

    if not clean_state_code:
        raise HTTPException(status_code=400, detail="State code is required for User ID generation")
    if not clean_brand_code:
        raise HTTPException(status_code=400, detail="Brand code is required for User ID generation")

    # Format: SS + StateCode + BrandCode + YYMMDD + running number
    # Example: SSKLHY26070501
    prefix = f"SS{clean_state_code}{clean_brand_code}{date_code}"

    count = await db.users.count_documents({
        "$or": [
            {"userId": {"$regex": f"^{re.escape(prefix)}"}},
            {"user_id": {"$regex": f"^{re.escape(prefix)}"}}
        ]
    })

    return f"{prefix}{str(count + 1).zfill(2)}"


@api_router.get("/users/generate-id")
async def generate_user_id(
    brand_code: str,
    state_code: str = "TN",
    current_user: UserResponse = Depends(get_current_user)
):
    await _ensure_master_or_admin(current_user)
    resolved_state_code = await _resolve_master_code("states", state_code, "State")
    resolved_brand_code = await _resolve_master_code("brands", brand_code, "Brand")
    return {"user_id": await _build_next_user_id(resolved_state_code, resolved_brand_code)}


@api_router.get("/users/list")
async def list_hub_users(current_user: UserResponse = Depends(get_current_user)):
    query = {}

    if current_user.role == "admin":
        query = {"brand": current_user.brand, "group": current_user.group}
    elif current_user.role == "user":
        query = {
            "brand": current_user.brand,
            "group": current_user.group,
            "location": current_user.location
        }

    return await db.users.find(query, {"_id": 0, "password": 0}).sort("created_at", -1).to_list(1000)


@api_router.post("/users/create")
async def create_hub_user(data: UserHubCreate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master_or_admin(current_user)

    if current_user.role == "admin" and data.role != "user":
        raise HTTPException(status_code=403, detail="Admin can create only users")

    if data.password != data.confirmPassword:
        raise HTTPException(status_code=400, detail="Password not matching")

    if current_user.role == "admin":
        allowed_permissions = set(_permission_list(current_user.permissions))
        requested_permissions = set(data.permissions or [])

        if not requested_permissions.issubset(allowed_permissions):
            raise HTTPException(status_code=403, detail="Admin cannot assign permissions not given by Master")

        data.brand = current_user.brand
        data.dealer = current_user.group
        if not data.branch:
            data.branch = current_user.location

    if data.role not in ["master", "admin", "user"]:
        raise HTTPException(status_code=400, detail="Invalid role")

    data.userId = (data.userId or "").strip()
    data.email = normalize_email(data.email)

    state_record = await _resolve_master_record("states", data.state, "State")
    brand_record = await _resolve_master_record("brands", data.brand, "Brand")

    if not data.userId:
        data.userId = await _build_next_user_id(state_record["code"], brand_record["code"])

    # Store display names in user records, but generate IDs only from master codes.
    data.state = state_record["name"]
    data.brand = brand_record["name"]

    if await db.users.find_one({"$or": [{"userId": data.userId}, {"user_id": data.userId}]}):
        raise HTTPException(status_code=400, detail="User ID already exists")

    escaped_email = re.escape(data.email)
    if await db.users.find_one({"email": {"$regex": f"^{escaped_email}$", "$options": "i"}}):
        raise HTTPException(status_code=400, detail="Email already exists")

    if await db.users.find_one({"$or": [{"mobile": data.mobile}, {"phone": data.mobile}]}):
        raise HTTPException(status_code=400, detail="Mobile already exists")

    hashed_password = pwd_context.hash(data.password)
    now = _now_iso()

    user_doc = {
        "id": str(uuid.uuid4()),
        "userId": data.userId,
        "user_id": data.userId,
        "username": data.name,
        "name": data.name,
        "mobile": data.mobile,
        "phone": data.mobile,
        "state": data.state,
        "email": data.email,
        "role": data.role,
        "brand": data.brand,
        "dealer": data.dealer,
        "group": data.dealer,
        "branch": data.branch,
        "location": data.branch,
        "password": hashed_password,
        "status": data.status,
        "permissions": normalize_permissions(data.permissions),
        "createdAt": now,
        "created_at": now,
        "createdBy": current_user.id,
        "lastLogin": None,
        "last_login": None
    }

    await db.users.insert_one(user_doc)

    await db.activity_logs.insert_one({
        "id": str(uuid.uuid4()),
        "action": "USER_CREATED",
        "user_id": current_user.id,
        "created_user_id": data.userId,
        "created_user_email": data.email,
        "created_at": now
    })

    return {"message": "User created successfully"}


@api_router.put("/users/hub/{user_id}")
async def update_hub_user(user_id: str, data: UserHubUpdate, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master_or_admin(current_user)

    target = await db.users.find_one({"$or": [{"id": user_id}, {"userId": user_id}, {"user_id": user_id}]})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    if current_user.role == "admin":
        if target.get("role") != "user":
            raise HTTPException(status_code=403, detail="Admin can update only users")
        if target.get("brand") != current_user.brand or target.get("group") != current_user.group:
            raise HTTPException(status_code=403, detail="Not authorized for this user")

    update_data = {}
    raw = data.model_dump()
    for key, value in raw.items():
        if value is not None:
            update_data[key] = value

    if "dealer" in update_data:
        update_data["group"] = update_data["dealer"]
    if "branch" in update_data:
        update_data["location"] = update_data["branch"]
    if "mobile" in update_data:
        update_data["phone"] = update_data["mobile"]
    if "name" in update_data:
        update_data["username"] = update_data["name"]

    update_data["updatedAt"] = _now_iso()
    update_data["updatedBy"] = current_user.id

    await db.users.update_one({"id": target["id"]}, {"$set": update_data})
    return {"message": "User updated successfully"}


@api_router.delete("/users/hub/{user_id}")
async def delete_hub_user(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    result = await db.users.delete_one({"$or": [{"id": user_id}, {"userId": user_id}, {"user_id": user_id}]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"message": "User deleted successfully"}


@api_router.get("/scope/options")
async def get_scope_options(current_user: UserResponse = Depends(get_current_user)):
    if current_user.role == "master":
        states = await db.states.find({"status": "active"}, {"_id": 0, "code": 1, "name": 1}).sort("name", 1).to_list(1000)
        if not states:
            states = [{"code": "TN", "name": "Tamil Nadu"}, {"code": "KL", "name": "Kerala"}]
        brands = await db.brands.find({"status": "active"}, {"_id": 0, "code": 1, "name": 1}).sort("name", 1).to_list(1000)
        dealers = await db.dealers.find({"status": "active"}, {"_id": 0, "name": 1, "brand": 1, "brand_name": 1}).sort("name", 1).to_list(1000)
        branches = await db.branches.find({"status": "active"}, {"_id": 0, "code": 1, "name": 1, "dealer": 1, "dealer_name": 1, "brand": 1, "brand_name": 1}).sort("name", 1).to_list(1000)
        return {"states": states, "brands": brands, "dealers": dealers, "branches": branches}

    if current_user.role == "admin":
        # Branch master records may contain either dealer/dealer_name and
        # brand/brand_name depending on when they were created. Match both
        # field variants case-insensitively so every branch belonging to the
        # Admin's assigned dealer is returned.
        dealer_value = (current_user.group or current_user.dealer or "").strip()
        brand_value = (current_user.brand or "").strip()
        dealer_rx = {"$regex": f"^{re.escape(dealer_value)}$", "$options": "i"}
        brand_rx = {"$regex": f"^{re.escape(brand_value)}$", "$options": "i"}

        branch_query = {
            "status": "active",
            "$and": [
                {"$or": [{"dealer": dealer_rx}, {"dealer_name": dealer_rx}]},
                {"$or": [
                    {"brand": {"$exists": False}},
                    {"brand": None},
                    {"brand": ""},
                    {"brand": brand_rx},
                    {"brand_name": brand_rx},
                ]},
            ],
        }
        branches = await db.branches.find(
            branch_query,
            {"_id": 0, "code": 1, "name": 1, "dealer": 1, "dealer_name": 1, "brand": 1, "brand_name": 1}
        ).sort("name", 1).to_list(1000)
        state_code = await _resolve_master_code("states", current_user.state or "Tamil Nadu", "State")
        brand_code = await _resolve_master_code("brands", current_user.brand or "Hyundai", "Brand")
        return {
            "states": [{"name": current_user.state or "Tamil Nadu", "code": state_code}],
            "brands": [{"name": current_user.brand, "code": brand_code}],
            "dealers": [{"name": current_user.group, "brand": current_user.brand, "brand_name": current_user.brand}],
            "branches": branches or [{"name": current_user.location, "dealer": current_user.group, "dealer_name": current_user.group, "brand": current_user.brand, "brand_name": current_user.brand}]
        }

    state_code = await _resolve_master_code("states", current_user.state or "Tamil Nadu", "State")
    brand_code = await _resolve_master_code("brands", current_user.brand or "Hyundai", "Brand")
    return {
        "states": [{"name": current_user.state or "Tamil Nadu", "code": state_code}],
        "brands": [{"name": current_user.brand, "code": brand_code}],
        "dealers": [{"name": current_user.group, "brand": current_user.brand, "brand_name": current_user.brand}],
        "branches": [{"name": current_user.location, "dealer": current_user.group, "dealer_name": current_user.group, "brand": current_user.brand, "brand_name": current_user.brand}]
    }


@api_router.get("/dashboard/summary")
async def get_dashboard_summary(
    brand: Optional[str] = None,
    dealer: Optional[str] = None,
    branch: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user)
):
    product_query = {}
    user_query = {}

    if current_user.role == "master":
        if brand and brand != "All Brands":
            product_query["brand_name"] = brand
            user_query["brand"] = brand
        if dealer and dealer != "All Dealers":
            product_query["dealer_name"] = dealer
            user_query["group"] = dealer
        if branch and branch != "All Branches":
            product_query["branch"] = branch
            user_query["location"] = branch
    elif current_user.role == "admin":
        product_query["brand_name"] = current_user.brand
        product_query["dealer_name"] = current_user.group
        user_query["brand"] = current_user.brand
        user_query["group"] = current_user.group
        if branch and branch != "All Branches":
            product_query["branch"] = branch
            user_query["location"] = branch
    else:
        product_query["brand_name"] = current_user.brand
        product_query["dealer_name"] = current_user.group
        product_query["branch"] = current_user.location
        user_query["brand"] = current_user.brand
        user_query["group"] = current_user.group
        user_query["location"] = current_user.location

    registered_dealers = await db.dealers.count_documents({})
    connected_branches = await db.branches.count_documents({})
    registered_users = await db.users.count_documents(user_query)
    total_parts = await db.products.count_documents(product_query)
    pending_requests = await db.orders.count_documents({"status": "pending"})

    pipeline = [
        {"$match": product_query},
        {"$group": {"_id": None, "value": {"$sum": {"$multiply": ["$quantity", "$mav_value"]}}}}
    ]
    value_result = await db.products.aggregate(pipeline).to_list(1)
    non_moving_value = value_result[0]["value"] if value_result else 0

    return {
        "registeredDealers": registered_dealers,
        "connectedBranches": connected_branches,
        "registeredUsers": registered_users,
        "totalParts": total_parts,
        "nonMovingValue": non_moving_value,
        "movedValue": 0,
        "pendingRequests": pending_requests
    }


@api_router.get("/uploads/master-summary")
async def get_master_upload_summary(
    brand: Optional[str] = None,
    dealer: Optional[str] = None,
    branch: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Master Admin / Admin summary cards for Upload Center / Dashboard.
    Refreshes automatically whenever the global header filters (Brand/Dealer/Branch) change,
    since the frontend passes the active scope values as query params. Admin is
    automatically scoped to their own Brand/Dealer by _apply_role_scope_v2 below —
    they can never see another Admin's data.
    """
    await _ensure_master_or_admin(current_user)

    date_key = _nmts_date_key()
    upload_query = {"date_key": date_key}
    _apply_role_scope_v2(upload_query, current_user, brand, dealer, branch)

    uploads_today = await db.uploads.find(upload_query, {"_id": 0, "raw_file_bytes": 0}).to_list(200000)

    active_uploads = [u for u in uploads_today if u.get("status") != "Cancelled"]
    brands_uploaded = len({u.get("brand_name") for u in active_uploads if u.get("brand_name")})
    dealers_uploaded = len({u.get("dealer_name") for u in active_uploads if u.get("dealer_name")})
    branches_uploaded = len({u.get("branch") for u in active_uploads if u.get("branch")})

    published = sum(1 for u in uploads_today if u.get("publish_status") == "Published")
    cancelled = sum(1 for u in uploads_today if u.get("status") == "Cancelled")
    pending = sum(1 for u in uploads_today if u.get("status") != "Cancelled" and u.get("publish_status") != "Published")

    # Quantity totals (not just counts) for the Published Quantity / Pending Quantity
    # / Cancelled Quantity summary cards. Additive fields — existing "published" /
    # "pending" / "cancelled" count fields above are left untouched.
    def _qty_of(u):
        return float(u.get("total_available_qty", 0) or 0)

    def _items_of(u):
        return int(u.get("item_count", u.get("rows_imported", 0)) or 0)

    def _value_of(u):
        return float(u.get("total_value", 0) or 0)

    uploaded_items = sum(_items_of(u) for u in active_uploads)
    uploaded_qty = sum(_qty_of(u) for u in active_uploads)
    uploaded_value = sum(_value_of(u) for u in active_uploads)

    published_items = sum(_items_of(u) for u in uploads_today if u.get("status") != "Cancelled" and u.get("publish_status") == "Published")
    published_qty = sum(_qty_of(u) for u in uploads_today if u.get("status") != "Cancelled" and u.get("publish_status") == "Published")
    published_value = sum(_value_of(u) for u in uploads_today if u.get("status") != "Cancelled" and u.get("publish_status") == "Published")

    cancelled_qty = sum(_qty_of(u) for u in uploads_today if u.get("status") == "Cancelled")
    pending_items = sum(_items_of(u) for u in uploads_today if u.get("status") != "Cancelled" and u.get("publish_status") != "Published")
    pending_qty = sum(_qty_of(u) for u in uploads_today if u.get("status") != "Cancelled" and u.get("publish_status") != "Published")
    pending_value = sum(_value_of(u) for u in uploads_today if u.get("status") != "Cancelled" and u.get("publish_status") != "Published")

    completed_pairs = {
        (u.get("dealer_name"), u.get("branch"))
        for u in active_uploads if u.get("dealer_name") and u.get("branch")
    }
    completed_uploads = len(completed_pairs)

    branch_query = {}
    if not _is_all_scope(dealer):
        branch_query["dealer"] = dealer
    elif current_user.role != "master":
        branch_query["dealer"] = current_user.group

    branch_docs = await db.branches.find(branch_query, {"_id": 0}).to_list(10000)
    expected_pairs = {(b.get("dealer"), b.get("name")) for b in branch_docs if b.get("dealer") and b.get("name")}
    if not _is_all_scope(branch):
        expected_pairs = {p for p in expected_pairs if p[1] == branch}
    expected_uploads = len(expected_pairs)
    balance_uploads = max(expected_uploads - completed_uploads, 0)

    return {
        "brandsUploaded": brands_uploaded,
        "dealersUploaded": dealers_uploaded,
        "branchesUploaded": branches_uploaded,
        "expectedUploads": expected_uploads,
        "completedUploads": completed_uploads,
        "balanceUploads": balance_uploads,
        "published": published,
        "pending": pending,
        "cancelled": cancelled,
        "uploadedItems": uploaded_items,
        "uploadedQty": uploaded_qty,
        "uploadedValue": uploaded_value,
        "publishedItems": published_items,
        "publishedQty": published_qty,
        "publishedValue": published_value,
        "pendingItems": pending_items,
        "pendingQty": pending_qty,
        "pendingValue": pending_value,
        "cancelledQty": cancelled_qty,
    }


@api_router.get("/uploads/master-summary/balance-details")
async def get_upload_balance_details(
    brand: Optional[str] = None,
    dealer: Optional[str] = None,
    branch: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user),
):
    """Dealer/branch pairs that have not uploaded today (scoped)."""
    await _ensure_master_or_admin(current_user)
    date_key = _nmts_date_key()
    upload_query = {"date_key": date_key}
    _apply_role_scope_v2(upload_query, current_user, brand, dealer, branch)
    uploads_today = await db.uploads.find(upload_query, {"_id": 0, "raw_file_bytes": 0}).to_list(200000)
    active_uploads = [u for u in uploads_today if u.get("status") != "Cancelled"]
    completed_pairs = {
        (u.get("dealer_name"), u.get("branch"))
        for u in active_uploads if u.get("dealer_name") and u.get("branch")
    }
    branch_query = {}
    if not _is_all_scope(dealer):
        branch_query["dealer"] = dealer
    elif current_user.role != "master":
        branch_query["dealer"] = current_user.group
    branch_docs = await db.branches.find(branch_query, {"_id": 0}).to_list(10000)
    expected_pairs = {(b.get("dealer"), b.get("name")) for b in branch_docs if b.get("dealer") and b.get("name")}
    if not _is_all_scope(branch):
        expected_pairs = {p for p in expected_pairs if p[1] == branch}
    pending = []
    completed = []
    for dealer_name, branch_name in sorted(expected_pairs):
        row = {
            "dealer": dealer_name,
            "branch": branch_name,
            "brand": brand if brand and not str(brand).startswith("All") else (current_user.brand if current_user.role != "master" else ""),
            "upload_status": "Completed" if (dealer_name, branch_name) in completed_pairs else "Pending",
        }
        if row["upload_status"] == "Completed":
            completed.append(row)
        else:
            pending.append(row)
    return {
        "date_key": date_key,
        "expected_uploads": len(expected_pairs),
        "completed_uploads": len(completed_pairs),
        "balance_uploads": max(len(expected_pairs) - len(completed_pairs), 0),
        "completed": completed,
        "pending": pending,
    }


@api_router.get("/uploads/today-summary")
async def get_today_upload_summary(
    brand: Optional[str] = None,
    dealer: Optional[str] = None,
    branch: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Today's own Upload Center summary. For role 'user' this is always scoped to
    that user's own uploads only, regardless of any filter passed in — a User must
    never see another user's uploaded data. For Master/Admin it follows the normal
    Brand/Dealer/Branch role scope (same as /uploads/master-summary) so the same
    endpoint can back a personal 'my today upload' card anywhere it's needed.
    Reads db.upload_items (the per-row upload records) directly so Available Items
    reflects actual rows with available quantity > 0, not just upload batch counts.
    """
    date_key = _nmts_date_key()
    item_query = {"upload_type": "product", "active_date_key": date_key, "publish_status": {"$ne": "Cancelled"}}

    if current_user.role == "user":
        item_query["uploaded_user_id"] = current_user.id
    else:
        _apply_role_scope_v2(item_query, current_user, brand, dealer, branch)

    pipeline = [
        {"$match": item_query},
        {"$group": {
            "_id": None,
            "total_items": {"$sum": 1},
            "available_items": {"$sum": {"$cond": [{"$gt": [{"$toDouble": {"$ifNull": ["$available_qty_number", 0]}}, 0]}, 1, 0]}},
            "available_qty": {"$sum": {"$toDouble": {"$ifNull": ["$available_qty_number", 0]}}},
            "total_value": {"$sum": {"$toDouble": {"$ifNull": ["$total_value_number", 0]}}},
            "published_items": {"$sum": {"$cond": [{"$eq": ["$publish_status", "Published"]}, 1, 0]}},
            "published_qty": {"$sum": {"$cond": [{"$eq": ["$publish_status", "Published"]}, {"$toDouble": {"$ifNull": ["$available_qty_number", 0]}}, 0]}},
            "published_value": {"$sum": {"$cond": [{"$eq": ["$publish_status", "Published"]}, {"$toDouble": {"$ifNull": ["$total_value_number", 0]}}, 0]}},
            "pending_items": {"$sum": {"$cond": [{"$ne": ["$publish_status", "Published"]}, 1, 0]}},
            "pending_qty": {"$sum": {"$cond": [{"$ne": ["$publish_status", "Published"]}, {"$toDouble": {"$ifNull": ["$available_qty_number", 0]}}, 0]}},
            "pending_value": {"$sum": {"$cond": [{"$ne": ["$publish_status", "Published"]}, {"$toDouble": {"$ifNull": ["$total_value_number", 0]}}, 0]}},
        }},
    ]
    result = await db.upload_items.aggregate(pipeline).to_list(1)
    row = result[0] if result else {}

    return {
        "todayUploadedItems": row.get("total_items", 0),
        "todayUploadedAvailableItems": row.get("available_items", 0),
        "todayUploadedAvailableQty": row.get("available_qty", 0.0),
        "todayUploadedValue": row.get("total_value", 0.0),
        "publishedItems": row.get("published_items", 0),
        "publishedQty": row.get("published_qty", 0.0),
        "publishedValue": row.get("published_value", 0.0),
        "pendingItems": row.get("pending_items", 0),
        "pendingQty": row.get("pending_qty", 0.0),
        "pendingValue": row.get("pending_value", 0.0),
    }


@api_router.post("/templates/upload")
async def upload_template(
    brand: str = Form(...),
    templateType: Optional[str] = Form(None),
    template_type: Optional[str] = Form(None),
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user)
):
    await _ensure_master(current_user)

    final_template_type = templateType or template_type
    if final_template_type not in ["Product Hub", "Order Desk"]:
        raise HTTPException(status_code=400, detail="Invalid template type")

    content = await file.read()
    template_id = str(uuid.uuid4())
    stored = await file_objects.store_bytes(
        module="templates",
        relative_key=f"{brand}/{template_id}_{file.filename}",
        data=content,
        original_filename=file.filename or "template.xlsx",
        content_type=file.content_type,
    )

    await db.templates.insert_one({
        "id": template_id,
        "brand": brand,
        "templateType": final_template_type,
        "template_type": final_template_type,
        "fileName": file.filename,
        "contentType": file.content_type,
        # New templates store binaries in S3/object-store; keep fileBytes absent.
        "storage_provider": stored["storage_provider"],
        "storage_key": stored["storage_key"],
        "file_size": stored["file_size"],
        "sha256": stored["sha256"],
        "archived_at": stored["archived_at"],
        "uploadedBy": current_user.username,
        "uploadedById": current_user.id,
        "uploadedAt": _now_iso(),
        "status": "active"
    })

    return {"message": "Template uploaded successfully"}


@api_router.get("/templates")
async def list_templates(current_user: UserResponse = Depends(get_current_user)):
    query = {}
    if current_user.role != "master":
        query["brand"] = current_user.brand

    return await db.templates.find(query, {"_id": 0, "fileBytes": 0}).sort("uploadedAt", -1).to_list(1000)


@api_router.get("/templates/download/{template_id}")
async def download_template(template_id: str, current_user: UserResponse = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse

    template = await db.templates.find_one({"id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    if current_user.role != "master" and template.get("brand") != current_user.brand:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Prefer S3/object-store; fall back to legacy fileBytes for old templates.
    if template.get("storage_key") or template.get("fileBytes") is not None:
        return file_objects.streaming_response_from_meta(
            {
                **template,
                "original_filename": template.get("fileName") or "template.xlsx",
                "content_type": template.get("contentType"),
            },
            filename=template.get("fileName") or "template.xlsx",
        )

    return StreamingResponse(
        BytesIO(b""),
        media_type=template.get("contentType") or "application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={template.get('fileName', 'template.xlsx')}"}
    )


@api_router.delete("/templates/{template_id}")
async def delete_template(template_id: str, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)

    result = await db.templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")

    return {"message": "Template deleted successfully"}



# ==================== NMTS PRODUCT UPLOAD PIPELINE V2 ====================

NMTS_TIMEZONE = ZoneInfo("Asia/Kolkata")


def _nmts_now():
    """Return the business clock in India Standard Time.

    Product Hub is a daily operational view, so using UTC here caused the
    previous day's published stock to remain visible until 05:30 IST.
    """
    return datetime.now(NMTS_TIMEZONE)


def _nmts_date_key(dt=None):
    value = dt or _nmts_now()
    if value.tzinfo is None:
        value = value.replace(tzinfo=NMTS_TIMEZONE)
    else:
        value = value.astimezone(NMTS_TIMEZONE)
    return value.strftime("%Y%m%d")


def _nmts_display_date(dt=None):
    return (dt or _nmts_now()).strftime("%d-%m-%Y")


def _nmts_display_time(dt=None):
    return (dt or _nmts_now()).strftime("%H:%M:%S")


def _is_all_scope(value: str) -> bool:
    return not value or str(value).startswith("All ") or value == "N/A" or value == "all"


# Allowed Part Type values for Product Hub + Analytics. Upload never rejects a
# row for category reasons — aliases normalize common spellings (including the
# legacy "Genuine Parts" / "Non OEM parts" labels) to:
#   OE Parts | Accessories | Others
try:
    from part_category import (  # type: ignore
        PART_CATEGORY_OPTIONS,
        PART_TYPE_OPTIONS,
        is_all_part_type as _is_all_part_type,
        normalize_part_category as _normalize_part_category,
        part_type_mongo_clause as _part_type_mongo_clause,
    )
except ImportError:
    from .part_category import (  # type: ignore
        PART_CATEGORY_OPTIONS,
        PART_TYPE_OPTIONS,
        is_all_part_type as _is_all_part_type,
        normalize_part_category as _normalize_part_category,
        part_type_mongo_clause as _part_type_mongo_clause,
    )


def _format_uploaded_date(created_at) -> str:
    """Format the ISO `created_at` timestamp (set once at upload time) as a
    dd-mm-yyyy display date for the Uploaded Date column."""
    if not created_at:
        return ""
    text = str(created_at).strip()
    try:
        return datetime.fromisoformat(text).strftime("%d-%m-%Y")
    except ValueError:
        return format_date_for_display(text)


def _safe_float(value, default=0.0):
    try:
        if value in (None, ""):
            return default
        return float(str(value).replace(",", "").strip())
    except Exception:
        return default


def _format_excel_date_v2(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    text = str(value).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d-%m-%Y")
        except ValueError:
            pass
    return text


async def _build_upload_context_v2(current_user: UserResponse):
    brand_name = current_user.brand or ""
    brand_code = await resolve_brand_code_for_upload(brand_name)
    dealer_name = current_user.group or ""
    branch_name = current_user.location or ""
    return {
        "brand": brand_name,
        "brand_name": brand_name,
        "brand_code": brand_code,
        "dealer_name": dealer_name,
        "dealer_code": dealer_name or (current_user.user_id or current_user.id),
        "branch": branch_name,
        "uploaded_user_name": current_user.username,
        "uploaded_user_id": current_user.id,
        "uploaded_by": current_user.id,
        "user_code": current_user.user_id or current_user.id,
    }


def _apply_role_scope_v2(query: dict, current_user: UserResponse, brand=None, dealer=None, branch=None):
    if current_user.role == "master":
        if not _is_all_scope(brand):
            query["brand_name"] = brand
        if not _is_all_scope(dealer):
            query["dealer_name"] = dealer
        if not _is_all_scope(branch):
            query["branch"] = branch
    elif current_user.role == "admin":
        query["brand_name"] = current_user.brand
        query["dealer_name"] = current_user.group
        if not _is_all_scope(branch):
            query["branch"] = branch
    else:
        query["brand_name"] = current_user.brand
        query["dealer_name"] = current_user.group
        query["branch"] = current_user.location
    return query


@api_router.post("/upload/v2")
async def upload_product_center_v2(file: UploadFile = File(...), current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin", "user"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")

    now = _nmts_now()
    context = await _build_upload_context_v2(current_user)
    upload_no = await generate_upload_no("product", context["brand_code"])
    raw_bytes = await file.read()
    file_size = len(raw_bytes)

    # Store original Product Excel in S3/object-store (never embed multi-MB blobs in Mongo).
    stored_excel = None
    try:
        date_iso = _nmts_display_date(now)
        # _nmts_display_date may return DD-MM-YYYY; prefer ISO for S3 key path
        try:
            date_iso = now.astimezone(NMTS_TIMEZONE).date().isoformat()
        except Exception:
            date_iso = datetime.now(NMTS_TIMEZONE).date().isoformat()
        stored_excel = await file_objects.store_bytes(
            module="uploads",
            relative_key=f"{date_iso}/product-hub/{upload_no}_{file.filename}",
            data=raw_bytes,
            original_filename=file.filename or "product_upload.xlsx",
            content_type=file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        logger.error("Product Excel S3 store failed (upload continues): %s", exc)

    wb = openpyxl.load_workbook(BytesIO(raw_bytes), data_only=True)
    # Prefer a worksheet explicitly named "Inventory" (case-insensitive) since that is
    # the required worksheet name for the Product Hub inventory upload. Fall back to
    # the active sheet for any older workbook that doesn't use that sheet name, so
    # existing uploads never break.
    ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == "inventory":
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    rows_processed = 0
    rows_imported = 0
    rejected = []
    total_available_qty = 0.0
    total_value = 0.0
    upload_id = str(uuid.uuid4())
    item_docs = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        rows_processed += 1
        # Rule: reject a row ONLY when Part Number (column A) is missing. Part Name,
        # Quantity, Category, Brand, Branch, and Dealer must never cause a row to be
        # rejected — every other valid row is still imported.
        part_no = str(row[0] or "").strip() if len(row) > 0 else ""
        if not part_no:
            rejected.append({"row": row_idx, "reason": "Part No Missing"})
            continue
        part_name = str(row[1] or "").strip() if len(row) > 1 else ""
        loc = str(row[2] or "").strip() if len(row) > 2 else ""
        qty = _safe_float(row[3] if len(row) > 3 else 0, 0.0)
        last_receipt = _format_excel_date_v2(row[4] if len(row) > 4 else "")
        last_sales = _format_excel_date_v2(row[5] if len(row) > 5 else "")
        mav = _safe_float(row[6] if len(row) > 6 else 0, 0.0)
        # Column H — Part Category. Never required; stored as-is (trimmed) even if
        # it doesn't match one of the standard values below, so no row is ever
        # rejected because of it.
        part_category_raw = str(row[7] or "").strip() if len(row) > 7 else ""
        part_category = _normalize_part_category(part_category_raw)
        line_value = qty * mav
        total_available_qty += qty
        total_value += line_value
        rows_imported += 1
        item_docs.append({
            "id": str(uuid.uuid4()),
            "upload_id": upload_id,
            "upload_no": upload_no,
            "upload_type": "product",
            "part_number": part_no,
            "item_name": part_name,
            "loc": loc,
            "bin_location": loc,
            "quantity": qty,
            "available_qty": qty,
            "last_receipt_date": last_receipt,
            "last_sales_date": last_sales,
            "mav": mav,
            "mav_value": mav,
            "total_value": line_value,
            "part_category": part_category,
            # Explicit numeric fields (guaranteed float, never string) used for all
            # summary/aggregation math so quantity/value can never be mis-summed
            # due to string concatenation or mixed types coming from Excel.
            "available_qty_number": float(qty),
            "unit_value_number": float(mav),
            "total_value_number": float(line_value),
            "publish_status": "Waiting",
            "active_date_key": _nmts_date_key(now),
            "created_at": now.isoformat(),
            **context,
        })

    if item_docs:
        await db.upload_items.insert_many(item_docs)

    status_text = "Uploaded" if rows_imported else "Failed"
    upload_doc = {
        "id": upload_id,
        "upload_no": upload_no,
        "original_upload_no": upload_no,
        "upload_type": "product",
        "file_name": file.filename,
        "file_size": file_size,
        # Do not embed multi-MB Excel blobs in MongoDB — they blow Atlas free-tier
    # quotas and leave half-published Product Hub state when writes get blocked.
    # Keep file metadata + S3/object-store pointer; legacy uploads without storage_key still work.
    "raw_file_content_type": file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "storage_provider": (stored_excel or {}).get("storage_provider"),
    "storage_key": (stored_excel or {}).get("storage_key"),
    "original_filename": file.filename,
    "content_type": (stored_excel or {}).get("content_type") or (file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "sha256": (stored_excel or {}).get("sha256"),
    "archived_at": (stored_excel or {}).get("archived_at"),
    "rows_processed": rows_processed,
        "rows_imported": rows_imported,
        "failed_count": len(rejected),
        "item_count": rows_imported,
        "total_available_qty": total_available_qty,
        "total_value": total_value,
        "status": status_text,
        "publish_status": "Waiting" if rows_imported else "Failed",
        "errors": rejected,
        "upload_date": _nmts_display_date(now),
        "upload_time": _nmts_display_time(now),
        "date_key": _nmts_date_key(now),
        "created_at": now.isoformat(),
        **context,
    }
    await db.uploads.insert_one(upload_doc)

    return {
        "message": "Upload completed",
        "upload_no": upload_no,
        "upload_id": upload_id,
        "rows_processed": rows_processed,
        "rows_imported": rows_imported,
        "failed_count": len(rejected),
        "total_available_qty": total_available_qty,
        "total_value": total_value,
        "errors": rejected,
    }


@api_router.get("/uploads/v2")
async def get_uploads_v2(type: str = None, brand: str = None, dealer: str = None, branch: str = None, current_user: UserResponse = Depends(get_current_user)):
    query = {}
    if type:
        query["upload_type"] = type
    _apply_role_scope_v2(query, current_user, brand, dealer, branch)
    return await db.uploads.find(query, {"_id": 0, "raw_file_bytes": 0}).sort("created_at", -1).to_list(1000)


def _atlas_or_write_error_detail(exc: Exception) -> str:
    msg = str(exc or "")
    if "space quota" in msg.lower() or "8000" in msg:
        return (
            "Publish failed: MongoDB Atlas storage quota exceeded (writes blocked). "
            "Free cluster storage or upgrade the tier, then retry Publish. "
            "Existing Product Hub rows were left unchanged."
        )
    return f"Publish failed due to a database write error: {msg[:300]}"


def _product_hub_batch_totals_from_items(items: list) -> dict:
    batch_total_item = 0
    batch_available_item = 0
    batch_available_qty = 0.0
    batch_total_value = 0.0
    for item in items:
        qty_num = float(item.get("available_qty_number", item.get("quantity", 0)) or 0)
        unit_val_num = float(item.get("unit_value_number", item.get("mav_value", item.get("mav", 0)) or 0) or 0)
        total_val_num = float(item.get("total_value_number", qty_num * unit_val_num) or 0)
        batch_total_item += 1
        if qty_num > 0:
            batch_available_item += 1
        batch_available_qty += qty_num
        batch_total_value += total_val_num
    return {
        "total_item": batch_total_item,
        "available_item": batch_available_item,
        "available_qty": batch_available_qty,
        "total_value": batch_total_value,
    }


async def _finalize_product_publish(
    upload: dict,
    upload_id: str,
    date_key: str,
    now,
    current_user: UserResponse,
    totals: dict,
    *,
    already_published_products: bool,
):
    """Write batch_summaries + upload status. Safe to call when products already exist."""
    await db.batch_summaries.update_one(
        {
            "brand_name": upload.get("brand_name"),
            "dealer_name": upload.get("dealer_name"),
            "branch": upload.get("branch"),
            "active_date_key": date_key,
        },
        {"$set": {
            "brand_name": upload.get("brand_name"),
            "dealer_name": upload.get("dealer_name"),
            "branch": upload.get("branch"),
            "active_date_key": date_key,
            "upload_id": upload_id,
            "upload_no": upload.get("upload_no"),
            "total_item": totals["total_item"],
            "available_item": totals["available_item"],
            "available_qty": totals["available_qty"],
            "total_value": totals["total_value"],
            "published_at": now.isoformat(),
            "uploaded_user_name": upload.get("uploaded_user_name") or current_user.username,
        }},
        upsert=True,
    )
    await db.upload_items.update_many(
        {"upload_id": upload_id},
        {"$set": {"publish_status": "Published", "published_at": now.isoformat()}},
    )
    await db.uploads.update_one({"id": upload_id}, {"$set": {
        "publish_status": "Published",
        "status": "Ready to Send",
        "published_at": now.isoformat(),
        "published_by": current_user.id,
        "published_user_name": current_user.username,
        # Drop embedded Excel blob after successful publish to protect Atlas free-tier quota.
        "raw_file_cleared_reason": "cleared_after_publish",
    }, "$unset": {"raw_file_bytes": ""}})
    return {
        "message": (
            "Already published — Product Hub/status reconciled successfully"
            if already_published_products else
            "Published successfully"
        ),
        "items": totals["total_item"],
        "available_qty": totals["available_qty"],
        "total_value": totals["total_value"],
        "reconciled": already_published_products,
    }


@api_router.put("/uploads/{upload_id}/publish-v2")
async def publish_upload_v2(upload_id: str, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin", "user"]:
        raise HTTPException(status_code=403, detail="You are not allowed to publish uploads")
    upload = await db.uploads.find_one({"id": upload_id}, {"_id": 0, "raw_file_bytes": 0})
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    if upload.get("status") == "Cancelled" or upload.get("publish_status") == "Cancelled":
        raise HTTPException(status_code=400, detail="Cancelled upload cannot be published")
    if current_user.role == "admin" and (upload.get("brand_name") != current_user.brand or upload.get("dealer_name") != current_user.group):
        raise HTTPException(status_code=403, detail="Not allowed to publish this upload")
    if current_user.role == "user" and str(upload.get("uploaded_user_id") or "") != str(current_user.id):
        raise HTTPException(status_code=403, detail="Users can publish only their own uploads")

    now = _nmts_now()
    # Prefer the upload's own business date so late publish still lands on the
    # intended day; fall back to "today" for legacy rows without date_key.
    date_key = str(upload.get("date_key") or _nmts_date_key(now))
    existing_products = await db.products.count_documents({"upload_id": upload_id, "publish_status": "Published"})
    items = await db.upload_items.find({"upload_id": upload_id}, {"_id": 0}).to_list(200000)
    if not items and existing_products <= 0:
        raise HTTPException(status_code=400, detail="No upload items found")

    # Idempotent success when upload is already marked Published.
    if upload.get("publish_status") == "Published":
        batch = await db.batch_summaries.find_one({
            "brand_name": upload.get("brand_name"),
            "dealer_name": upload.get("dealer_name"),
            "branch": upload.get("branch"),
            "active_date_key": date_key,
            "upload_id": upload_id,
        }, {"_id": 0})
        if batch:
            return {
                "message": "Already published",
                "items": int(batch.get("total_item") or existing_products or len(items)),
                "available_qty": float(batch.get("available_qty") or 0),
                "total_value": float(batch.get("total_value") or 0),
                "reconciled": False,
                "already_published": True,
            }
        # Published flag set but summary missing — reconcile without re-insert.
        totals = _product_hub_batch_totals_from_items(items) if items else {
            "total_item": existing_products,
            "available_item": await db.products.count_documents({"upload_id": upload_id, "available_qty_number": {"$gt": 0}}),
            "available_qty": float(upload.get("total_available_qty") or 0),
            "total_value": float(upload.get("total_value") or 0),
        }
        try:
            return await _finalize_product_publish(
                upload, upload_id, date_key, now, current_user, totals, already_published_products=True,
            )
        except OperationFailure as exc:
            raise HTTPException(status_code=507, detail=_atlas_or_write_error_detail(exc))

    totals = _product_hub_batch_totals_from_items(items) if items else {
        "total_item": existing_products,
        "available_item": 0,
        "available_qty": float(upload.get("total_available_qty") or 0),
        "total_value": float(upload.get("total_value") or 0),
    }

    try:
        if upload.get("upload_type") == "product":
            # Partial-publish reconcile: products already exist for this upload_id.
            if existing_products > 0:
                await db.products.update_many(
                    {
                        "brand_name": upload.get("brand_name"),
                        "dealer_name": upload.get("dealer_name"),
                        "branch": upload.get("branch"),
                        "is_active_today": True,
                        "upload_id": {"$ne": upload_id},
                    },
                    {"$set": {"is_active_today": False}},
                )
                await db.products.update_many(
                    {"upload_id": upload_id},
                    {"$set": {
                        "is_active_today": True,
                        "active_date_key": date_key,
                        "publish_status": "Published",
                        "published_at": now.isoformat(),
                        "published_by": current_user.id,
                        "published_user_name": current_user.username,
                    }},
                )
                return await _finalize_product_publish(
                    upload, upload_id, date_key, now, current_user, totals, already_published_products=True,
                )

            # Fresh publish: deactivate prior active stock for this branch scope
            # (any previous business date), then insert today's batch once.
            await db.products.update_many({
                "brand_name": upload.get("brand_name"),
                "dealer_name": upload.get("dealer_name"),
                "branch": upload.get("branch"),
                "is_active_today": True,
            }, {"$set": {"is_active_today": False}})

            product_docs = []
            for item in items:
                doc = dict(item)
                qty_num = float(item.get("available_qty_number", item.get("quantity", 0)) or 0)
                unit_val_num = float(item.get("unit_value_number", item.get("mav_value", item.get("mav", 0)) or 0) or 0)
                total_val_num = float(item.get("total_value_number", qty_num * unit_val_num) or 0)
                if (doc.get("loc") or doc.get("bin_location")) and doc.get("location") is not None:
                    branch_meta = str(doc.get("branch") or "").strip()
                    location_meta = str(doc.get("location") or "").strip()
                    if branch_meta and location_meta.casefold() == branch_meta.casefold():
                        doc.pop("location", None)
                doc.update({
                    "id": str(uuid.uuid4()),
                    "upload_id": upload_id,
                    "upload_no": upload.get("upload_no"),
                    "published_at": now.isoformat(),
                    "published_by": current_user.id,
                    "published_user_name": current_user.username,
                    "publish_status": "Published",
                    "is_active_today": True,
                    "active_date_key": date_key,
                    "price": unit_val_num,
                    "category": "",
                    "upload_method": "excel",
                    "updated_at": now.isoformat(),
                    "available_qty_number": qty_num,
                    "unit_value_number": unit_val_num,
                    "total_value_number": total_val_num,
                    "quantity": qty_num,
                    "available_qty": qty_num,
                    "mav_value": unit_val_num,
                    "total_value": total_val_num,
                })
                product_docs.append(doc)

            if product_docs:
                await db.products.insert_many(product_docs)

            return await _finalize_product_publish(
                upload, upload_id, date_key, now, current_user, totals, already_published_products=False,
            )

        # Non-product uploads: mark published only.
        await db.upload_items.update_many(
            {"upload_id": upload_id},
            {"$set": {"publish_status": "Published", "published_at": now.isoformat()}},
        )
        await db.uploads.update_one({"id": upload_id}, {"$set": {
            "publish_status": "Published",
            "status": "Ready to Send",
            "published_at": now.isoformat(),
            "published_by": current_user.id,
            "published_user_name": current_user.username,
        }, "$unset": {"raw_file_bytes": ""}})
        return {"message": "Published successfully", "items": len(items)}
    except OperationFailure as exc:
        raise HTTPException(status_code=507, detail=_atlas_or_write_error_detail(exc))
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("publish-v2 failed for upload %s", upload_id)
        raise HTTPException(status_code=500, detail=f"Publish failed: {str(exc)[:300]}")


class CancelUploadRequest(BaseModel):
    reason: str


@api_router.put("/uploads/{upload_id}/cancel-v2")
async def cancel_upload_v2(upload_id: str, data: CancelUploadRequest, current_user: UserResponse = Depends(get_current_user)):
    if current_user.role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Only Master/Admin can cancel")
    upload = await db.uploads.find_one({"id": upload_id}, {"_id": 0})
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    if current_user.role == "admin" and (upload.get("brand_name") != current_user.brand or upload.get("dealer_name") != current_user.group):
        raise HTTPException(status_code=403, detail="Not allowed to cancel this upload")
    old_no = upload.get("upload_no") or ""
    cancel_no = old_no
    if old_no.startswith(("PU", "OU")):
        cancel_no = "CN" + old_no[2:]
    now = _nmts_now()
    reason = (data.reason or "Other").strip() or "Other"
    await db.uploads.update_one({"id": upload_id}, {"$set": {
        "upload_no": cancel_no,
        "cancelled_upload_no": cancel_no,
        "status": "Cancelled",
        "publish_status": "Cancelled",
        "cancel_reason": reason,
        "cancelled_at": now.isoformat(),
        "cancelled_by": current_user.id,
        "cancelled_user_name": current_user.username,
    }})
    await db.upload_items.update_many({"upload_id": upload_id}, {"$set": {"publish_status": "Cancelled", "upload_no": cancel_no, "cancel_reason": reason}})
    await db.products.update_many({"upload_id": upload_id}, {"$set": {"is_active_today": False, "publish_status": "Cancelled", "cancel_reason": reason}})
    await db.batch_summaries.delete_one({"upload_id": upload_id})
    return {"message": "Upload cancelled", "upload_no": cancel_no}


@api_router.get("/uploads/{upload_id}/raw-file")
async def download_raw_upload_file(upload_id: str, current_user: UserResponse = Depends(get_current_user)):
    excel_permissions.require_excel_export(current_user)
    upload = await db.uploads.find_one({"id": upload_id})
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    if current_user.role == "admin" and (upload.get("brand_name") != current_user.brand or upload.get("dealer_name") != current_user.group):
        raise HTTPException(status_code=403, detail="Not authorized")
    if current_user.role == "user" and upload.get("uploaded_user_id") != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    if not file_objects.meta_has_readable_bytes(upload):
        raise HTTPException(status_code=404, detail="Raw file not stored for this upload")
    return file_objects.streaming_response_from_meta(
        upload,
        filename=upload.get("file_name") or upload.get("original_filename") or "raw_upload.xlsx",
    )


@api_router.get("/products/active")
async def get_active_products_v2(brand: str = None, dealer: str = None, branch: str = None, current_user: UserResponse = Depends(get_current_user)):
    query = {"publish_status": "Published", "is_active_today": True, "active_date_key": _nmts_date_key()}
    _apply_role_scope_v2(query, current_user, brand, dealer, branch)
    return await db.products.find(query, {"_id": 0}).sort("part_number", 1).to_list(200000)


@api_router.get("/product-hub-history")
async def list_product_hub_history(
    date_key: str = None, from_date: str = None, to_date: str = None,
    brand: str = None, dealer: str = None, branch: str = None,
    current_user: UserResponse = Depends(get_current_user)
):
    scope = {}
    _apply_role_scope_v2(scope, current_user, brand, dealer, branch)
    try:
        return await hybrid_history.summarize_product_history(
            db,
            date_key=date_key,
            from_date=from_date,
            to_date=to_date,
            brand=scope.get("brand_name"),
            dealer=scope.get("dealer_name"),
            branch=scope.get("branch"),
        )
    except Exception as exc:
        logger.warning("Hybrid history list failed, using Mongo-only path: %s", exc)

    query = {"publish_status": "Published"}
    if date_key:
        query["active_date_key"] = date_key.replace("-", "")
    elif from_date or to_date:
        date_range = {}
        if from_date:
            date_range["$gte"] = from_date.replace("-", "")
        if to_date:
            date_range["$lte"] = to_date.replace("-", "")
        query["active_date_key"] = date_range
    _apply_role_scope_v2(query, current_user, brand, dealer, branch)
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": {"date": "$active_date_key", "brand": "$brand_name", "dealer": "$dealer_name", "branch": "$branch"},
            "records": {"$sum": 1},
            "total_available_qty": {"$sum": "$quantity"},
            "total_value": {"$sum": "$total_value"},
            "last_published_at": {"$max": "$published_at"},
        }},
        {"$sort": {"_id.date": -1, "_id.brand": 1, "_id.dealer": 1, "_id.branch": 1}},
    ]
    rows = await db.products.aggregate(pipeline).to_list(10000)
    return [{
        "date_key": r["_id"].get("date"),
        "brand": r["_id"].get("brand"),
        "dealer": r["_id"].get("dealer"),
        "branch": r["_id"].get("branch"),
        "records": r.get("records", 0),
        "total_available_qty": r.get("total_available_qty", 0),
        "total_value": r.get("total_value", 0),
        "last_published_at": r.get("last_published_at"),
    } for r in rows]


@api_router.get("/product-hub-history/download")
async def download_product_hub_history(date_key: str = None, from_date: str = None, to_date: str = None, brand: str = None, dealer: str = None, branch: str = None, current_user: UserResponse = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    excel_permissions.require_excel_export(current_user)
    scope = {}
    _apply_role_scope_v2(scope, current_user, brand, dealer, branch)
    result = await hybrid_history.read_product_history(
        db,
        date_key=date_key,
        from_date=from_date,
        to_date=to_date,
        brand=scope.get("brand_name"),
        dealer=scope.get("dealer_name"),
        branch=scope.get("branch"),
    )
    rows = result.get("rows") or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Hub History"
    headers = ["Part No", "Part Name", "LOC", "Available Qty", "Last Receipt Date", "Last Sales Date", "MAV", "Total Value", "Upload No", "Brand", "Dealer", "Branch"]
    ws.append(headers)
    for p in rows:
        ws.append([
            p.get("part_number", ""), p.get("item_name", ""), p.get("loc") or p.get("location", ""), p.get("quantity", 0),
            p.get("last_receipt_date", ""), p.get("last_sales_date", ""), p.get("mav_value", 0), p.get("total_value", 0),
            p.get("upload_no", ""), p.get("brand_name", ""), p.get("dealer_name", ""), p.get("branch", ""),
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    date_label = date_key or f"{from_date or 'START'}_to_{to_date or 'END'}"
    filename = f"Product_Hub_History_{date_label}_{brand or 'ALL'}_{branch or 'ALL'}.xlsx".replace(" ", "_")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@api_router.get("/product-hub-history/rows")
async def list_product_hub_history_rows(
    date_key: str = None, from_date: str = None, to_date: str = None,
    brand: str = None, dealer: str = None, branch: str = None,
    part_number: str = None, search: str = None,
    page: int = 1, page_size: int = 50,
    current_user: UserResponse = Depends(get_current_user),
):
    """Paginated historical product rows (Mongo hot or S3 cold) for on-screen viewing."""
    scope = {}
    _apply_role_scope_v2(scope, current_user, brand, dealer, branch)
    result = await hybrid_history.read_product_history(
        db,
        date_key=date_key,
        from_date=from_date,
        to_date=to_date,
        brand=scope.get("brand_name"),
        dealer=scope.get("dealer_name"),
        branch=scope.get("branch"),
        page=page,
        page_size=page_size,
        part_number=part_number,
        search=search,
    )
    return {
        "rows": result.get("rows") or [],
        "count": result.get("count") or 0,
        "total": result.get("total") or result.get("count") or 0,
        "sources": result.get("sources") or {},
        "page": result.get("page"),
        "mongo_count": result.get("mongo_count"),
        "s3_count": result.get("s3_count"),
    }


# ==================== PRODUCT HUB V3 - SCALABLE (LAKHS OF RECORDS) ====================
# Redesigned to avoid ever loading/scanning the full raw stock collection for
# summary numbers, and to avoid ever sending more than one page of raw rows
# to the frontend. All quantity/value math uses the pre-calculated numeric
# fields (available_qty_number / unit_value_number / total_value_number)
# written at upload time, so there is no string-concatenation risk.

def _product_hub_active_query(current_user: UserResponse, brand=None, dealer=None, branch=None):
    query = {"publish_status": "Published", "is_active_today": True, "active_date_key": _nmts_date_key()}
    _apply_role_scope_v2(query, current_user, brand, dealer, branch)
    return query


@api_router.get("/product-hub/summary")
async def product_hub_summary(
    brand: str = None, dealer: str = None, branch: str = None, search: str = None,
    category: str = None, stock_status: str = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Fast summary cards: Total Item, Total Available Item, Total Available Quantity, Total Value.
    When there is no search/Part Type/stock filter, this reads the small pre-aggregated
    batch_summaries collection. Part Type / search / stock filters require a products scan."""
    date_key = _nmts_date_key()
    search = (search or "").strip()
    needs_product_scan = bool(search) or (not _is_all_part_type(category) and not _is_all_scope(category)) or (
        (stock_status or "all").strip().lower() not in {"", "all"}
    )

    if not needs_product_scan:
        batch_query = {"active_date_key": date_key}
        if not _is_all_scope(brand):
            batch_query["brand_name"] = brand
        elif current_user.role != "master":
            batch_query["brand_name"] = current_user.brand
        if not _is_all_scope(dealer):
            batch_query["dealer_name"] = dealer
        elif current_user.role != "master":
            batch_query["dealer_name"] = current_user.group
        if not _is_all_scope(branch):
            batch_query["branch"] = branch
        elif current_user.role == "user":
            batch_query["branch"] = current_user.location

        rows = await db.batch_summaries.find(batch_query, {"_id": 0}).to_list(10000)
        summary = {
            "totalItem": sum(int(r.get("total_item", 0)) for r in rows),
            "totalAvailableItem": sum(int(r.get("available_item", 0)) for r in rows),
            "totalAvailableQty": sum(float(r.get("available_qty", 0)) for r in rows),
            "totalValue": sum(float(r.get("total_value", 0)) for r in rows),
        }
        # If pre-agg is missing (partial publish / quota failure), fall back to
        # the same active products query used by /product-hub/records.
        if rows and (summary["totalItem"] > 0 or summary["totalAvailableQty"] > 0 or summary["totalValue"] > 0):
            return summary
        needs_product_scan = True

    query = _product_hub_active_query(current_user, brand, dealer, branch)
    _apply_category_filter(query, category)
    _apply_stock_status_filter(query, stock_status)
    if search:
        safe_search = re.escape(search)
        search_clause = {
            "$or": [
                {"part_number": {"$regex": safe_search, "$options": "i"}},
                {"item_name": {"$regex": safe_search, "$options": "i"}},
            ]
        }
        if "$and" in query:
            query["$and"].append(search_clause)
        elif any(k.startswith("$") for k in query.keys()):
            existing = {k: v for k, v in list(query.items())}
            query.clear()
            query["$and"] = [existing, search_clause]
        else:
            query["$or"] = search_clause["$or"]
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": None,
            "total_item": {"$sum": 1},
            "total_available_item": {"$sum": {"$cond": [{"$gt": [{"$toDouble": {"$ifNull": ["$available_qty_number", 0]}}, 0]}, 1, 0]}},
            "total_available_qty": {"$sum": {"$toDouble": {"$ifNull": ["$available_qty_number", 0]}}},
            "total_value": {"$sum": {"$toDouble": {"$ifNull": ["$total_value_number", 0]}}},
        }},
    ]
    result = await db.products.aggregate(pipeline, allowDiskUse=True).to_list(1)
    row = result[0] if result else {}
    return {
        "totalItem": row.get("total_item", 0),
        "totalAvailableItem": row.get("total_available_item", 0),
        "totalAvailableQty": row.get("total_available_qty", 0.0),
        "totalValue": row.get("total_value", 0.0),
    }


@api_router.get("/product-hub/branch-summary")
async def product_hub_branch_summary(
    brand: str = None, dealer: str = None, branch: str = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Master Product Hub landing view: branch-wise summary cards (not raw records).
    Reads only the small pre-aggregated batch_summaries collection, so this stays
    fast no matter how many lakhs of rows exist in the raw products collection."""
    date_key = _nmts_date_key()
    query = {"active_date_key": date_key}
    if not _is_all_scope(brand):
        query["brand_name"] = brand
    elif current_user.role != "master":
        query["brand_name"] = current_user.brand
    if not _is_all_scope(dealer):
        query["dealer_name"] = dealer
    elif current_user.role != "master":
        query["dealer_name"] = current_user.group
    if not _is_all_scope(branch):
        query["branch"] = branch
    elif current_user.role == "user":
        query["branch"] = current_user.location

    rows = await db.batch_summaries.find(query, {"_id": 0}).sort("branch", 1).to_list(10000)
    return rows


def _apply_category_filter(query: dict, category: str = None):
    """Part Type filter — additive, never required.

    Matches the final Part Type options (OE Parts / Accessories / Others) and
    all known legacy aliases (Genuine Parts, Non OEM parts, etc.).
    """
    if _is_all_part_type(category) or _is_all_scope(category):
        return
    clause = _part_type_mongo_clause(category)
    if not clause:
        return
    # Merge into existing query without dropping other $or/$and conditions.
    if "$and" in query:
        query["$and"].append(clause)
    elif any(k.startswith("$") for k in query.keys()):
        existing = {k: v for k, v in list(query.items())}
        query.clear()
        query["$and"] = [existing, clause]
    else:
        # Simple field query — attach $or via $and with a copy of fields.
        fields = {k: v for k, v in list(query.items())}
        query.clear()
        query["$and"] = [fields, clause]


def _apply_stock_status_filter(query: dict, stock_status: str = None):
    """Filter Product Hub rows by available quantity without changing stored totals."""
    status = (stock_status or "all").strip().lower()
    qty_expr = {"$toDouble": {"$ifNull": ["$available_qty_number", 0]}}
    if status in {"available", "available_items", "in_stock"}:
        query["$expr"] = {"$gt": [qty_expr, 0]}
    elif status in {"zero", "zero_quantity", "not_available", "out_of_stock"}:
        query["$expr"] = {"$eq": [qty_expr, 0]}


def _apply_uploaded_date_range_filter(query: dict, from_date: str = None, to_date: str = None):
    """From Date / To Date filter on the item's original Uploaded Date
    (`created_at`, set once at upload time and carried through publish).
    Dates are plain YYYY-MM-DD strings from the date-picker; ISO datetime strings
    sort lexicographically so range comparison works without parsing."""
    from_date = (from_date or "").strip()
    to_date = (to_date or "").strip()
    if not from_date and not to_date:
        return
    date_range = {}
    if from_date:
        date_range["$gte"] = f"{from_date}T00:00:00"
    if to_date:
        date_range["$lte"] = f"{to_date}T23:59:59.999999+00:00"
    if date_range:
        query["created_at"] = date_range


@api_router.get("/product-hub/records")
async def product_hub_records(
    brand: str = None, dealer: str = None, branch: str = None, search: str = None,
    category: str = None, stock_status: str = None, from_date: str = None, to_date: str = None,
    page: int = 1, page_size: int = 300,
    current_user: UserResponse = Depends(get_current_user)
):
    """Paginated Product Hub raw records. Search/filter happen entirely on the
    backend so the frontend never needs to hold more than one page in memory."""
    page = max(page, 1)
    page_size = max(1, min(page_size, 1000))

    query = _product_hub_active_query(current_user, brand, dealer, branch)
    _apply_category_filter(query, category)
    _apply_stock_status_filter(query, stock_status)
    _apply_uploaded_date_range_filter(query, from_date, to_date)
    search = (search or "").strip()
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"part_number": {"$regex": safe_search, "$options": "i"}},
            {"item_name": {"$regex": safe_search, "$options": "i"}},
        ]

    total = await db.products.count_documents(query)
    rows = await db.products.find(query, {"_id": 0}).sort("part_number", 1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    for row in rows:
        # Backward compatible: compute from stored numeric fields first, then
        # from the last receipt/sales date, without ever overwriting whichever
        # value is already stored (purchase and sales aging stay independent).
        row["purchase_aging_days"] = row.get("purchase_aging_days", _order_stock_purchase_aging_days(row))
        row["sales_aging_days"] = row.get("sales_aging_days", _order_stock_sales_aging_days(row))
        row["purchase_aging"] = row["purchase_aging_days"]
        row["sales_aging"] = row["sales_aging_days"]

    return {
        "records": rows,
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": max(1, (total + page_size - 1) // page_size),
    }


def _write_products_sheet(ws, rows):
    # Note: existing columns/order are kept exactly as-is (export contract preserved).
    # Part Category, Uploaded Date, Uploaded User, Active Status, Purchase Aging
    # and Sales Aging are appended at the end so any existing consumer reading
    # the first 12 columns is unaffected.
    ws.append([
        "Part Number", "Part Name", "Location", "Available Qty", "Last Receipt Date", "Last Sales Date", "MAV", "Total Value", "Upload No", "Brand", "Dealer", "Branch",
        "Part Category", "Uploaded Date", "Uploaded User", "Active Status", "Purchase Aging Days", "Sales Aging Days",
    ])
    for p in rows:
        ws.append([
            p.get("part_number", ""), p.get("item_name", ""), p.get("loc") or p.get("location", ""),
            float(p.get("available_qty_number", p.get("quantity", 0)) or 0),
            p.get("last_receipt_date", ""), p.get("last_sales_date", ""),
            float(p.get("unit_value_number", p.get("mav_value", 0)) or 0),
            float(p.get("total_value_number", p.get("total_value", 0)) or 0),
            p.get("upload_no", ""), p.get("brand_name", ""), p.get("dealer_name", ""), p.get("branch", ""),
            p.get("part_category", ""), _format_uploaded_date(p.get("created_at", "")), p.get("uploaded_user_name", ""),
            "Active" if p.get("is_active_today") else "Inactive",
            p.get("purchase_aging_days", _order_stock_purchase_aging_days(p)),
            p.get("sales_aging_days", _order_stock_sales_aging_days(p)),
        ])


@api_router.get("/product-hub/export/branch")
async def export_product_hub_branch(
    brand: str, dealer: str, branch: str,
    category: str = None, stock_status: str = None, from_date: str = None, to_date: str = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Single-branch Excel export, generated entirely on the backend and streamed
    to the browser as one file — never sends raw rows to the frontend for export."""
    from fastapi.responses import StreamingResponse
    excel_permissions.require_excel_export(current_user)

    query = _product_hub_active_query(current_user, brand, dealer, branch)
    _apply_category_filter(query, category)
    _apply_stock_status_filter(query, stock_status)
    _apply_uploaded_date_range_filter(query, from_date, to_date)
    cursor = db.products.find(query, {"_id": 0}).sort("part_number", 1).batch_size(1000)
    rows = await cursor.to_list(300000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = branch[:31] if branch else "Branch"
    _write_products_sheet(ws, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ProductHub_{brand}_{dealer}_{branch}.xlsx".replace(" ", "_")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@api_router.get("/product-hub/export/master")
async def export_product_hub_master(
    brand: str = None, dealer: str = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Master-only full export. Streams one ZIP containing a separate Excel
    per branch plus one Master Summary Excel, so large exports never freeze the
    browser and never require holding lakhs of rows in memory at once (only one
    branch's rows are held in memory at any given time)."""
    import zipfile
    from fastapi.responses import StreamingResponse

    await _ensure_master(current_user)
    date_key = _nmts_date_key()

    summary_query = {"active_date_key": date_key}
    if not _is_all_scope(brand):
        summary_query["brand_name"] = brand
    if not _is_all_scope(dealer):
        summary_query["dealer_name"] = dealer

    branch_summaries = await db.batch_summaries.find(summary_query, {"_id": 0}).sort([("brand_name", 1), ("dealer_name", 1), ("branch", 1)]).to_list(10000)

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Master Summary sheet
        summary_wb = openpyxl.Workbook()
        summary_ws = summary_wb.active
        summary_ws.title = "Master Summary"
        summary_ws.append(["Brand", "Dealer", "Branch", "Total Item", "Available Item", "Available Qty", "Total Value", "Last Upload No", "Uploaded User"])
        for b in branch_summaries:
            summary_ws.append([
                b.get("brand_name", ""), b.get("dealer_name", ""), b.get("branch", ""),
                int(b.get("total_item", 0)), int(b.get("available_item", 0)),
                float(b.get("available_qty", 0)), float(b.get("total_value", 0)),
                b.get("upload_no", ""), b.get("uploaded_user_name", ""),
            ])
        summary_out = BytesIO()
        summary_wb.save(summary_out)
        summary_out.seek(0)
        zf.writestr("Master_Summary.xlsx", summary_out.read())

        # One Excel per branch, streamed/queried one at a time to bound memory use.
        for b in branch_summaries:
            branch_query = {
                "publish_status": "Published", "is_active_today": True, "active_date_key": date_key,
                "brand_name": b.get("brand_name"), "dealer_name": b.get("dealer_name"), "branch": b.get("branch"),
            }
            cursor = db.products.find(branch_query, {"_id": 0}).sort("part_number", 1).batch_size(1000)
            rows = await cursor.to_list(300000)

            branch_wb = openpyxl.Workbook()
            branch_ws = branch_wb.active
            branch_ws.title = str(b.get("branch") or "Branch")[:31]
            _write_products_sheet(branch_ws, rows)
            branch_out = BytesIO()
            branch_wb.save(branch_out)
            branch_out.seek(0)

            safe_name = f"{b.get('brand_name','')}_{b.get('dealer_name','')}_{b.get('branch','')}".replace(" ", "_").replace("/", "-")
            zf.writestr(f"{safe_name}.xlsx", branch_out.read())

    zip_buffer.seek(0)
    return StreamingResponse(zip_buffer, media_type="application/zip", headers={"Content-Disposition": "attachment; filename=ProductHub_Full_Export.zip"})


# ==================== ORDER DESK V2 ROUTES ====================

def _order_clean_text(value):
    return str(value or '').strip()


def _order_stock_purchase_aging_days(stock: dict):
    """Purchase (receipt) aging in days, from stored numeric fields first, then
    derived from the last receipt date if no numeric value was stored."""
    candidate_fields = (
        'purchase_ageing_days', 'purchase_aging_days',
        'receipt_ageing', 'receipt_aging',
        'aging_days', 'ageing_days',  # legacy single-value field, historically meant purchase aging
    )
    for field in candidate_fields:
        value = stock.get(field)
        if value in (None, ''):
            continue
        try:
            return max(0, int(float(str(value).replace(',', '').strip())))
        except (TypeError, ValueError):
            continue
    for field in ('last_receipt_date', 'last_purchase_date', 'receipt_date'):
        value = stock.get(field)
        if value in (None, ''):
            continue
        calculated = calculate_ageing_days(value)
        if calculated is not None:
            return int(calculated)
    return 0


def _order_stock_sales_aging_days(stock: dict):
    """Sales aging in days, from stored numeric fields first, then derived
    from the last sales date if no numeric value was stored."""
    candidate_fields = ('sales_ageing_days', 'sales_aging_days', 'sale_ageing', 'sale_aging')
    for field in candidate_fields:
        value = stock.get(field)
        if value in (None, ''):
            continue
        try:
            return max(0, int(float(str(value).replace(',', '').strip())))
        except (TypeError, ValueError):
            continue
    for field in ('last_sales_date', 'last_sale_date', 'sales_date', 'sale_date'):
        value = stock.get(field)
        if value in (None, ''):
            continue
        calculated = calculate_ageing_days(value)
        if calculated is not None:
            return int(calculated)
    return 0


def _order_stock_aging_days(stock: dict):
    """Backward-compatible single aging value. Defaults to Purchase Aging,
    kept for any existing frontend logic or API consumer that still reads
    the plain 'aging_days' field on an availability source."""
    return _order_stock_purchase_aging_days(stock)


async def _reservation_qty_map(stock_ids: list) -> dict:
    """Sum of active reserved qty per stock_id across the whole system, so a
    unit of stock already reserved by any sent request (this order or any
    other) is never suggested or double-allocated again until it is released
    by a Reject/Cancel."""
    ids = [s for s in set(stock_ids) if s]
    if not ids:
        return {}
    totals = {}
    cursor = db.stock_reservations.aggregate([
        {'$match': {'stock_id': {'$in': ids}, 'status': 'active'}},
        {'$group': {'_id': '$stock_id', 'qty': {'$sum': '$qty'}}},
    ])
    async for doc in cursor:
        totals[doc['_id']] = float(doc.get('qty') or 0)
    return totals


def _with_reservation_adjustment(entry: dict, reserved_map: dict) -> dict:
    """Annotate an availability source with reserved_qty / net_available_qty
    without touching its original (gross) available_qty field, so existing
    manual-selection UI and API consumers keep seeing the same available_qty
    they always have."""
    stock_id = entry.get('stock_id')
    reserved = float(reserved_map.get(stock_id, 0) or 0)
    gross = float(entry.get('available_qty') or 0)
    net = max(0.0, gross - reserved)
    return {**entry, 'reserved_qty': reserved, 'net_available_qty': net}


def _aging_value_for_suggest(entry: dict, aging_type: str) -> float:
    if aging_type == 'sales':
        return float(entry.get('sales_aging_days') or 0)
    return float(entry.get('purchase_aging_days') or entry.get('aging_days') or 0)


def _auto_suggest_sort_key(entry: dict, aging_type: str):
    """Highest Aging First; if equal, Highest Available Quantity first; if
    still equal, Oldest Uploaded Stock first; if still equal, Alphabetical
    Branch. Built as an ascending sort key (negate the descending fields)."""
    aging = _aging_value_for_suggest(entry, aging_type)
    qty = float(entry.get('net_available_qty', entry.get('available_qty', 0)) or 0)
    last_upload = entry.get('last_upload') or ''
    branch = _order_clean_text(entry.get('branch')).lower()
    return (-aging, -qty, last_upload, branch)


def _order_number_prefix(brand_code: str):
    clean = re.sub(r'[^A-Za-z0-9]', '', str(brand_code or 'XX')).upper()[:2]
    return clean or 'XX'


async def _generate_order_number(brand_code: str):
    """Generate a permanent order number without hyphens, e.g. ORHY260712001."""
    date_key = datetime.now(timezone.utc).strftime('%y%m%d')
    brand = _order_number_prefix(brand_code)
    counter_id = f'order_OR_{brand}_{date_key}'
    counter = await db.counters.find_one_and_update(
        {'_id': counter_id},
        {'$inc': {'seq': 1}, '$setOnInsert': {'date_key': date_key, 'brand_code': brand}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return f"OR{brand}{date_key}{int(counter.get('seq', 1)):03d}"


async def _generate_request_number(brand_code: str) -> dict:
    """Concurrency-safe Parts Transfer Request number:
    RQ + BRAND_CODE + YYMMDD + 4-digit daily serial, e.g. RQHY2607150001.
    The counter is a per-brand, per-day atomic $inc (same pattern as
    _generate_order_number / generate_upload_no), so two simultaneous
    requests for the same brand/day can never receive the same number, and
    the daily serial naturally restarts at 0001 for a new date or brand."""
    date_key = datetime.now(timezone.utc).strftime('%y%m%d')
    clean_brand = re.sub(r'[^A-Za-z0-9]', '', str(brand_code or 'XX')).upper() or 'XX'
    counter_id = f'request_RQ_{clean_brand}_{date_key}'
    counter = await db.counters.find_one_and_update(
        {'_id': counter_id},
        {'$inc': {'seq': 1}, '$setOnInsert': {'date_key': date_key, 'brand_code': clean_brand, 'type': 'RQ'}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    seq = int(counter.get('seq', 1))
    if seq > 9999:
        # Never silently wrap into a colliding number.
        raise HTTPException(status_code=500, detail=f'Daily request sequence exhausted for brand {clean_brand}. Contact support.')
    return {
        'request_number': f'RQ{clean_brand}{date_key}{seq:04d}',
        'brand_code': clean_brand, 'date_key': date_key, 'sequence': seq,
    }


def _send_requests_fingerprint(order_id: str, items: list) -> str:
    """SHA-256 content fingerprint of an order's current allocations, used to
    detect a repeated Send Request click / frontend retry / API retry with
    identical content so a second request/request-number is never created."""
    parts = [order_id]
    for item in sorted(items, key=lambda i: i.get('id', '')):
        sources = sorted(
            item.get('allocations', []) or [],
            key=lambda s: (_order_clean_text(s.get('dealer_name')), _order_clean_text(s.get('branch'))),
        )
        for source in sources:
            parts.append('|'.join([
                item.get('id', ''),
                _order_clean_text(source.get('dealer_name')),
                _order_clean_text(source.get('branch')),
                str(source.get('request_qty', '')),
            ]))
    raw = '::'.join(parts)
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


async def _resolve_request_receiver_email(brand: str, dealer: str, branch: str) -> str:
    """Best single receiver for a Requested-To destination: an active Admin
    scoped to the exact supplying dealer+branch, else an active Master for
    the supplying brand. Never guesses — returns '' if nothing matches."""
    query = {
        'status': {'$regex': '^active$', '$options': 'i'},
        'group': dealer, 'location': branch, 'role': 'admin',
    }
    user = await db.users.find_one(query, {'_id': 0, 'email': 1, 'id': 1})
    if not user or not (user.get('email') or '').strip():
        master_query = {'status': {'$regex': '^active$', '$options': 'i'}, 'role': 'master'}
        if brand:
            master_query['brand'] = brand
        user = await db.users.find_one(master_query, {'_id': 0, 'email': 1, 'id': 1})
    return ((user or {}).get('email') or '').strip()


async def _send_request_group_email(group_doc: dict):
    """Best-effort PDF + Gmail dispatch for one Requested-To destination
    group. Never raises — the request/request number/items are already
    saved before this runs, so any failure here only updates email_* status
    and notification_logs, and never rolls back the saved request."""
    now = datetime.now(timezone.utc).isoformat()
    receiver_email = await _resolve_request_receiver_email(
        group_doc.get('supplying_brand'), group_doc.get('supplying_dealer'), group_doc.get('supplying_branch'),
    )
    log_id = str(uuid.uuid4())
    subject = f"Parts Transfer Request - {group_doc['request_number']}"
    base_log = {
        'id': log_id, 'request_id': group_doc['id'], 'request_number': group_doc['request_number'],
        'order_id': group_doc['order_id'], 'receiver_user_id': '', 'receiver_email': receiver_email or '',
        'notification_type': 'parts_transfer_request', 'channel': 'email', 'subject': subject,
        'attachment_filename': group_doc['pdf_filename'], 'retry_count': group_doc.get('retry_count', 0),
        'created_at': now,
    }

    if not receiver_email or not notifications.is_valid_email(receiver_email):
        await db.request_headers.update_one({'id': group_doc['id']}, {'$set': {
            'email_sent': False, 'email_status': 'failed', 'email_error': 'Receiver email not configured',
            'receiver_email': receiver_email or '', 'updated_at': now,
        }})
        await db.notification_logs.insert_one({**base_log, 'status': 'failed', 'attempted_at': now, 'sent_at': None,
                                                'failed_at': now, 'error_message': 'Receiver email not configured'})
        return

    try:
        pdf_bytes = notifications.build_request_pdf(group_doc)
    except Exception as exc:  # noqa: BLE001 — a PDF failure must never break the saved request
        safe_error = f'PDF generation failed: {str(exc)[:250]}'
        await db.request_headers.update_one({'id': group_doc['id']}, {'$set': {
            'email_sent': False, 'email_status': 'failed', 'email_error': safe_error,
            'receiver_email': receiver_email, 'updated_at': now,
        }})
        await db.notification_logs.insert_one({**base_log, 'status': 'failed', 'attempted_at': now, 'sent_at': None,
                                                'failed_at': now, 'error_message': safe_error})
        return

    requester_cc = (group_doc.get('requester_email') or '').strip()
    if not notifications.is_valid_email(requester_cc) or requester_cc.lower() == receiver_email.lower(): requester_cc = ''
    result = await asyncio.get_event_loop().run_in_executor(
        None, notifications.send_request_pdf_email, receiver_email, group_doc, pdf_bytes, requester_cc,
    )
    sent = result.get('status') == 'sent'
    await db.request_headers.update_one({'id': group_doc['id']}, {'$set': {
        'email_sent': sent, 'email_status': result.get('status'), 'email_error': (result.get('error') or None) if not sent else None,
        'email_sent_at': now if sent else None, 'receiver_email': receiver_email,
        'notification_log_id': log_id, 'updated_at': now,
    }})
    await db.notification_logs.insert_one({
        **base_log, 'status': result.get('status'), 'attempted_at': now,
        'sent_at': now if sent else None, 'failed_at': None if sent else now,
        'error_message': result.get('error') or '', 'provider_message_id': result.get('provider_response', ''),
    })


def _send_requests_response(group_docs: list, duplicate: bool):
    groups = [g for g in group_docs if g]
    request_numbers = [g.get('request_number') for g in groups]
    any_email_sent = any(g.get('email_sent') for g in groups)
    any_email_failed = any(not g.get('email_sent') for g in groups)
    errors = [g.get('email_error') for g in groups if g.get('email_error')]

    if duplicate:
        message = f'Request already created for this order ({len(request_numbers)} request number(s)). No duplicate was created.'
    elif not groups:
        message = 'No requests were created.'
    elif any_email_sent and not any_email_failed:
        message = f'{len(request_numbers)} request(s) created and emailed successfully.'
    elif any_email_sent and any_email_failed:
        message = f'{len(request_numbers)} request(s) created. Some emails failed to send.'
    else:
        message = f'{len(request_numbers)} request(s) created successfully, but email could not be sent.'

    return {
        'request_created': bool(groups), 'message': message, 'request_numbers': request_numbers,
        'email_sent': any_email_sent, 'email_error': errors[0] if errors else None,
        'groups': groups, 'duplicate': duplicate,
    }


def _parse_order_rows_from_worksheet(ws):
    headers = [_order_clean_text(c.value).lower() for c in ws[1]]
    aliases = {
        'part_number': ['part number', 'part no', 'part_no', 'partno'],
        'quantity': ['quantity', 'qty', 'order qty', 'required qty'],
        'description': ['description', 'part name', 'part_name', 'part description'],
        'value': ['value', 'mav', 'mav value', 'unit value', 'price'],
    }
    mapping = {}
    for field, names in aliases.items():
        for idx, header in enumerate(headers):
            if header in names:
                mapping[field] = idx
                break
    missing = [name for name in ('part_number', 'quantity', 'description', 'value') if name not in mapping]
    if missing:
        raise HTTPException(status_code=400, detail='Required columns: Part Number, Quantity, Description, Value')

    rows = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        part_number = _order_clean_text(row[mapping['part_number']])
        description = _order_clean_text(row[mapping['description']])
        raw_value = row[mapping['value']]

        missing = []
        if not part_number:
            missing.append('Part Number')
        if raw_value is None or str(raw_value).strip() == '':
            missing.append('Value')
        if missing:
            field_text = ' and '.join(missing)
            verb = 'are' if len(missing) > 1 else 'is'
            raise HTTPException(status_code=400, detail=f'Row {row_number}: {field_text} {verb} required')

        try:
            quantity = float(row[mapping['quantity']] or 0)
            value = float(str(raw_value).replace(',', '').replace('₹', '').strip())
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f'Row {row_number}: Invalid quantity or value')
        if quantity <= 0:
            raise HTTPException(status_code=400, detail=f'Row {row_number}: Quantity must be greater than zero')
        rows.append({'part_number': part_number, 'quantity': quantity, 'description': description, 'value': value})
    if not rows:
        raise HTTPException(status_code=400, detail='No valid order rows found')
    return rows


def _deduplicate_order_rows(rows):
    """Remove exact duplicate input rows while preserving their first occurrence."""
    unique_rows = []
    seen = set()
    for row in rows or []:
        part_number = _order_clean_text(row.get('part_number')).upper()
        description = _order_clean_text(row.get('description'))
        try:
            quantity = float(row.get('quantity') or 0)
            value = float(row.get('value') or 0)
        except (TypeError, ValueError):
            quantity, value = 0.0, 0.0
        key = (part_number, quantity, description.casefold(), value)
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append({
            'part_number': _order_clean_text(row.get('part_number')),
            'quantity': quantity,
            'description': description,
            'value': value,
        })
    return unique_rows


async def _create_order_v2(rows, current_user: UserResponse, source: str, file_name: str = '', upload_fingerprint: str = '', scope: dict = None):
    rows = _deduplicate_order_rows(rows)
    if not rows:
        raise HTTPException(status_code=400, detail='No valid unique order rows found')
    ctx = await get_user_upload_context(current_user)
    scope = scope or {}
    role = (current_user.role or '').lower()
    if role == 'master':
        if scope.get('brand'): ctx['brand_name'] = _order_clean_text(scope.get('brand'))
        if scope.get('dealer'): ctx['dealer_name'] = _order_clean_text(scope.get('dealer'))
        if scope.get('branch'): ctx['branch'] = _order_clean_text(scope.get('branch'))
    elif role == 'admin' and scope.get('branch'):
        ctx['branch'] = _order_clean_text(scope.get('branch'))
    now = datetime.now(timezone.utc)

    # Browser double-clicks/retries must return the already-created order instead
    # of inserting a second header and a second set of items.
    if upload_fingerprint:
        existing = await db.order_headers.find_one({
            'created_by': current_user.id,
            'upload_fingerprint': upload_fingerprint,
        }, {'_id': 0})
        if existing:
            existing_items = await db.order_items.find(
                {'order_id': existing.get('id')}, {'_id': 0}
            ).to_list(10000)
            return {'order': existing, 'items': existing_items, 'duplicate': True}

    order_number = await _generate_order_number(ctx.get('brand_code'))
    order_id = str(uuid.uuid4())
    items = []
    total_value = 0.0
    total_qty = 0.0
    for row in rows:
        item_value = float(row.get('value', 0) or 0)
        qty = float(row.get('quantity', 0) or 0)
        total_value += item_value * qty
        total_qty += qty
        items.append({
            'id': str(uuid.uuid4()), 'order_id': order_id, 'order_number': order_number,
            'part_number': _order_clean_text(row.get('part_number')),
            'description': _order_clean_text(row.get('description')),
            'required_qty': qty, 'unit_value': item_value,
            'allocated_qty': 0.0, 'balance_qty': qty,
            'availability_status': 'Not Checked', 'allocations': [],
            'status': 'Order Created', 'created_at': now.isoformat(), 'updated_at': now.isoformat(),
        })
    order_doc = {
        'id': order_id, 'order_number': order_number, 'brand_name': ctx.get('brand_name', ''),
        'brand_code': ctx.get('brand_code', ''), 'dealer_name': ctx.get('dealer_name', ''),
        'branch': ctx.get('branch', ''), 'created_by': current_user.id,
        'created_user_name': current_user.username, 'source': source, 'file_name': file_name,
        'item_count': len(items), 'total_required_qty': total_qty, 'total_order_value': total_value,
        'status': 'Order Created', 'availability_checked': False,
        'upload_fingerprint': upload_fingerprint or None,
        'created_at': now.isoformat(), 'updated_at': now.isoformat(),
    }
    await db.order_headers.insert_one(order_doc.copy())
    await db.order_items.insert_many([dict(i) for i in items])
    await db.order_activity.insert_one({
        'id': str(uuid.uuid4()), 'order_id': order_id, 'order_number': order_number,
        'action': 'Order Created', 'performed_by': current_user.id,
        'performed_user_name': current_user.username, 'created_at': now.isoformat(),
    })
    order_doc.pop('_id', None)
    return {'order': order_doc, 'items': items}


@api_router.post('/order-desk/upload')
async def order_desk_upload(file: UploadFile = File(...), brand: str = Form(''), dealer: str = Form(''), branch: str = Form(''), current_user: UserResponse = Depends(get_current_user)):
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail='Only Excel files are allowed')
    content = await file.read()
    fingerprint_source = b'|'.join([
        current_user.id.encode('utf-8'),
        (file.filename or '').encode('utf-8'),
        content,
    ])
    upload_fingerprint = hashlib.sha256(fingerprint_source).hexdigest()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    rows = _parse_order_rows_from_worksheet(wb.active)
    return await _create_order_v2(
        rows, current_user, 'Excel Upload', file.filename, upload_fingerprint, {'brand': brand, 'dealer': dealer, 'branch': branch}
    )


@api_router.post('/order-desk/paste')
async def order_desk_paste(payload: dict, current_user: UserResponse = Depends(get_current_user)):
    rows = payload.get('rows') or []
    clean_rows = []
    for row_number, row in enumerate(rows, start=1):
        part_number = _order_clean_text(row.get('part_number'))
        description = _order_clean_text(row.get('description'))
        raw_value = row.get('value')

        missing = []
        if not part_number:
            missing.append('Part Number')
        if raw_value is None or str(raw_value).strip() == '':
            missing.append('Value')
        if missing:
            field_text = ' and '.join(missing)
            verb = 'are' if len(missing) > 1 else 'is'
            raise HTTPException(status_code=400, detail=f'Row {row_number}: {field_text} {verb} required')

        try:
            qty = float(row.get('quantity') or 0)
            value = float(str(raw_value).replace(',', '').replace('₹', '').strip())
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f'Row {row_number}: Invalid quantity or value')
        if qty <= 0:
            raise HTTPException(status_code=400, detail=f'Row {row_number}: Quantity must be greater than zero')
        clean_rows.append({'part_number': part_number, 'quantity': qty, 'description': description, 'value': value})
    if not clean_rows:
        raise HTTPException(status_code=400, detail='No valid rows found')
    return await _create_order_v2(clean_rows, current_user, 'Copy From Excel', scope={'brand': payload.get('brand'), 'dealer': payload.get('dealer'), 'branch': payload.get('branch')})


@api_router.get('/order-desk/orders')
async def order_desk_orders(brand: Optional[str] = None, dealer: Optional[str] = None, branch: Optional[str] = None, current_user: UserResponse = Depends(get_current_user)):
    role = (current_user.role or '').lower()
    query = {} if role == 'master' else ({'dealer_name': current_user.group} if role == 'admin' else {'created_by': current_user.id})
    if brand and not str(brand).startswith('All '): query['brand_name'] = brand
    if dealer and not str(dealer).startswith('All '): query['dealer_name'] = dealer
    if branch and not str(branch).startswith('All '): query['branch'] = branch
    if role == 'admin': query['dealer_name'] = current_user.group
    if role == 'user': query.update({'created_by': current_user.id, 'branch': current_user.location})
    rows = await db.order_headers.find(query, {'_id': 0}).sort('created_at', -1).limit(1000).to_list(1000)
    return rows


@api_router.get('/order-desk/orders/{order_id}')
async def order_desk_order_detail(
    order_id: str,
    branch_aging_type: Optional[str] = 'purchase',
    branch_min_aging: float = 0,
    dealer_aging_type: Optional[str] = 'purchase',
    dealer_min_aging: float = 0,
    current_user: UserResponse = Depends(get_current_user),
):
    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail='Order not found')
    role = (current_user.role or '').lower()
    if role != 'master' and order.get('created_by') != current_user.id and order.get('dealer_name') != current_user.group:
        raise HTTPException(status_code=403, detail='Not authorized')
    items = await db.order_items.find({'order_id': order_id}, {'_id': 0}).to_list(10000)
    items = await odw.enrich_order_items(
        db, order, items,
        branch_aging_type=(branch_aging_type or 'purchase'),
        branch_min_aging=float(branch_min_aging or 0),
        dealer_aging_type=(dealer_aging_type or 'purchase'),
        dealer_min_aging=float(dealer_min_aging or 0),
    )
    order_stage = odw.compute_order_stage(items)
    return {
        'order': {**order, **order_stage},
        'items': items,
        'stage': order_stage,
    }


@api_router.get('/order-desk/template')
async def order_desk_template(current_user: UserResponse = Depends(get_current_user)):
    """Generic Order Desk Excel template (Part Number, Quantity, Description, Value)."""
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Order'
    ws.append(['Part Number', 'Quantity', 'Description', 'Value'])
    ws.append(['86511B4000', 2, 'FRONT BUMPER', 1500])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=Order_Desk_Template.xlsx'},
    )


@api_router.get('/order-desk/orders/{order_id}/export')
async def order_desk_order_export(order_id: str, current_user: UserResponse = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    excel_permissions.require_excel_export(current_user)

    detail = await order_desk_order_detail(order_id, current_user=current_user)
    order = detail['order']
    items = detail['items']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Order Desk'
    headers = [
        'Part Number', 'Part Name', 'Requested Qty', 'Available Qty', 'Allocated Qty', 'Balance Qty',
        'Value', 'LOC', 'Purchase Aging', 'Sales Aging', 'Availability Status', 'Source Branch', 'Source Dealer',
    ]
    ws.append(headers)
    for item in items:
        alloc_qty = sum(float(s.get('allocated_qty') or 0) for s in (item.get('selected_sources') or []))
        req = float(item.get('required_qty') or item.get('quantity') or 0)
        avail = float(item.get('available_qty') or 0)
        sources = item.get('selected_sources') or item.get('same_dealer_sources') or []
        src_branch = sources[0].get('branch') if sources else ''
        src_dealer = sources[0].get('dealer_name') if sources else ''
        ws.append([
            item.get('part_number'),
            item.get('description') or item.get('part_name'),
            req,
            avail,
            alloc_qty,
            max(req - alloc_qty, 0),
            item.get('value') or item.get('line_value'),
            item.get('loc') or item.get('location'),
            item.get('purchase_aging_days'),
            item.get('sales_aging_days'),
            item.get('availability_status'),
            src_branch,
            src_dealer,
        ])
    meta = wb.create_sheet('Summary')
    meta.append(['Order Number', order.get('order_number')])
    meta.append(['Brand', order.get('brand_name')])
    meta.append(['Dealer', order.get('dealer_name')])
    meta.append(['Branch', order.get('branch')])
    meta.append(['Status', order.get('status')])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Order_Desk_{order.get('order_number') or order_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fname}'},
    )


@api_router.post('/order-desk/orders/{order_id}/check-availability')
async def order_desk_check_availability(order_id: str, brand: Optional[str] = None, dealer: Optional[str] = None, branch: Optional[str] = None, current_user: UserResponse = Depends(get_current_user)):
    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail='Order not found')
    if not brand or str(brand).startswith('All ') or not dealer or str(dealer).startswith('All ') or not branch or str(branch).startswith('All '):
        raise HTTPException(status_code=400, detail='Select a valid Brand, Dealer and Branch before Check Availability')
    role = (current_user.role or '').lower()
    if role != 'master' and dealer != current_user.group:
        raise HTTPException(status_code=403, detail='Selected dealer is outside your permitted scope')
    if role == 'user' and branch != current_user.location:
        raise HTTPException(status_code=403, detail='Selected branch is outside your permitted scope')
    if order.get('brand_name') != brand or order.get('dealer_name') != dealer or order.get('branch') != branch:
        raise HTTPException(status_code=400, detail='Selected scope does not match this order')
    items = await db.order_items.find({'order_id': order_id}, {'_id': 0}).to_list(10000)
    checked_at = datetime.now(timezone.utc).isoformat()

    # Pass 1: build raw availability entries per item (unchanged from before).
    per_item_sources = []
    all_stock_ids = []
    for item in items:
        part_no = item.get('part_number')
        product_query = {
            'part_number': {'$regex': f'^{re.escape(part_no)}$', '$options': 'i'},
            'brand_name': order.get('brand_name'),
            'publish_status': 'Published',
            'is_active_today': True,
            # Order Desk must use only stock from today's published batches.
            # Older rows can still have is_active_today=True in legacy data,
            # so active_date_key is the authoritative day boundary.
            'active_date_key': _nmts_date_key(),
            'available_qty_number': {'$gt': 0},
        }
        stocks = await db.products.find(product_query, {'_id': 0}).to_list(5000)
        same_dealer, other_dealers = [], []
        for stock in stocks:
            purchase_aging = _order_stock_purchase_aging_days(stock)
            sales_aging = _order_stock_sales_aging_days(stock)
            entry = {
                'stock_id': stock.get('id'), 'dealer_name': stock.get('dealer_name', ''),
                'branch': stock.get('branch', ''),
                'available_qty': float(stock.get('available_qty_number', stock.get('quantity', 0)) or 0),
                'purchase_aging_days': purchase_aging,
                'sales_aging_days': sales_aging,
                'aging_days': purchase_aging,  # backward compatible: defaults to Purchase Aging
                'unit_value': float(stock.get('unit_value_number', stock.get('mav_value', 0)) or 0),
                'loc': stock.get('location', ''),
                'last_upload': stock.get('created_at', ''),
            }
            all_stock_ids.append(entry['stock_id'])
            if _order_clean_text(stock.get('dealer_name')).lower() == _order_clean_text(order.get('dealer_name')).lower():
                same_dealer.append(entry)
            else:
                other_dealers.append(entry)
        per_item_sources.append((item, same_dealer, other_dealers))

    # Pass 2: one reservation lookup for the whole order, then annotate every
    # source with reserved_qty / net_available_qty so already-reserved stock
    # is visibly excluded from what Auto Suggest (and a careful human) would
    # allocate again, without changing the original available_qty field.
    reserved_map = await _reservation_qty_map(all_stock_ids)

    response_items = []
    for item, same_dealer, other_dealers in per_item_sources:
        same_dealer = [_with_reservation_adjustment(e, reserved_map) for e in same_dealer]
        other_dealers = [_with_reservation_adjustment(e, reserved_map) for e in other_dealers]
        status = 'Available' if same_dealer or other_dealers else 'Not Available'
        await db.order_items.update_one({'id': item['id']}, {'$set': {
            'availability_status': status, 'same_dealer_sources': same_dealer,
            'other_dealer_sources': other_dealers, 'availability_checked_at': checked_at,
            'updated_at': checked_at,
        }})
        response_items.append({**item, 'availability_status': status, 'same_dealer_sources': same_dealer, 'other_dealer_sources': other_dealers})
    await db.order_headers.update_one({'id': order_id}, {'$set': {'availability_checked': True, 'status': 'Availability Checked', 'updated_at': checked_at}})
    await db.order_activity.insert_one({'id': str(uuid.uuid4()), 'order_id': order_id, 'order_number': order.get('order_number'), 'action': 'Availability Checked', 'performed_by': current_user.id, 'performed_user_name': current_user.username, 'created_at': checked_at})
    return {'order_number': order.get('order_number'), 'items': response_items}


@api_router.post('/order-desk/orders/{order_id}/allocate')
async def order_desk_allocate(order_id: str, payload: dict, current_user: UserResponse = Depends(get_current_user)):
    allocations = payload.get('allocations') or []
    now = datetime.now(timezone.utc).isoformat()
    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail='Order not found')
    for entry in allocations:
        item_id = entry.get('item_id')
        selected = entry.get('sources') or []
        item = await db.order_items.find_one({'id': item_id, 'order_id': order_id}, {'_id': 0})
        if not item:
            continue
        # Preserve already-sent allocations; only replace draft/unsent selections.
        preserved = [
            a for a in (item.get('allocations') or [])
            if a.get('request_no') or a.get('request_number')
            or str(a.get('status') or '').lower() in (
                'request sent', 'requested', 'accepted', 'partially accepted',
                'rejected', 'cancelled', 'completed', 'dispatched', 'received',
            )
        ]
        preserved_keys = {
            (str(a.get('dealer_name') or '').lower(), str(a.get('branch') or '').lower())
            for a in preserved
        }
        total_preserved = sum(max(0, float(a.get('request_qty') or 0)) for a in preserved)
        cleaned = list(preserved)
        for source in selected:
            qty = max(0, float(source.get('request_qty') or 0))
            if qty <= 0:
                continue
            key = (str(source.get('dealer_name') or '').lower(), str(source.get('branch') or '').lower())
            if key in preserved_keys:
                continue  # do not overwrite a source that already has a sent request
            level = odw.allocation_level({**source, 'level': source.get('level') or source.get('source_type')}, order)
            cleaned.append({
                'dealer_name': _order_clean_text(source.get('dealer_name')),
                'branch': _order_clean_text(source.get('branch')),
                'request_qty': qty, 'available_qty_at_selection': float(source.get('available_qty') or 0),
                'purchase_aging_days': source.get('purchase_aging_days', source.get('aging_days', '')),
                'sales_aging_days': source.get('sales_aging_days', ''),
                'aging_days': source.get('aging_days', ''), 'loc': source.get('loc', ''),
                'stock_id': source.get('stock_id'),
                'status': odw.REQUEST_STATUS_READY, 'selected_at': now, 'origin': 'manual',
                'level': level, 'source_type': level,
                'source_dealer': _order_clean_text(source.get('dealer_name')),
                'source_branch': _order_clean_text(source.get('branch')),
                'requested_qty': qty, 'accepted_qty': 0, 'remaining_qty': qty,
                'request_status': odw.REQUEST_STATUS_READY,
            })
        total = sum(max(0, float(s.get('request_qty') or 0)) for s in cleaned)
        required = float(item.get('required_qty', 0) or 0)
        # Cap against already-accepted + active remaining requirement
        existing_reqs = await db.order_requests.find({'order_item_id': item_id}, {'_id': 0}).to_list(1000)
        wf = odw.compute_item_workflow(item, order, existing_reqs)
        # Draft allocations may only cover remaining (+ currently unsent replaced drafts)
        if total - total_preserved > wf['remaining_qty'] + 1e-9 and total > required:
            raise HTTPException(status_code=400, detail=f'Request quantity exceeds required quantity for {item.get("part_number")}')
        if total > required:
            raise HTTPException(status_code=400, detail=f'Request quantity exceeds required quantity for {item.get("part_number")}')
        # A manual save (even clearing the row back to empty) always wins over
        # Auto Suggest — once allocation_source is 'manual', Auto Suggest will
        # skip this item entirely until the user clears the selection, at
        # which point Auto Suggest is free to run for it again.
        await db.order_items.update_one({'id': item_id}, {'$set': {
            'allocations': cleaned, 'allocated_qty': total, 'balance_qty': max(0, required - total),
            'remaining_qty': max(0, required - float(wf.get('accepted_qty') or 0) - sum(
                float(a.get('request_qty') or 0) for a in cleaned
                if a.get('request_no') or a.get('request_number') or str(a.get('status') or '').lower() in (
                    'request sent', 'requested', 'accepted', 'partially accepted', 'completed', 'dispatched', 'received',
                )
            )),
            'status': 'Source Selected' if cleaned else 'Availability Checked', 'updated_at': now,
            'allocation_source': 'manual' if any(not (a.get('request_no') or a.get('request_number')) for a in cleaned) else item.get('allocation_source'),
            'request_status': odw.REQUEST_STATUS_READY if any(not (a.get('request_no') or a.get('request_number')) for a in cleaned) else wf.get('request_status'),
        }})
    await db.order_headers.update_one({'id': order_id}, {'$set': {'status': 'Source Selected', 'updated_at': now}})
    await odw.append_order_audit(db, order, 'Manual source selected / Qty changed', current_user)
    return {'message': 'Source selections saved'}


@api_router.post('/order-desk/orders/{order_id}/auto-suggest')
async def order_desk_auto_suggest(order_id: str, payload: dict, current_user: UserResponse = Depends(get_current_user)):
    """Smart Auto Suggest Allocation Engine.

    level='branch': suggests only from same_dealer_sources (Branch
    Availability), sorted by the chosen aging type (Purchase/Sales), highest
    aging first, ties broken by highest available qty, then oldest uploaded
    stock, then alphabetical branch. Allocates from a single branch if it can
    cover the full requirement, otherwise splits across branches — never
    exceeding the requested quantity.

    level='dealer': same ranking, but pulls only from other_dealer_sources,
    is refused until at least one request has been sent for this order, and
    only ever targets the Pending Qty = Requested Qty - Already Requested
    Branch Qty (computed live from non-Rejected/Cancelled order_requests).

    Both levels are reservation-aware (already-reserved stock is excluded)
    and both skip any item the user has manually overridden.
    """
    level = str((payload or {}).get('level') or '').strip().lower()
    aging_type = str((payload or {}).get('aging_type') or 'purchase').strip().lower()
    try:
        min_aging_days = float((payload or {}).get('min_aging_days') or (payload or {}).get('aging_min_days') or 0)
    except (TypeError, ValueError):
        min_aging_days = 0.0
    if level not in ('branch', 'dealer'):
        raise HTTPException(status_code=400, detail="level must be 'branch' or 'dealer'")
    if aging_type not in ('purchase', 'sales'):
        aging_type = 'purchase'

    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail='Order not found')
    role = (current_user.role or '').lower()
    if role != 'master' and order.get('created_by') != current_user.id and order.get('dealer_name') != current_user.group:
        raise HTTPException(status_code=403, detail='Not authorized for this order')

    items = await db.order_items.find({'order_id': order_id}, {'_id': 0}).to_list(10000)
    if not items:
        raise HTTPException(status_code=400, detail='Order has no items')
    if not any(item.get('availability_checked_at') for item in items):
        raise HTTPException(status_code=400, detail='Run Check Availability before using Auto Suggest')

    # Stage gate: Dealer Auto Suggest only after Branch stage exhausted for remaining items
    if level == 'dealer':
        freezes_gate = await odw.load_today_freezes(db, [i.get('part_number') for i in items], order.get('brand_name') or '')
        still_branch = False
        for item in items:
            rem = float(item.get('remaining_qty') if item.get('remaining_qty') is not None else item.get('required_qty') or 0)
            # subtract accepted/pending roughly via enrichment later; use flag + pool
            if item.get('branch_stage_exhausted'):
                continue
            pool = odw.eligible_pool(item, order, 'branch', freezes_gate, 'purchase', 0)
            if pool and rem > 0 and not item.get('branch_stage_exhausted'):
                # If any active branch request pending, also block
                still_branch = True
                break
        sent_already = await db.request_headers.find_one({'order_id': order_id}, {'_id': 0, 'id': 1})
        if not sent_already and still_branch:
            raise HTTPException(status_code=400, detail='Dealer Auto Suggest opens only after Branch stage is exhausted for remaining quantity')
        if still_branch:
            raise HTTPException(status_code=400, detail='Complete or exhaust Branch stage before Dealer Auto Suggest')

    selected_item_ids = {str(x) for x in ((payload or {}).get('item_ids') or []) if x}
    pool_field = 'same_dealer_sources' if level == 'branch' else 'other_dealer_sources'

    all_stock_ids = [s.get('stock_id') for item in items for s in (item.get(pool_field) or [])]
    reserved_map = await _reservation_qty_map(all_stock_ids)
    freezes = await odw.load_today_freezes(db, [i.get('part_number') for i in items], order.get('brand_name') or '')

    now = datetime.now(timezone.utc).isoformat()
    result_items = []
    for item in items:
        if selected_item_ids and item.get('id') not in selected_item_ids:
            result_items.append(item)
            continue
        # Locked items with no remaining qty: skip
        active_pending = await db.order_requests.find_one(
            {'order_item_id': item.get('id'), 'status': 'Requested'}, {'_id': 0, 'id': 1},
        )
        if active_pending:
            result_items.append({**item, 'auto_suggest_skipped': 'awaiting_response_lock'})
            continue
        if item.get('allocation_source') == 'manual' and not item.get('retry_required'):
            # Allow auto-suggest when remaining exists after partial/reject even if prior manual
            existing_reqs = await db.order_requests.find({'order_item_id': item.get('id')}, {'_id': 0}).to_list(100)
            if not existing_reqs:
                result_items.append({**item, 'auto_suggest_skipped': 'manual_override'})
                continue

        required_qty = float(item.get('required_qty') or 0)
        active_requests = await db.order_requests.find(
            {'order_item_id': item.get('id'), 'status': {'$nin': ['Rejected', 'Cancelled']}},
            {'_id': 0, 'requested_qty': 1, 'accepted_qty': 1, 'approved_qty': 1, 'status': 1},
        ).to_list(1000)
        already_requested_qty = sum(
            float((r.get('accepted_qty') if r.get('status') == 'Approved' else r.get('requested_qty')) or 0)
            for r in active_requests
        )
        pending_qty = max(0.0, required_qty - already_requested_qty)
        if pending_qty <= 0:
            result_items.append({**item, 'auto_suggest_skipped': 'fully_locked'})
            continue

        if level == 'dealer' and not item.get('branch_stage_exhausted'):
            branch_left = odw.eligible_pool(item, order, 'branch', freezes, aging_type, min_aging_days)
            if branch_left:
                result_items.append({**item, 'auto_suggest_skipped': 'branch_stage_open'})
                continue

        pool = odw.eligible_pool(item, order, level, freezes, aging_type, min_aging_days)
        pool = [_with_reservation_adjustment(s, reserved_map) for s in pool]
        pool = [s for s in pool if s.get('net_available_qty', 0) > 0]
        pool.sort(key=lambda s: _auto_suggest_sort_key(s, aging_type))

        picked = []
        left = pending_qty
        for source in pool:
            if left <= 0:
                break
            take = min(left, float(source.get('net_available_qty') or 0))
            if take <= 0:
                continue
            picked.append({
                'dealer_name': _order_clean_text(source.get('dealer_name')),
                'branch': _order_clean_text(source.get('branch')),
                'request_qty': take, 'available_qty_at_selection': float(source.get('available_qty') or 0),
                'purchase_aging_days': source.get('purchase_aging_days'),
                'sales_aging_days': source.get('sales_aging_days'),
                'aging_days': source.get('aging_days'), 'loc': source.get('loc', ''),
                'stock_id': source.get('stock_id'),
                'status': odw.REQUEST_STATUS_READY, 'selected_at': now, 'level': level, 'origin': 'auto',
                'source_type': level, 'source_dealer': _order_clean_text(source.get('dealer_name')),
                'source_branch': _order_clean_text(source.get('branch')),
                'requested_qty': take, 'accepted_qty': 0, 'remaining_qty': take,
                'request_status': odw.REQUEST_STATUS_READY,
            })
            left -= take
            reserved_map[source.get('stock_id')] = reserved_map.get(source.get('stock_id'), 0) + take

        preserved = [
            a for a in (item.get('allocations') or [])
            if a.get('request_no') or a.get('request_number')
            or str(a.get('status') or '').lower() in (
                'request sent', 'awaiting response', 'requested', 'accepted', 'partially accepted',
                'rejected', 'cancelled', 'completed', 'dispatched', 'received', 'response time expired',
            )
        ]
        if level == 'branch':
            new_allocations = preserved + picked
        else:
            existing = [
                a for a in preserved
                if not (a.get('level') == 'dealer' and a.get('origin') == 'auto' and not (a.get('request_no') or a.get('request_number')))
            ]
            # drop unsent dealer drafts then add new
            existing = [a for a in (item.get('allocations') or []) if a.get('request_no') or a.get('request_number') or a.get('level') != 'dealer']
            new_allocations = existing + picked

        total_allocated = sum(float(a.get('request_qty') or 0) for a in new_allocations)
        await db.order_items.update_one({'id': item['id']}, {'$set': {
            'allocations': new_allocations, 'allocated_qty': total_allocated,
            'balance_qty': max(0.0, required_qty - total_allocated),
            'remaining_qty': pending_qty - sum(p['request_qty'] for p in picked),
            'status': 'Source Selected' if new_allocations else item.get('status', 'Availability Checked'),
            'allocation_source': 'auto', 'retry_required': False if picked else item.get('retry_required', False),
            'request_status': odw.REQUEST_STATUS_READY if picked else item.get('request_status'),
            'updated_at': now,
        }})
        updated_item = await db.order_items.find_one({'id': item['id']}, {'_id': 0})
        result_items.append({
            **updated_item,
            'auto_suggest_new_qty': sum(p['request_qty'] for p in picked),
            'auto_suggest_unfulfilled': max(0.0, left),
        })

    await db.order_headers.update_one({'id': order_id}, {'$set': {
        'updated_at': now,
        f'{level}_aging_type': aging_type,
        f'{level}_min_aging_days': min_aging_days,
    }})
    await db.order_activity.insert_one({
        'id': str(uuid.uuid4()), 'order_id': order_id, 'order_number': order.get('order_number'),
        'action': f'Auto Suggest ({level.title()}, {aging_type.title()} Aging ≥{int(min_aging_days)}d)',
        'performed_by': current_user.id,
        'performed_user_name': current_user.username, 'created_at': now,
    })
    enriched = await odw.enrich_order_items(db, order, result_items, aging_type, min_aging_days, aging_type, min_aging_days)
    return {'order_number': order.get('order_number'), 'level': level, 'aging_type': aging_type,
            'min_aging_days': min_aging_days, 'items': enriched}


async def _active_recipients_for_scope(brand: str = None, dealer: str = None, branch: str = None, roles=('admin', 'master')):
    """Active users (role admin/master) matching the given Brand/Dealer/Branch
    scope — used to notify the supplying side of a request. Master users are
    always included for the given brand since they can act on any dealer."""
    query = {'status': {'$regex': '^active$', '$options': 'i'}, 'role': {'$in': list(roles)}}
    scope_or = []
    if dealer and branch:
        scope_or.append({'group': dealer, 'location': branch})
    if brand:
        scope_or.append({'brand': brand, 'role': 'master'})
    if scope_or:
        query['$or'] = scope_or
    users = await db.users.find(query, {'_id': 0, 'email': 1, 'mobile': 1, 'phone': 1, 'name': 1, 'username': 1}).to_list(200)
    seen, recipients = set(), []
    for u in users:
        key = (u.get('email') or '') + '|' + (u.get('mobile') or u.get('phone') or '')
        if key in seen or key == '|':
            continue
        seen.add(key)
        recipients.append({'email': u.get('email'), 'mobile': u.get('mobile') or u.get('phone'), 'name': u.get('name') or u.get('username')})
    return recipients


async def _create_request_group(order: dict, pairs: list, current_user: UserResponse, now: str):
    """Create one Parts Transfer Request (one request_number, one PDF, one
    email) for a single Requested-To destination (supplying dealer+branch).
    `pairs` is the list of (order_item, allocation_source) tuples going to
    that destination. Existing per-item order_requests rows are still
    created — one per (item, source) — so Approve/Reject/Cancel and the
    Request Center list continue to work exactly as before; they simply now
    share a common request_number and request_group_id."""
    first_item, first_source = pairs[0]
    supplying_brand = order.get('brand_name')
    supplying_dealer = _order_clean_text(first_source.get('dealer_name'))
    supplying_branch = _order_clean_text(first_source.get('branch'))

    brand_code = await resolve_brand_code_for_upload(order.get('brand_name'))
    numbering = await _generate_request_number(brand_code)
    request_number = numbering['request_number']
    group_id = str(uuid.uuid4())

    line_items, order_request_ids = [], []
    total_qty, total_value = 0.0, 0.0
    for item, source in pairs:
        qty = float(source.get('request_qty') or 0)
        unit_value = float(item.get('unit_value', 0) or 0)
        line_value = qty * unit_value
        total_qty += qty
        total_value += line_value
        order_request_id = str(uuid.uuid4())
        order_request_ids.append(order_request_id)
        request_doc = {
            'id': order_request_id, 'order_id': order.get('id'), 'order_number': order.get('order_number'),
            'order_item_id': item.get('id'), 'part_number': item.get('part_number'),
            'description': item.get('description'), 'requesting_brand': order.get('brand_name'),
            'requesting_dealer': order.get('dealer_name'), 'requesting_branch': order.get('branch'),
            'supplying_brand': supplying_brand, 'supplying_dealer': supplying_dealer, 'supplying_branch': supplying_branch,
            'requested_qty': qty, 'available_qty_at_request': source.get('available_qty_at_selection'),
            'unit_value_at_request': unit_value, 'value_at_request': line_value,
            'aging_days_at_request': source.get('aging_days'),
            'purchase_aging_days_at_request': source.get('purchase_aging_days', source.get('purchase_aging', source.get('aging_days'))),
            'sales_aging_days_at_request': source.get('sales_aging_days', source.get('sales_aging')),
            'loc_at_request': source.get('loc', ''),
            'status': 'Requested', 'remarks': '', 'approval_remarks': '',
            'requested_by': current_user.id, 'requested_user_name': current_user.username, 'requested_at': now,
            'requester_email': current_user.email, 'requester_mobile': getattr(current_user, 'phone', ''),
            'request_number': request_number, 'request_group_id': group_id,
        }
        await db.order_requests.insert_one(dict(request_doc))
        line_items.append({
            'order_request_id': order_request_id, 'part_number': item.get('part_number'),
            'description': item.get('description'), 'requested_qty': qty,
            'available_qty_at_request': source.get('available_qty_at_selection'),
            'unit_value': unit_value, 'value': line_value,
            'purchase_aging_days': source.get('purchase_aging_days', source.get('purchase_aging', source.get('aging_days'))),
            'sales_aging_days': source.get('sales_aging_days', source.get('sales_aging')),
            'loc': source.get('loc', ''),
        })

    group_doc = {
        'id': group_id, 'request_number': request_number, 'request_number_date': numbering['date_key'],
        'request_sequence': numbering['sequence'], 'brand_code': numbering['brand_code'],
        'order_id': order.get('id'), 'order_number': order.get('order_number'),
        'requesting_brand': order.get('brand_name'), 'requesting_dealer': order.get('dealer_name'),
        'requesting_branch': order.get('branch'),
        'supplying_brand': supplying_brand, 'supplying_dealer': supplying_dealer, 'supplying_branch': supplying_branch,
        'requested_by': current_user.id, 'requested_user_name': current_user.username,
        'requester_email': current_user.email, 'requester_mobile': getattr(current_user, 'phone', ''),
        'order_request_ids': order_request_ids, 'items': line_items,
        'total_items': len(line_items), 'total_qty': total_qty, 'total_value': total_value,
        'status': 'Requested', 'created_at': now, 'updated_at': now, 'retry_count': 0,
        'email_sent': False, 'email_status': 'pending', 'email_sent_at': None, 'email_error': None,
        'receiver_email': '', 'pdf_filename': f'{request_number}.pdf', 'notification_log_id': None,
    }
    # Response SLA timer — based on THIS request group's line-item count only.
    schedule = odw.compute_response_schedule(len(line_items), datetime.now(timezone.utc))
    group_doc.update(schedule)

    try:
        await db.request_headers.insert_one(dict(group_doc))
    except DuplicateKeyError:
        # Concurrent duplicate ACTIVE request to the same destination
        # (partial unique index on status=Requested). Reuse the winner.
        # Terminal history for the same destination is allowed and does not
        # hit this path.
        existing = await db.request_headers.find_one(
            {
                'order_id': order.get('id'),
                'supplying_dealer': supplying_dealer,
                'supplying_branch': supplying_branch,
                'status': 'Requested',
            },
            {'_id': 0},
        )
        await db.order_requests.delete_many({'id': {'$in': order_request_ids}})
        if existing:
            return existing
        # Fallback: any matching destination header (should be rare).
        return await db.request_headers.find_one(
            {'order_id': order.get('id'), 'supplying_dealer': supplying_dealer, 'supplying_branch': supplying_branch},
            {'_id': 0},
        )

    # Reservation-aware allocation: only once the request group is durably
    # saved (never on the DuplicateKeyError/retry path above) do we reserve
    # the underlying stock rows, so the same physical stock can never be
    # suggested or double-allocated again — by this order or any other —
    # until it is released by a Reject/Cancel on this request.
    reservation_docs = []
    for (item, source), order_request_id in zip(pairs, order_request_ids):
        stock_id = source.get('stock_id')
        if not stock_id:
            continue
        reservation_docs.append({
            'id': str(uuid.uuid4()), 'order_id': order.get('id'), 'order_item_id': item.get('id'),
            'order_request_id': order_request_id, 'request_group_id': group_id,
            'part_number': item.get('part_number'), 'brand_name': order.get('brand_name'),
            'supplying_dealer': supplying_dealer, 'supplying_branch': supplying_branch,
            'stock_id': stock_id, 'qty': float(source.get('request_qty') or 0),
            'status': 'active', 'created_at': now, 'released_at': None,
        })
    if reservation_docs:
        await db.stock_reservations.insert_many(reservation_docs)

    # Email is best-effort only: the request/request number/items above are
    # already saved and must never be rolled back by a PDF or SMTP failure.
    await _send_request_group_email(group_doc)
    return await db.request_headers.find_one({'id': group_id}, {'_id': 0})


@api_router.post('/order-desk/orders/{order_id}/send-requests')
async def order_desk_send_requests_v2(order_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    """Send requests for selected allocations.

    Body (optional):
      level: 'branch' | 'dealer'  — send only that stage's unsent allocations.
      If omitted, defaults to sending all unsent allocations (legacy behaviour),
      but the UI always passes an explicit level so Branch and Dealer stay separate.
    Auto Suggest never calls this endpoint.
    """
    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail='Order not found')

    role = (current_user.role or '').lower()
    if role != 'master' and order.get('created_by') != current_user.id and order.get('dealer_name') != current_user.group:
        raise HTTPException(status_code=403, detail='Not authorized to send requests for this order')

    level = str(((payload or {}).get('level') or '')).strip().lower()
    if level and level not in ('branch', 'dealer'):
        raise HTTPException(status_code=400, detail="level must be 'branch' or 'dealer'")

    items = await db.order_items.find({'order_id': order_id}, {'_id': 0}).to_list(10000)
    now = datetime.now(timezone.utc).isoformat()

    # Build the candidate set of unsent allocations (optionally filtered by level)
    # BEFORE fingerprinting so re-sends of already-sent content stay idempotent
    # while NEW remaining-qty allocations can still create fresh requests.
    all_existing_requests = await db.order_requests.find({'order_id': order_id}, {'_id': 0}).to_list(20000)
    reqs_by_item = {}
    for r in all_existing_requests:
        reqs_by_item.setdefault(r.get('order_item_id'), []).append(r)

    groups = {}
    for item in items:
        item_reqs = reqs_by_item.get(item.get('id'), [])
        if level:
            sources = odw.filter_unsent_allocations(item, order, level, item_reqs)
        else:
            sources = []
            for lvl in ('branch', 'dealer'):
                sources.extend(odw.filter_unsent_allocations(item, order, lvl, item_reqs))
        for source in sources:
            key = (_order_clean_text(source.get('dealer_name')), _order_clean_text(source.get('branch')))
            groups.setdefault(key, []).append((item, source))

    if not groups:
        raise HTTPException(
            status_code=400,
            detail=f"No unsent {'branch' if level == 'branch' else 'dealer' if level == 'dealer' else ''} source selections to send".strip()
            or 'Select at least one source before sending requests',
        )

    # Content fingerprint of THIS send batch (not the whole order) for duplicate protection.
    fingerprint_parts = [order_id, level or 'all']
    for key, pairs in sorted(groups.items(), key=lambda kv: kv[0]):
        for item, source in pairs:
            fingerprint_parts.append('|'.join([
                item.get('id', ''), key[0], key[1], str(source.get('request_qty', '')),
            ]))
    fingerprint = hashlib.sha256('::'.join(fingerprint_parts).encode('utf-8')).hexdigest()
    prior_fps = set(order.get('send_requests_fingerprints') or [])
    if order.get('send_requests_fingerprint') == fingerprint or fingerprint in prior_fps:
        # Return existing groups that match these destinations if present
        existing_groups = await db.request_headers.find({'order_id': order_id}, {'_id': 0}).sort('created_at', 1).to_list(1000)
        matched = [
            g for g in existing_groups
            if (_order_clean_text(g.get('supplying_dealer')), _order_clean_text(g.get('supplying_branch'))) in groups
        ]
        if matched:
            return _send_requests_response(matched, duplicate=True)

    created_group_docs = []
    for _key, pairs in groups.items():
        group_doc = await _create_request_group(order, pairs, current_user, now)
        created_group_docs.append(group_doc)
        request_number = (group_doc or {}).get('request_number')
        touched_item_ids = set()
        for item, source in pairs:
            touched_item_ids.add(item.get('id'))
            # Stamp allocation rows with request number / Request Sent status
            fresh = await db.order_items.find_one({'id': item.get('id')}, {'_id': 0})
            if not fresh:
                continue
            stamped = odw.mark_allocations_sent(fresh, [source], request_number, now)
            await db.order_items.update_one({'id': item.get('id')}, {'$set': {
                'allocations': stamped,
                'status': 'Requested',
                'request_status': odw.REQUEST_STATUS_SENT,
                'retry_required': False,
                're_enquire': False,
                'updated_at': now,
            }})
        for item_id in touched_item_ids:
            # Store source_type on the underlying order_requests for this batch
            await db.order_requests.update_many(
                {'order_item_id': item_id, 'request_number': request_number},
                {'$set': {'source_type': level or odw.allocation_level(
                    {'dealer_name': _key[0], 'branch': _key[1]}, order
                ), 'level': level or odw.allocation_level(
                    {'dealer_name': _key[0], 'branch': _key[1]}, order
                )}},
            )

    prior_fps.add(fingerprint)
    await db.order_headers.update_one({'id': order_id}, {'$set': {
        'status': 'Requested', 'updated_at': now,
        'send_requests_fingerprint': fingerprint,
        'send_requests_fingerprints': list(prior_fps)[-50:],
        'last_send_level': level or 'all',
    }})
    action = f"{'Branch' if level == 'branch' else 'Dealer' if level == 'dealer' else ''} Request sent".strip()
    await db.order_activity.insert_one({
        'id': str(uuid.uuid4()), 'order_id': order_id, 'order_number': order.get('order_number'),
        'action': action or 'Requests Sent', 'performed_by': current_user.id, 'performed_user_name': current_user.username,
        'created_at': now, 'details': {'level': level or 'all', 'request_numbers': [g.get('request_number') for g in created_group_docs if g]},
    })

    return _send_requests_response(created_group_docs, duplicate=False)


@api_router.post('/order-desk/orders/{order_id}/add-items')
async def order_desk_add_items(order_id: str, payload: dict, current_user: UserResponse = Depends(get_current_user)):
    """Add more parts under an existing Order Number. Never creates a new order number."""
    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail='Order not found')
    role = (current_user.role or '').lower()
    if role != 'master' and order.get('created_by') != current_user.id and order.get('dealer_name') != current_user.group:
        raise HTTPException(status_code=403, detail='Not authorized')
    rows = payload.get('rows') or payload.get('items') or []
    if not rows:
        raise HTTPException(status_code=400, detail='No items to add')
    now = datetime.now(timezone.utc).isoformat()
    original_created_at = order.get('created_at')
    new_items = []
    add_qty = 0.0
    add_value = 0.0
    for row_number, row in enumerate(rows, start=1):
        part_number = _order_clean_text(row.get('part_number'))
        description = _order_clean_text(row.get('description') or row.get('part_name'))
        raw_value = row.get('value', row.get('unit_value'))
        if not part_number:
            raise HTTPException(status_code=400, detail=f'Row {row_number}: Part Number is required')
        try:
            qty = float(row.get('quantity') or row.get('required_qty') or 0)
            value = float(str(raw_value if raw_value is not None else 0).replace(',', '').replace('₹', '').strip() or 0)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f'Row {row_number}: Invalid quantity or value')
        if qty <= 0:
            raise HTTPException(status_code=400, detail=f'Row {row_number}: Quantity must be greater than zero')
        add_qty += qty
        add_value += value * qty
        new_items.append({
            'id': str(uuid.uuid4()), 'order_id': order_id, 'order_number': order.get('order_number'),
            'part_number': part_number, 'description': description,
            'required_qty': qty, 'unit_value': value,
            'allocated_qty': 0.0, 'balance_qty': qty, 'accepted_qty': 0.0, 'remaining_qty': qty,
            'availability_status': 'Not Checked', 'allocations': [],
            'status': 'Order Created', 'request_status': 'Order Created',
            'added_after_order_creation': True,
            'added_by': current_user.id, 'added_by_name': current_user.username, 'added_at': now,
            'original_order_created_at': original_created_at,
            'created_at': now, 'updated_at': now,
        })
    await db.order_items.insert_many([dict(i) for i in new_items])
    await db.order_headers.update_one({'id': order_id}, {'$set': {
        'item_count': int(order.get('item_count') or 0) + len(new_items),
        'total_required_qty': float(order.get('total_required_qty') or 0) + add_qty,
        'total_order_value': float(order.get('total_order_value') or 0) + add_value,
        'availability_checked': False,
        'updated_at': now,
    }})
    await odw.append_order_audit(db, order, 'Item Added', current_user, {
        'count': len(new_items),
        'parts': [i['part_number'] for i in new_items],
    })
    detail = await order_desk_order_detail(order_id, current_user=current_user)
    return {'message': f'{len(new_items)} item(s) added to {order.get("order_number")}', **detail}


@api_router.post('/order-desk/orders/{order_id}/re-enquire')
async def order_desk_re_enquire(order_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    """Mark selected / remaining items for the next enquiry cycle.

    Does NOT auto-suggest and does NOT send any request. Remaining qty only.
    Accepted quantities and prior request history stay intact.
    """
    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail='Order not found')
    role = (current_user.role or '').lower()
    if role != 'master' and order.get('created_by') != current_user.id and order.get('dealer_name') != current_user.group:
        raise HTTPException(status_code=403, detail='Not authorized')

    item_ids = [str(x) for x in ((payload or {}).get('item_ids') or []) if x]
    items = await db.order_items.find({'order_id': order_id}, {'_id': 0}).to_list(10000)
    now = datetime.now(timezone.utc).isoformat()
    enriched = await odw.enrich_order_items(db, order, items)
    updated = 0
    for item in enriched:
        if item_ids and item.get('id') not in item_ids:
            continue
        remaining = float(item.get('remaining_qty') or 0)
        if remaining <= 0 and not item.get('retry_required'):
            continue
        # Determine next stage: branch first, then dealer, then factory.
        excluded = {
            (str(x.get('dealer_name') or '').lower(), str(x.get('branch') or '').lower())
            for x in (item.get('excluded_sources') or [])
        }
        branch_left = [
            s for s in (item.get('same_dealer_sources') or [])
            if (str(s.get('dealer_name') or '').lower(), str(s.get('branch') or '').lower()) not in excluded
            and float(s.get('net_available_qty', s.get('available_qty') or 0) or 0) > 0
        ]
        dealer_left = [
            s for s in (item.get('other_dealer_sources') or [])
            if (str(s.get('dealer_name') or '').lower(), str(s.get('branch') or '').lower()) not in excluded
            and float(s.get('net_available_qty', s.get('available_qty') or 0) or 0) > 0
        ]
        if branch_left:
            next_stage = 'branch'
            factory_qty = 0
            no_further = False
            status = odw.REQUEST_STATUS_REMAINING
        elif dealer_left:
            next_stage = 'dealer'
            factory_qty = 0
            no_further = False
            status = odw.REQUEST_STATUS_REMAINING
        else:
            next_stage = 'factory'
            factory_qty = remaining
            no_further = True
            status = odw.REQUEST_STATUS_FACTORY

        await db.order_items.update_one({'id': item['id']}, {'$set': {
            'retry_required': not no_further,
            're_enquire': not no_further,
            'retry_selected': True if (payload or {}).get('select', True) else item.get('retry_selected', False),
            'enquiry_stage': next_stage,
            'remaining_qty': remaining,
            'factory_order_qty': factory_qty,
            'no_further_stock': no_further,
            'request_status': status,
            'status': status if no_further else 'Pending Retry',
            # Clear only unsent draft allocations so Auto Suggest can refill remaining.
            'allocations': [
                a for a in (item.get('allocations') or [])
                if a.get('request_no') or a.get('request_number')
                or str(a.get('status') or '').lower() in (
                    'request sent', 'requested', 'accepted', 'partially accepted',
                    'rejected', 'cancelled', 'completed', 'dispatched', 'received',
                )
            ],
            'updated_at': now,
        }})
        updated += 1

    await odw.append_order_audit(db, order, 'Re-Enquire Remaining Qty', current_user, {
        'item_ids': item_ids or 'all_remaining', 'updated': updated,
    })
    detail = await order_desk_order_detail(order_id, current_user=current_user)
    return {'message': f'{updated} item(s) marked for re-enquiry', 'updated': updated, **detail}


CANCELLATION_REASONS = list(odw.CANCELLATION_REASONS)


@api_router.post('/order-desk/orders/{order_id}/items/{item_id}/request-cancellation')
async def order_desk_request_cancellation(order_id: str, item_id: str, payload: dict, current_user: UserResponse = Depends(get_current_user)):
    """Request cancellation of an order item. No physical delete. Reason required."""
    order = await db.order_headers.find_one({'id': order_id}, {'_id': 0})
    item = await db.order_items.find_one({'id': item_id, 'order_id': order_id}, {'_id': 0})
    if not order or not item:
        raise HTTPException(status_code=404, detail='Order item not found')
    role = (current_user.role or '').lower()
    if role != 'master' and order.get('created_by') != current_user.id and order.get('dealer_name') != current_user.group:
        raise HTTPException(status_code=403, detail='Not authorized')

    reason = _order_clean_text((payload or {}).get('reason'))
    remarks = _order_clean_text((payload or {}).get('remarks'))
    if reason not in CANCELLATION_REASONS:
        raise HTTPException(status_code=400, detail=f'Cancellation reason is required. Allowed: {", ".join(CANCELLATION_REASONS)}')
    if reason == 'Other' and not remarks:
        raise HTTPException(status_code=400, detail='Remarks are required when reason is Other')

    now = datetime.now(timezone.utc).isoformat()
    # Safe auto-approval: no request sent, no reservation, no acceptance, no dispatch/receive.
    item_reqs = await db.order_requests.find({'order_item_id': item_id}, {'_id': 0}).to_list(1000)
    # Block cancellation of locked awaiting requests before timeout — use cancel-timeout instead.
    pending = [r for r in item_reqs if r.get('status') == 'Requested']
    if pending and reason != 'Cancelled – No Response':
        for r in pending:
            header = await db.request_headers.find_one({'request_number': r.get('request_number')}, {'_id': 0})
            timer = odw.evaluate_group_timer(header or {})
            if not timer.get('cancel_allowed'):
                raise HTTPException(
                    status_code=400,
                    detail='Request Sent – Awaiting Response. Cancellation is disabled until the response timer expires.',
                )
    # Accepted qty is never cancelled via this path for the accepted portion
    if any(float(r.get('accepted_qty') or r.get('approved_qty') or 0) > 0 and r.get('status') in (
        'Approved', 'Partially Approved', 'Dispatched', 'Received', 'Completed'
    ) for r in item_reqs) and float(item.get('remaining_qty') or 0) <= 0:
        raise HTTPException(status_code=400, detail='Accepted quantity is locked and cannot be cancelled')

    active_reservations = await db.stock_reservations.count_documents({'order_item_id': item_id, 'status': 'active'})
    has_sent = any(r.get('status') not in (None, '') for r in item_reqs)
    has_acceptance = any(float(r.get('accepted_qty') or r.get('approved_qty') or 0) > 0 for r in item_reqs)
    has_logistics = any(r.get('status') in ('Dispatched', 'Received', 'Completed') for r in item_reqs)
    safe_auto = (not has_sent) and active_reservations == 0 and (not has_acceptance) and (not has_logistics)

    cancel_doc = {
        'id': str(uuid.uuid4()),
        'order_id': order_id,
        'order_number': order.get('order_number'),
        'order_item_id': item_id,
        'part_number': item.get('part_number'),
        'part_name': item.get('description'),
        'qty': item.get('required_qty'),
        'reason': reason,
        'remarks': remarks,
        'cancellation_requested_by': current_user.id,
        'cancellation_requested_by_name': current_user.username,
        'cancellation_requested_at': now,
        'approval_status': 'approved' if safe_auto else 'pending',
        'auto_approved': safe_auto,
        'purchased_outside': reason == 'Purchased Outside',
        'purchased_outside_qty': float(item.get('required_qty') or 0) if reason == 'Purchased Outside' else 0,
        'brand': order.get('brand_name'),
        'dealer': order.get('dealer_name'),
        'branch': order.get('branch'),
        'created_at': now,
    }
    if safe_auto:
        cancel_doc.update({
            'approved_by': 'system', 'approved_by_name': 'Auto Approval',
            'approved_at': now, 'cancelled_by': 'system', 'cancelled_at': now,
        })

    await db.order_cancellation_requests.insert_one(dict(cancel_doc))

    item_update = {
        'cancellation_requested': True,
        'cancellation_reason': reason,
        'cancellation_remarks': remarks,
        'cancellation_requested_by': current_user.id,
        'cancellation_requested_at': now,
        'cancellation_status': 'approved' if safe_auto else 'pending',
        'cancellation_request_id': cancel_doc['id'],
        'request_status': odw.REQUEST_STATUS_CANCELLED if safe_auto else odw.REQUEST_STATUS_CANCEL_REQ,
        'status': odw.REQUEST_STATUS_CANCELLED if safe_auto else odw.REQUEST_STATUS_CANCEL_REQ,
        'updated_at': now,
    }
    if reason == 'Purchased Outside':
        item_update['purchased_outside'] = True
        item_update['purchased_outside_qty'] = float(item.get('required_qty') or 0)
        item_update['purchased_outside_value'] = float(item.get('unit_value') or 0) * float(item.get('required_qty') or 0)
    if safe_auto:
        item_update.update({'cancelled_by': 'system', 'cancelled_at': now})
        # Clear unsent draft allocations only; keep any historical allocations if somehow present
        item_update['allocations'] = [
            a for a in (item.get('allocations') or [])
            if a.get('request_no') or a.get('request_number')
        ]

    await db.order_items.update_one({'id': item_id}, {'$set': item_update})
    await odw.append_order_audit(db, order,
        'Cancellation approved (auto)' if safe_auto else 'Cancellation requested',
        current_user,
        {'reason': reason, 'remarks': remarks, 'item_id': item_id, 'auto': safe_auto, 'purchased_outside': reason == 'Purchased Outside'},
    )
    return {
        'message': 'Item cancelled automatically (safe — no request/reservation).' if safe_auto
            else 'Cancellation requested — awaiting Admin/Master approval.',
        'auto_approved': safe_auto,
        'cancellation': {k: v for k, v in cancel_doc.items() if k != '_id'},
    }


@api_router.post('/order-desk/cancellations/{cancellation_id}/decide')
async def order_desk_decide_cancellation(cancellation_id: str, payload: dict, current_user: UserResponse = Depends(get_current_user)):
    """Admin/Master approve or reject a cancellation request."""
    role = (current_user.role or '').lower()
    if role not in ('master', 'admin'):
        raise HTTPException(status_code=403, detail='Only Admin/Master can approve cancellations')
    doc = await db.order_cancellation_requests.find_one({'id': cancellation_id}, {'_id': 0})
    if not doc:
        raise HTTPException(status_code=404, detail='Cancellation request not found')
    if doc.get('approval_status') != 'pending':
        return {'message': f'Already {doc.get("approval_status")}', 'cancellation': doc}

    decision = str((payload or {}).get('decision') or '').strip().lower()
    if decision not in ('approve', 'approved', 'reject', 'rejected'):
        raise HTTPException(status_code=400, detail="decision must be 'approve' or 'reject'")
    now = datetime.now(timezone.utc).isoformat()
    approved = decision in ('approve', 'approved')
    await db.order_cancellation_requests.update_one({'id': cancellation_id}, {'$set': {
        'approval_status': 'approved' if approved else 'rejected',
        'approved_by' if approved else 'rejected_by': current_user.id,
        'approved_by_name' if approved else 'rejected_by_name': current_user.username,
        'approved_at' if approved else 'rejected_at': now,
        'decision_remarks': _order_clean_text((payload or {}).get('remarks')),
        'updated_at': now,
    }})
    item_set = {
        'cancellation_status': 'approved' if approved else 'rejected',
        'updated_at': now,
    }
    if approved:
        item_set.update({
            'request_status': odw.REQUEST_STATUS_CANCELLED,
            'status': odw.REQUEST_STATUS_CANCELLED,
            'cancelled_by': current_user.id,
            'cancelled_at': now,
        })
        # Also cancel any still-open Request Center lines for this item
        open_reqs = await db.order_requests.find(
            {'order_item_id': doc.get('order_item_id'), 'status': {'$in': ['Requested', 'Approved', 'Partially Approved']}},
            {'_id': 0, 'id': 1},
        ).to_list(1000)
        for r in open_reqs:
            await _request_center_transition(r['id'], 'Cancelled', f"Order item cancellation: {doc.get('reason')}", current_user)
    else:
        item_set.update({
            'cancellation_requested': False,
            'request_status': 'Request Sent',  # will be re-enriched on next load
        })
    await db.order_items.update_one({'id': doc.get('order_item_id')}, {'$set': item_set})
    order = await db.order_headers.find_one({'id': doc.get('order_id')}, {'_id': 0}) or {}
    await odw.append_order_audit(
        db, order,
        'Cancellation approved' if approved else 'Cancellation rejected',
        current_user,
        {'cancellation_id': cancellation_id, 'reason': doc.get('reason')},
    )
    updated = await db.order_cancellation_requests.find_one({'id': cancellation_id}, {'_id': 0})
    return {'message': f'Cancellation {updated.get("approval_status")}', 'cancellation': updated}


@api_router.post('/requests/group/{request_number}/resend-email')
async def resend_request_group_email(request_number: str, current_user: UserResponse = Depends(get_current_user)):
    """Safe, explicit retry for a request whose email previously failed.
    Never creates a new request or a new request number."""
    group_doc = await db.request_headers.find_one({'request_number': request_number}, {'_id': 0})
    if not group_doc:
        raise HTTPException(status_code=404, detail='Request not found')
    role = (current_user.role or '').lower()
    is_supplier = role == 'master' or group_doc.get('supplying_dealer') == current_user.group
    is_requester = group_doc.get('requested_by') == current_user.id or (role != 'user' and group_doc.get('requesting_dealer') == current_user.group)
    if not (is_supplier or is_requester):
        raise HTTPException(status_code=403, detail='Not authorized for this request')
    if group_doc.get('email_sent'):
        return {'message': 'Email already sent successfully; no resend needed.', 'email_sent': True, 'email_status': 'Email Sent'}

    await db.request_headers.update_one({'id': group_doc['id']}, {'$set': {'retry_count': int(group_doc.get('retry_count', 0)) + 1}})
    refreshed = await db.request_headers.find_one({'id': group_doc['id']}, {'_id': 0})
    await _send_request_group_email(refreshed)
    updated = await db.request_headers.find_one({'id': group_doc['id']}, {'_id': 0})
    order = await db.order_headers.find_one({'id': updated.get('order_id')}, {'_id': 0}) or {}
    await odw.append_order_audit(
        db, order,
        'Email sent' if updated.get('email_sent') else 'Email failed/retried',
        current_user,
        {'request_number': request_number, 'email_status': updated.get('email_status'), 'email_error': updated.get('email_error')},
    )
    return {
        'message': 'Resend attempted.',
        'email_sent': updated.get('email_sent'),
        'email_error': updated.get('email_error'),
        'email_status': odw.email_status_label(updated.get('email_status') if not updated.get('email_sent') else 'sent'),
    }


@api_router.post('/requests/group/{request_number}/cancel-timeout')
async def cancel_request_group_timeout(request_number: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    """Cancel unanswered items after response deadline. Does not unlock Accepted qty.
    Cancellation reason: Cancelled – No Response. Does NOT create a Rejected-Today freeze.
    """
    group_doc = await db.request_headers.find_one({'request_number': request_number}, {'_id': 0})
    if not group_doc:
        raise HTTPException(status_code=404, detail='Request not found')
    role = (current_user.role or '').lower()
    is_requester = group_doc.get('requested_by') == current_user.id or (role != 'user' and group_doc.get('requesting_dealer') == current_user.group)
    if role != 'master' and not is_requester:
        raise HTTPException(status_code=403, detail='Not authorized to cancel this request')

    timer = odw.evaluate_group_timer(group_doc)
    if not timer.get('cancel_allowed'):
        raise HTTPException(status_code=400, detail='Cancel is only allowed after the response deadline expires')

    now = datetime.now(timezone.utc).isoformat()
    pending = await db.order_requests.find(
        {'request_number': request_number, 'status': 'Requested'}, {'_id': 0},
    ).to_list(10000)
    if not pending:
        return {'message': 'No unanswered items to cancel', 'cancelled': 0, 'cancel_allowed': False}

    cancelled = 0
    for req in pending:
        updated, changed = await _request_center_transition(
            req['id'], 'Cancelled', 'Cancelled – No Response', current_user,
        )
        if changed:
            cancelled += 1
            await db.order_requests.update_one({'id': req['id']}, {'$set': {
                'cancellation_reason': 'Cancelled – No Response',
                'timeout_cancelled': True,
                'timeout_cancelled_at': now,
            }})
            # Prefer next source — exclude this source for current order attempts,
            # but do NOT write a day-wide Rejected Today freeze.
            await db.order_items.update_one({'id': req.get('order_item_id')}, {
                '$addToSet': {'excluded_sources': {
                    'dealer_name': req.get('supplying_dealer', ''),
                    'branch': req.get('supplying_branch', ''),
                    'reason': 'no_response',
                }},
                '$set': {'retry_required': True, 'updated_at': now},
            })

    await db.request_headers.update_one({'id': group_doc['id']}, {'$set': {
        'response_status': 'cancelled',
        'timeout_cancelled': True,
        'timeout_cancelled_at': now,
        'updated_at': now,
    }})
    order = await db.order_headers.find_one({'id': group_doc.get('order_id')}, {'_id': 0}) or {}
    await odw.append_order_audit(db, order, 'Cancelled – No Response', current_user, {
        'request_number': request_number, 'cancelled': cancelled,
    })
    return {
        'message': f'{cancelled} unanswered item(s) cancelled – no response.',
        'cancelled': cancelled,
        'cancel_allowed': False,
        'request_number': request_number,
    }


@api_router.get('/requests/group/{request_number}')
async def request_group_detail(request_number: str, current_user: UserResponse = Depends(get_current_user)):
    group_doc = await db.request_headers.find_one({'request_number': request_number}, {'_id': 0})
    if not group_doc:
        raise HTTPException(status_code=404, detail='Request not found')
    role = (current_user.role or '').lower()
    is_supplier = role == 'master' or group_doc.get('supplying_dealer') == current_user.group
    is_requester = group_doc.get('requested_by') == current_user.id or (role != 'user' and group_doc.get('requesting_dealer') == current_user.group)
    if not (is_supplier or is_requester):
        raise HTTPException(status_code=403, detail='Not authorized for this request')
    timer = odw.evaluate_group_timer(group_doc)
    group_doc = {
        **group_doc,
        **timer,
        'email_status_label': odw.email_status_label(
            'sent' if group_doc.get('email_sent') else group_doc.get('email_status')
        ),
    }
    return group_doc


REQUEST_CENTER_TRANSITIONS = {
    'Requested': {'Approved', 'Rejected', 'Cancelled'},
    'Approved': {'Dispatched', 'Cancelled'},
    'Partially Approved': {'Dispatched', 'Cancelled'},
    'Dispatched': {'Received'},
    'Received': {'Completed'},
}


@api_router.get('/requests')
async def request_center_list(
    view: str = 'all', status_filter: Optional[str] = None, search: Optional[str] = None,
    brand: Optional[str] = None, dealer: Optional[str] = None, branch: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user),
):
    """Request Center listing, scoped by role like the rest of the app.
    view=incoming -> requests this user's dealer/branch must supply.
    view=outgoing -> requests this user/dealer raised. view=all -> both."""
    role = (current_user.role or '').lower()
    outgoing_query, incoming_query = {}, {}
    if role == 'master':
        pass
    elif role == 'admin':
        outgoing_query['requesting_dealer'] = current_user.group
        incoming_query['supplying_dealer'] = current_user.group
    else:
        outgoing_query['requested_by'] = current_user.id
        incoming_query = {'supplying_dealer': current_user.group, 'supplying_branch': current_user.location}

    if view == 'incoming':
        query = incoming_query
    elif view == 'outgoing':
        query = outgoing_query
    else:
        clauses = [q for q in (outgoing_query, incoming_query) if q]
        query = {'$or': clauses} if (role != 'master' and clauses) else {}

    # Global dashboard scope is applied directionally for each Request Center view:
    # incoming = supplying/source scope, outgoing = requesting/destination scope,
    # all = either side. This prevents a selected destination branch from leaking
    # into Requests To Me merely because it appears on the opposite side.
    scope_clauses = []
    def _selected(value, all_label):
        value = (value or '').strip()
        return value if value and value.lower() != all_label.lower() else ''

    selected_brand = _selected(brand, 'All Brands')
    selected_dealer = _selected(dealer, 'All Dealers')
    selected_branch = _selected(branch, 'All Branches')
    if role != 'master':
        # Never allow a client to broaden or replace the authenticated assignment.
        assigned_dealer = (current_user.group or '').strip()
        assigned_branch = (current_user.location or '').strip()
        if assigned_dealer:
            selected_dealer = assigned_dealer
        if role == 'user' and assigned_branch:
            selected_branch = assigned_branch
    def _scope_match(field_suffix: str, value: str):
        if view == 'incoming':
            return {f'supplying_{field_suffix}': value}
        if view == 'outgoing':
            return {f'requesting_{field_suffix}': value}
        return {'$or': [
            {f'requesting_{field_suffix}': value},
            {f'supplying_{field_suffix}': value},
        ]}

    if selected_brand:
        scope_clauses.append(_scope_match('brand', selected_brand))
    if selected_dealer:
        scope_clauses.append(_scope_match('dealer', selected_dealer))
    if selected_branch:
        scope_clauses.append(_scope_match('branch', selected_branch))
    if scope_clauses:
        query = {'$and': [query, *scope_clauses]} if query else {'$and': scope_clauses}

    if status_filter:
        query = {'$and': [query, {'status': status_filter}]} if query else {'status': status_filter}
    if search:
        escaped = re.escape(search.strip())
        search_clause = {'$or': [
            {'order_number': {'$regex': escaped, '$options': 'i'}},
            {'part_number': {'$regex': escaped, '$options': 'i'}},
            {'request_number': {'$regex': escaped, '$options': 'i'}},
        ]}
        query = {'$and': [query, search_clause]} if query else search_clause

    rows = await db.order_requests.find(query, {'_id': 0}).sort('requested_at', -1).limit(2000).to_list(2000)

    # Enrich the list with user-facing sender/receiver identities without
    # changing the stored request workflow. A request is addressed to the
    # supplying branch team, so all active users in that exact scope are
    # returned as potential receivers. LOC remains the item-level snapshot
    # stored in loc_at_request and is never replaced by the branch name.
    supplier_scopes = {
        (str(r.get('supplying_brand') or r.get('requesting_brand') or '').strip(),
         str(r.get('supplying_dealer') or '').strip(),
         str(r.get('supplying_branch') or '').strip())
        for r in rows
    }
    receiver_map = {}
    for brand, dealer, branch in supplier_scopes:
        uq = {'status': 'active'}
        if brand:
            uq['brand'] = brand
        if dealer:
            uq['group'] = dealer
        if branch:
            uq['location'] = branch
        users = await db.users.find(uq, {'_id': 0, 'id': 1, 'user_id': 1, 'username': 1, 'name': 1}).to_list(200)
        receiver_map[(brand, dealer, branch)] = [{
            'id': u.get('user_id') or u.get('id') or '',
            'name': u.get('name') or u.get('username') or '',
        } for u in users]

    for row in rows:
        scope = (
            str(row.get('supplying_brand') or row.get('requesting_brand') or '').strip(),
            str(row.get('supplying_dealer') or '').strip(),
            str(row.get('supplying_branch') or '').strip(),
        )
        row['requested_user_id'] = row.get('requested_by') or ''
        row['receiver_users'] = receiver_map.get(scope, [])
    return rows


@api_router.get('/requests/{request_id}')
async def request_center_detail(request_id: str, current_user: UserResponse = Depends(get_current_user)):
    req = await db.order_requests.find_one({'id': request_id}, {'_id': 0})
    if not req:
        raise HTTPException(status_code=404, detail='Request not found')
    role = (current_user.role or '').lower()
    is_supplier = role == 'master' or req.get('supplying_dealer') == current_user.group
    is_requester = req.get('requested_by') == current_user.id or (role != 'user' and req.get('requesting_dealer') == current_user.group)
    if not (is_supplier or is_requester):
        raise HTTPException(status_code=403, detail='Not authorized for this request')
    return req


async def _sync_request_header_after_item_decision(req: dict, now: str):
    """Keep the request header/PDF summary aligned with item-wise decisions."""
    request_number = req.get('request_number')
    if not request_number:
        return
    items = await db.order_requests.find({'request_number': request_number}, {'_id': 0}).to_list(10000)
    if not items:
        return
    statuses = [i.get('status', 'Requested') for i in items]
    accepted_items = [i for i in items if float(i.get('accepted_qty', i.get('approved_qty', 0)) or 0) > 0]
    accepted_statuses = [i.get('status', 'Approved') for i in accepted_items]
    if any(st == 'Requested' for st in statuses):
        header_status = 'Requested'
    elif accepted_statuses and all(st == 'Completed' for st in accepted_statuses):
        header_status = 'Completed'
    elif accepted_statuses and all(st in ('Received', 'Completed') for st in accepted_statuses):
        header_status = 'Received'
    elif accepted_statuses and all(st in ('Dispatched', 'Received', 'Completed') for st in accepted_statuses):
        header_status = 'Dispatched'
    elif all(st == 'Rejected' for st in statuses):
        header_status = 'Rejected'
    elif all(st == 'Cancelled' for st in statuses):
        header_status = 'Cancelled'
    elif accepted_items:
        is_partial = any(
            i.get('status') in ('Rejected', 'Cancelled') or
            float(i.get('accepted_qty', i.get('approved_qty', 0)) or 0) < float(i.get('requested_qty') or 0)
            for i in items
        )
        header_status = 'Partially Approved' if is_partial else 'Approved'
    else:
        header_status = 'Partially Approved'

    header_items = []
    accepted_total = 0.0
    for i in items:
        accepted = float(i.get('accepted_qty', i.get('approved_qty', 0)) or 0)
        accepted_total += accepted
        header_items.append({
            'order_request_id': i.get('id'), 'part_number': i.get('part_number'),
            'description': i.get('description'), 'requested_qty': i.get('requested_qty'),
            'accepted_qty': accepted, 'approved_qty': accepted,
            'available_qty_at_request': i.get('available_qty_at_request'),
            'unit_value': i.get('unit_value_at_request'), 'value': i.get('value_at_request'),
            'purchase_aging_days': i.get('purchase_aging_days_at_request'),
            'sales_aging_days': i.get('sales_aging_days_at_request'),
            'loc': i.get('loc_at_request', ''), 'status': i.get('status'),
            'remarks': i.get('approval_remarks') or i.get('remarks') or '',
        })
    await db.request_headers.update_one(
        {'request_number': request_number},
        {'$set': {'status': header_status, 'items': header_items, 'accepted_total_qty': accepted_total, 'updated_at': now}},
    )


async def _request_center_transition(request_id: str, new_status: str, remarks: str, current_user: UserResponse, accepted_qty=None):
    req = await db.order_requests.find_one({'id': request_id}, {'_id': 0})
    if not req:
        raise HTTPException(status_code=404, detail='Request not found')

    role = (current_user.role or '').lower()
    is_supplier = role == 'master' or req.get('supplying_dealer') == current_user.group
    if new_status != 'Cancelled' and not is_supplier:
        raise HTTPException(status_code=403, detail='Only the supplying Admin/Master can approve or reject a request')
    is_requester = req.get('requested_by') == current_user.id or (role != 'user' and req.get('requesting_dealer') == current_user.group)
    if new_status == 'Cancelled' and not (is_supplier or is_requester):
        raise HTTPException(status_code=403, detail='Not authorized to cancel this request')

    old_status = req.get('status')
    allowed = REQUEST_CENTER_TRANSITIONS.get(old_status, set())
    if new_status not in allowed:
        if old_status == new_status:
            return req, False
        raise HTTPException(status_code=400, detail=f'Cannot move a request from {old_status} to {new_status}')

    requested_qty = float(req.get('requested_qty') or 0)
    accepted = 0.0
    if new_status == 'Approved':
        try:
            accepted = requested_qty if accepted_qty is None else float(accepted_qty)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail='Accept Quantity must be a valid number')
        if accepted <= 0:
            raise HTTPException(status_code=400, detail='Accept Quantity must be greater than zero. Use Reject for unavailable parts.')
        if accepted > requested_qty:
            raise HTTPException(status_code=400, detail='Accept Quantity cannot exceed Request Quantity')

    now = datetime.now(timezone.utc).isoformat()
    update = {
        'status': new_status, 'updated_at': now, 'approval_remarks': sanitize_text_safe(remarks),
        'accepted_qty': accepted, 'approved_qty': accepted,
        'decision_type': ('Partial' if new_status == 'Approved' and accepted < requested_qty else new_status),
        'decided_by': current_user.id, 'decided_user_name': current_user.username, 'decided_at': now,
    }
    await db.order_requests.update_one({'id': request_id}, {'$set': update})

    if new_status in ('Rejected', 'Cancelled') or (new_status == 'Approved' and accepted < requested_qty):
        # Preserve accepted qty + request history; mark remaining for Re-Enquire.
        # Do NOT wipe allocations / accepted quantities (Order Desk audit requirement).
        await odw.sync_order_item_after_request_decision(db, {**req, **update}, now)
    elif new_status == 'Approved':
        await odw.sync_order_item_after_request_decision(db, {**req, **update}, now)

    if new_status in ('Rejected', 'Cancelled'):
        await db.stock_reservations.update_many(
            {'order_request_id': request_id, 'status': 'active'},
            {'$set': {'status': 'released', 'released_at': now, 'released_qty': requested_qty}},
        )
    elif new_status == 'Approved':
        # Keep only the accepted portion reserved and release the balance for
        # the next Auto Suggest / order without changing the original request quantity.
        released_qty = max(0.0, requested_qty - accepted)
        await db.stock_reservations.update_many(
            {'order_request_id': request_id, 'status': 'active'},
            {'$set': {'qty': accepted, 'accepted_qty': accepted, 'released_qty': released_qty,
                      'partially_released_at': now if released_qty else None}},
        )

    await db.order_activity.insert_one({
        'id': str(uuid.uuid4()), 'order_id': req.get('order_id'), 'order_number': req.get('order_number'),
        'request_id': request_id, 'action': f'Request Item {new_status}', 'performed_by': current_user.id,
        'performed_user_name': current_user.username, 'role': role, 'old_status': old_status,
        'new_status': new_status, 'requested_qty': requested_qty, 'accepted_qty': accepted,
        'remarks': sanitize_text_safe(remarks), 'created_at': now,
    })
    updated = await db.order_requests.find_one({'id': request_id}, {'_id': 0})
    await _sync_request_header_after_item_decision(updated, now)
    return updated, True

def sanitize_text_safe(value):
    return notifications.sanitize_text(value, 500) if value else ''


async def _notify_request_status_change(req: dict, event: str):
    try:
        recipients = []
        if req.get('requester_email') or req.get('requester_mobile'):
            recipients.append({'email': req.get('requester_email'), 'mobile': req.get('requester_mobile'), 'name': req.get('requested_user_name')})
        supplier_recipients = await _active_recipients_for_scope(req.get('supplying_brand'), req.get('supplying_dealer'), req.get('supplying_branch'))
        recipients.extend(supplier_recipients)
        if recipients:
            await notifications.notify_request_event(db, event, req, recipients, remarks=req.get('approval_remarks', ''))
    except Exception as exc:  # noqa: BLE001
        logging.getLogger('nmts.notifications').warning('%s notification dispatch failed: %s', event, str(exc)[:300])


@api_router.post('/requests/{request_id}/approve')
async def request_center_approve(request_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    remarks = (payload or {}).get('remarks', '') if payload else ''
    accepted_qty = (payload or {}).get('accepted_qty', (payload or {}).get('accept_qty', (payload or {}).get('approved_qty')))
    req = await db.order_requests.find_one({'id': request_id}, {'_id': 0})
    if not req: raise HTTPException(status_code=404, detail='Request not found')
    try: accepted = float(accepted_qty)
    except (TypeError, ValueError): raise HTTPException(status_code=400, detail='Accepted Quantity must be numeric')
    requested = float(req.get('requested_qty') or 0)
    if accepted < 0: raise HTTPException(status_code=400, detail='Accepted Quantity cannot be negative')
    if accepted > requested: raise HTTPException(status_code=400, detail='Accepted Quantity cannot exceed Requested Quantity')
    if accepted < requested and not str(remarks).strip(): raise HTTPException(status_code=400, detail='Remark is required for Partial or Rejected responses')
    target = 'Rejected' if accepted == 0 else 'Approved'
    updated, changed = await _request_center_transition(request_id, target, remarks, current_user, accepted_qty=accepted)
    if changed: await _notify_request_status_change(updated, 'Request Rejected' if target == 'Rejected' else ('Request Partially Accepted' if accepted < requested else 'Request Accepted'))
    return updated


@api_router.post('/requests/{request_id}/reject')
async def request_center_reject(request_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    remarks = (payload or {}).get('remarks', '') if payload else ''
    if not str(remarks).strip():
        raise HTTPException(status_code=400, detail='A rejection reason or remark is required')
    updated, changed = await _request_center_transition(request_id, 'Rejected', remarks, current_user)
    if changed:
        await _notify_request_status_change(updated, 'Request Rejected')
    return updated


@api_router.post('/requests/{request_id}/cancel')
async def request_center_cancel(request_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    remarks = (payload or {}).get('remarks', '') if payload else ''
    updated, changed = await _request_center_transition(request_id, 'Cancelled', remarks, current_user)
    if changed:
        await _notify_request_status_change(updated, 'Request Cancelled')
    return updated


async def _request_logistics_transition(request_id: str, new_status: str, remarks: str, current_user: UserResponse):
    req = await db.order_requests.find_one({'id': request_id}, {'_id': 0})
    if not req:
        raise HTTPException(status_code=404, detail='Request not found')
    role = (current_user.role or '').lower()
    is_supplier = role == 'master' or req.get('supplying_dealer') == current_user.group
    is_requester = role == 'master' or req.get('requested_by') == current_user.id or (role != 'user' and req.get('requesting_dealer') == current_user.group)
    if new_status == 'Dispatched' and not is_supplier:
        raise HTTPException(status_code=403, detail='Only the supplying scope can dispatch accepted parts')
    if new_status in ('Received', 'Completed') and not is_requester:
        raise HTTPException(status_code=403, detail='Only the requesting scope can receive or complete accepted parts')
    old_status = req.get('status')
    if old_status == new_status:
        return req, False
    if new_status not in REQUEST_CENTER_TRANSITIONS.get(old_status, set()):
        raise HTTPException(status_code=400, detail=f'Cannot move a request from {old_status} to {new_status}')
    accepted = float(req.get('accepted_qty', req.get('approved_qty', 0)) or 0)
    if accepted <= 0:
        raise HTTPException(status_code=400, detail='Only accepted quantities can continue in the transfer workflow')
    now = datetime.now(timezone.utc).isoformat()
    update = {'status': new_status, 'updated_at': now, 'workflow_remarks': sanitize_text_safe(remarks),
              f'{new_status.lower()}_by': current_user.id, f'{new_status.lower()}_at': now}
    await db.order_requests.update_one({'id': request_id}, {'$set': update})
    if new_status == 'Completed':
        await db.stock_reservations.update_many(
            {'order_request_id': request_id, 'status': 'active'},
            {'$set': {'status': 'completed', 'completed_at': now, 'completed_qty': accepted}},
        )
    await db.order_activity.insert_one({
        'id': str(uuid.uuid4()), 'order_id': req.get('order_id'), 'order_number': req.get('order_number'),
        'request_id': request_id, 'action': f'Request Item {new_status}', 'performed_by': current_user.id,
        'performed_user_name': current_user.username, 'role': role, 'old_status': old_status,
        'new_status': new_status, 'accepted_qty': accepted, 'remarks': sanitize_text_safe(remarks), 'created_at': now,
    })
    updated = await db.order_requests.find_one({'id': request_id}, {'_id': 0})
    await _sync_request_header_after_item_decision(updated, now)
    return updated, True


@api_router.post('/requests/{request_id}/dispatch')
async def request_center_dispatch(request_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    updated, changed = await _request_logistics_transition(request_id, 'Dispatched', (payload or {}).get('remarks', ''), current_user)
    if changed:
        await _notify_request_status_change(updated, 'Request Dispatched')
    return updated


@api_router.post('/requests/{request_id}/receive')
async def request_center_receive(request_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    updated, changed = await _request_logistics_transition(request_id, 'Received', (payload or {}).get('remarks', ''), current_user)
    if changed:
        await _notify_request_status_change(updated, 'Request Received')
    return updated


@api_router.post('/requests/{request_id}/complete')
async def request_center_complete(request_id: str, payload: dict = None, current_user: UserResponse = Depends(get_current_user)):
    updated, changed = await _request_logistics_transition(request_id, 'Completed', (payload or {}).get('remarks', ''), current_user)
    if changed:
        await _notify_request_status_change(updated, 'Request Completed')
    return updated


# ==================== SLEEPING STOCK MOBILE ====================

class StockVerificationCreate(BaseModel):
    part_number: str
    physical_quantity: float
    scanned_location: str = ""
    remarks: str = ""
    brand: str = ""
    dealer: str = ""
    branch: str = ""


class StockVerificationItemCreate(BaseModel):
    part_number: str
    physical_quantity: float
    scanned_location: str = ""
    remarks: str = ""


class StockVerificationSessionCreate(BaseModel):
    brand: str
    dealer: str
    branch: str
    items: List[StockVerificationItemCreate]


class StockVerificationCorrection(BaseModel):
    correction_status: str = "pending"  # pending, corrected, not_required
    correction_method: str = ""  # system_corrected, physical_relocated, both, no_action
    correction_remarks: str = ""


def _mobile_dashboard_scope_query(
    current_user: UserResponse,
    brand: str = None,
    dealer: str = None,
    branch: str = None,
):
    """Use only exact Dashboard selections and enforce the user's permitted scope."""
    def exact(value):
        value = str(value or "").strip()
        return bool(value) and value != "N/A" and not value.lower().startswith("all ")

    if not exact(brand):
        raise HTTPException(status_code=400, detail="Select an exact Brand in the Dashboard filter")
    if not exact(dealer):
        raise HTTPException(status_code=400, detail="Select an exact Dealer in the Dashboard filter")
    if not exact(branch):
        raise HTTPException(status_code=400, detail="Select an exact Branch in the Dashboard filter")

    if current_user.role != "master":
        if current_user.brand and str(current_user.brand).casefold() != str(brand).strip().casefold():
            raise HTTPException(status_code=403, detail="Selected Brand is outside your permitted scope")
        if current_user.group and str(current_user.group).casefold() != str(dealer).strip().casefold():
            raise HTTPException(status_code=403, detail="Selected Dealer is outside your permitted scope")
        if current_user.role == "user" and current_user.location and str(current_user.location).casefold() != str(branch).strip().casefold():
            raise HTTPException(status_code=403, detail="Selected Branch is outside your permitted scope")

    return {
        "brand_name": _exact_ci(brand),
        "dealer_name": _exact_ci(dealer),
        "branch": _exact_ci(branch),
    }


# ==================== Legacy /mobile pairing, device, and stock-search routes removed ====================
# Superseded by the Mobile User + device-session model in mobile_api.py (see PART_26 rewrite).
# Perpetual Stock (web-authenticated verification/correction workflow) below is preserved unchanged.


def _stock_snapshot_quantity(product: dict) -> float:
    for field in ("available_qty_number", "available_quantity", "available_qty", "quantity"):
        value = product.get(field)
        if value not in (None, ""):
            try:
                return float(value)
            except (TypeError, ValueError):
                continue
    return 0.0


def _verification_scope_query(current_user: UserResponse, brand: str = None, dealer: str = None, branch: str = None):
    query = {}
    exact = lambda v: v and v != "N/A" and not str(v).startswith("All ")
    if current_user.role == "master":
        if exact(brand): query["brand_name"] = _exact_ci(brand)
        if exact(dealer): query["dealer_name"] = _exact_ci(dealer)
        if exact(branch): query["branch"] = _exact_ci(branch)
    elif current_user.role == "admin":
        if current_user.brand: query["brand_name"] = _exact_ci(current_user.brand)
        if current_user.group: query["dealer_name"] = _exact_ci(current_user.group)
        if exact(branch): query["branch"] = _exact_ci(branch)
    else:
        # Normal users must see every verification record for their assigned
        # branch, including records uploaded by paired mobile users.
        if current_user.brand: query["brand_name"] = _exact_ci(current_user.brand)
        if current_user.group: query["dealer_name"] = _exact_ci(current_user.group)
        if current_user.location: query["branch"] = _exact_ci(current_user.location)
    return query


@api_router.get("/mobile/perpetual-stock/lookup")
async def lookup_mobile_stock_snapshot(
    part_number: str,
    brand: str = None,
    dealer: str = None,
    branch: str = None,
    current_user: UserResponse = Depends(get_current_user),
):
    clean_part = (part_number or "").strip()
    if not clean_part:
        raise HTTPException(status_code=400, detail="Part number is required")

    query = _mobile_dashboard_scope_query(current_user, brand, dealer, branch)
    query["part_number"] = {"$regex": f"^{re.escape(clean_part)}$", "$options": "i"}
    product = await db.products.find_one(query, {"_id": 0}, sort=[("updated_at", -1)])
    if not product:
        raise HTTPException(
            status_code=404,
            detail=f"Part {clean_part} was not found under the selected Brand / Dealer / Branch",
        )

    from mobile_api import _product_pin_location

    bin_location = _product_pin_location(product)
    return {
        "product_id": product.get("id"),
        "part_number": product.get("part_number") or clean_part,
        "part_name": _resolve_verification_part_name(product),
        "system_quantity": _stock_snapshot_quantity(product),
        "system_location": bin_location,
        "pin_location": bin_location,
        "mav": float(product.get("mav") or product.get("MAV") or product.get("unit_value") or product.get("value") or 0),
        "brand_name": product.get("brand_name") or brand or current_user.brand,
        "dealer_name": product.get("dealer_name") or dealer or current_user.group,
        "branch": product.get("branch") or branch or current_user.location,
        "snapshot_at": datetime.now(timezone.utc),
        "information_only": True,
    }


async def _next_mops_verification_session_id() -> str:
    """MOPS + YYMMDD (IST) + 4-digit daily sequence. Atomic via db.counters."""
    india_now = datetime.now(ZoneInfo("Asia/Kolkata"))
    date_key = india_now.strftime("%y%m%d")
    counter_id = f"mops_verification_session_{date_key}"
    counter = await db.counters.find_one_and_update(
        {"_id": counter_id},
        {"$inc": {"seq": 1}, "$setOnInsert": {"date_key": date_key}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    seq = int(counter.get("seq", 1))
    if seq > 9999:
        raise HTTPException(status_code=500, detail="Daily MOPS verification session serial exhausted")
    return f"MOPS{date_key}{seq:04d}"


async def _next_perpetual_session_id(brand_name: str = "") -> str:
    """Verification session IDs for Perpetual Stock (web upload and mobile mirror)."""
    return await _next_mops_verification_session_id()


def _resolve_verification_part_name(product: dict) -> str:
    if not product:
        return ""
    return str(
        product.get("part_name")
        or product.get("item_name")
        or product.get("description")
        or product.get("part_description")
        or ""
    ).strip()


async def _enrich_verification_item_part_name(item: dict) -> dict:
    """API-only enrichment for legacy rows missing part_name (no MongoDB bulk update)."""
    if not item:
        return item
    if str(item.get("part_name") or "").strip():
        return item
    part_number = str(item.get("part_number") or "").strip()
    brand_name = item.get("brand_name")
    dealer_name = item.get("dealer_name")
    branch = item.get("branch")
    if not (part_number and brand_name and dealer_name and branch):
        return item
    product = await db.products.find_one(
        {
            "brand_name": brand_name,
            "dealer_name": dealer_name,
            "branch": branch,
            "part_number": {"$regex": f"^{re.escape(part_number)}$", "$options": "i"},
        },
        {"_id": 0},
        sort=[("updated_at", -1)],
    )
    resolved = _resolve_verification_part_name(product)
    if not resolved:
        return item
    enriched = dict(item)
    enriched["part_name"] = resolved
    return enriched


def _verification_calculations(system_qty: float, physical_qty: float, mav: float):
    difference = physical_qty - system_qty
    shortage_qty = abs(difference) if difference < 0 else 0.0
    excess_qty = difference if difference > 0 else 0.0
    return {
        "difference": difference,
        "shortage_qty": shortage_qty,
        "excess_qty": excess_qty,
        "shortage_value": shortage_qty * mav,
        "excess_value": excess_qty * mav,
        "quantity_status": "matched" if difference == 0 else ("shortage" if difference < 0 else "excess"),
    }


async def _build_verification_record(item, scope_query, current_user, session_id, now):
    clean_part = (item.part_number or "").strip()
    if not clean_part:
        raise HTTPException(status_code=400, detail="Part number is required for every item")
    query = dict(scope_query)
    query["part_number"] = {"$regex": f"^{re.escape(clean_part)}$", "$options": "i"}
    product = await db.products.find_one(query, {"_id": 0}, sort=[("updated_at", -1)])
    if not product:
        raise HTTPException(status_code=404, detail=f"Part {clean_part} not found in the selected Dashboard scope")
    system_qty = _stock_snapshot_quantity(product)
    physical_qty = float(item.physical_quantity)
    mav = float(product.get("mav") or product.get("MAV") or product.get("unit_value") or product.get("value") or 0)
    calculations = _verification_calculations(system_qty, physical_qty, mav)
    system_location = str(product.get("loc") or product.get("LOC") or product.get("bin_location") or product.get("location") or product.get("rack_location") or "").strip()
    physical_location = (item.scanned_location or "").strip()
    location_status = "matched" if system_location.casefold() == physical_location.casefold() else "mismatch"
    if calculations["quantity_status"] == "matched" and location_status == "matched":
        overall_status, correction_status = "matched", "not_required"
    elif calculations["quantity_status"] != "matched" and location_status != "matched":
        overall_status, correction_status = "quantity_and_location_mismatch", "pending"
    elif calculations["quantity_status"] != "matched":
        overall_status, correction_status = "quantity_mismatch", "pending"
    else:
        overall_status, correction_status = "location_mismatch", "pending"
    return {
        "id": str(uuid.uuid4()), "session_id": session_id,
        "part_number": product.get("part_number") or clean_part,
        "part_name": _resolve_verification_part_name(product),
        "product_id": product.get("id"), "mav": mav,
        "system_quantity": system_qty, "physical_quantity": physical_qty,
        **calculations,
        "system_location": system_location, "pin_location": system_location,
        "physical_location": physical_location, "scanned_location": physical_location,
        "location_status": location_status, "overall_status": overall_status,
        "remarks": (item.remarks or "").strip(),
        "brand_name": product.get("brand_name"), "dealer_name": product.get("dealer_name"), "branch": product.get("branch"),
        "verified_by": current_user.id, "verified_by_name": current_user.username,
        "snapshot_at": now, "created_at": now, "status": "submitted",
        "correction_status": correction_status, "correction_method": "", "correction_remarks": "",
        "information_only": True, "affects_stock": False,
    }


async def _get_or_create_web_physical_session(
    current_user: UserResponse,
    brand: str,
    dealer: str,
    branch: str,
) -> str:
    """One Physical Perpetual web session per NMTS user + branch + IST day."""
    india_now = datetime.now(ZoneInfo("Asia/Kolkata"))
    verification_date = india_now.strftime("%Y-%m-%d")
    scope = {
        "session_kind": "physical_web",
        "verification_date": verification_date,
        "verified_by": current_user.id,
        "brand_name": brand.strip(),
        "dealer_name": dealer.strip(),
        "branch": branch.strip(),
        "status": "ACTIVE",
    }
    existing = await db.stock_verification_sessions.find_one(scope, {"_id": 0, "session_id": 1})
    if existing and existing.get("session_id"):
        return existing["session_id"]
    session_id = await _next_mops_verification_session_id()
    now = datetime.now(timezone.utc)
    session_doc = {
        "id": str(uuid.uuid4()),
        "session_id": session_id,
        "session_kind": "physical_web",
        "verification_type": "physical",
        "verification_date": verification_date,
        "brand_name": brand.strip(),
        "dealer_name": dealer.strip(),
        "branch": branch.strip(),
        "verified_by": current_user.id,
        "verified_by_name": current_user.username,
        "status": "ACTIVE",
        "total_items": 0,
        "source": "WEB_PHYSICAL",
        "information_only": True,
        "affects_stock": False,
        "created_at": now,
        "updated_at": now,
    }
    try:
        await db.stock_verification_sessions.insert_one(session_doc)
    except Exception:
        existing = await db.stock_verification_sessions.find_one(scope, {"_id": 0, "session_id": 1})
        if existing and existing.get("session_id"):
            return existing["session_id"]
        raise
    return session_id


@api_router.post("/mobile/perpetual-stock/sessions")
async def create_mobile_stock_verification_session(payload: StockVerificationSessionCreate, current_user: UserResponse = Depends(get_current_user)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Add at least one verified item before upload")
    if len(payload.items) > 5000:
        raise HTTPException(status_code=400, detail="A verification session cannot exceed 5000 items")
    scope_query = _mobile_dashboard_scope_query(current_user, payload.brand, payload.dealer, payload.branch)
    session_id = await _get_or_create_web_physical_session(
        current_user, payload.brand, payload.dealer, payload.branch
    )
    now = datetime.now(timezone.utc)
    records = []
    seen = set()
    for item in payload.items:
        key = (item.part_number or "").strip().casefold()
        if key in seen:
            raise HTTPException(status_code=400, detail=f"Duplicate part number in upload list: {item.part_number}")
        seen.add(key)
        records.append(await _build_verification_record(item, scope_query, current_user, session_id, now))
    totals = {
        "total_items": len(records),
        "total_shortage_qty": sum(r["shortage_qty"] for r in records),
        "total_shortage_value": sum(r["shortage_value"] for r in records),
        "total_excess_qty": sum(r["excess_qty"] for r in records),
        "total_excess_value": sum(r["excess_value"] for r in records),
    }
    if records:
        await db.stock_verifications.insert_many([r.copy() for r in records])
    await db.stock_verification_sessions.update_one(
        {"session_id": session_id},
        {
            "$inc": {
                "total_items": totals["total_items"],
                "total_shortage_qty": totals["total_shortage_qty"],
                "total_shortage_value": totals["total_shortage_value"],
                "total_excess_qty": totals["total_excess_qty"],
                "total_excess_value": totals["total_excess_value"],
            },
            "$set": {"updated_at": now, "status": "submitted"},
        },
    )
    session = await db.stock_verification_sessions.find_one({"session_id": session_id}, {"_id": 0})
    return {**(session or {}), "items": records}


@api_router.get("/mobile/perpetual-stock/sessions")
async def list_mobile_stock_verification_sessions(
    brand: str, dealer: str, branch: str, date_from: str = None, date_to: str = None,
    user_filter: str = None, status_filter: str = None,
    current_user: UserResponse = Depends(get_current_user),
):
    query = _mobile_dashboard_scope_query(current_user, brand, dealer, branch)
    if user_filter:
        query["verified_by"] = user_filter
    if status_filter and status_filter != "all":
        query["status"] = status_filter
    if date_from or date_to:
        created = {}
        if date_from:
            created["$gte"] = datetime.fromisoformat(date_from).replace(tzinfo=ZoneInfo("Asia/Kolkata")).astimezone(timezone.utc)
        if date_to:
            created["$lt"] = (datetime.fromisoformat(date_to).replace(tzinfo=ZoneInfo("Asia/Kolkata")) + timedelta(days=1)).astimezone(timezone.utc)
        query["created_at"] = created
    return await db.stock_verification_sessions.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=10000)


@api_router.get("/mobile/perpetual-stock/verification-history")
async def list_verification_history_records(
    brand: str,
    dealer: str,
    branch: str,
    date_from: str = None,
    date_to: str = None,
    month: str = None,
    mobile_user_id: str = None,
    verification_type: str = None,
    result_filter: str = None,
    part_number: str = None,
    loc: str = None,
    limit: int = 500,
    current_user: UserResponse = Depends(get_current_user),
):
    """Line-level audit history (Physical / Auto / Recheck) with filters."""
    query = _mobile_dashboard_scope_query(current_user, brand, dealer, branch)
    if month and len(month) == 7:
        y, m = int(month[:4]), int(month[5:7])
        start = datetime(y, m, 1, tzinfo=ZoneInfo("Asia/Kolkata")).astimezone(timezone.utc)
        end_month = m + 1 if m < 12 else 1
        end_year = y if m < 12 else y + 1
        end = datetime(end_year, end_month, 1, tzinfo=ZoneInfo("Asia/Kolkata")).astimezone(timezone.utc)
        query["verified_at"] = {"$gte": start, "$lt": end}
    elif date_from or date_to:
        created = {}
        if date_from:
            created["$gte"] = datetime.fromisoformat(date_from).replace(tzinfo=ZoneInfo("Asia/Kolkata")).astimezone(timezone.utc)
        if date_to:
            created["$lt"] = (datetime.fromisoformat(date_to).replace(tzinfo=ZoneInfo("Asia/Kolkata")) + timedelta(days=1)).astimezone(timezone.utc)
        query["verified_at"] = created
    if mobile_user_id:
        query["mobile_user_id"] = mobile_user_id
    if part_number:
        query["part_number"] = {"$regex": f"^{re.escape(part_number.strip())}$", "$options": "i"}
    if loc:
        query["$or"] = [{"pin_location": loc}, {"system_location": loc}, {"location": loc}]
    if verification_type and verification_type.lower() != "all":
        vt = verification_type.lower()
        if vt == "recheck":
            query["coverage_kind"] = "recheck"
        else:
            query["verification_type"] = vt
    if result_filter and result_filter.lower() != "all":
        rf = result_filter.lower()
        if rf == "match":
            query["quantity_status"] = "matched"
        elif rf == "shortage":
            query["quantity_status"] = "shortage"
        elif rf == "excess":
            query["quantity_status"] = "excess"
        elif rf == "damage":
            query["has_damage"] = True
    rows = await db.stock_verification_history.find(query, {"_id": 0}).sort("verified_at", -1).limit(min(limit, 2000)).to_list(2000)
    return rows


@api_router.get("/mobile/perpetual-stock/export-all/excel")
async def export_all_mobile_stock_verifications(
    brand: str = None, dealer: str = None, branch: str = None,
    date_from: str = None, date_to: str = None,
    current_user: UserResponse = Depends(get_current_user),
):
    from fastapi.responses import StreamingResponse
    excel_permissions.require_excel_export(current_user)
    """Download the complete permitted Perpetual Stock list.

    Master: all brands/dealers/branches (optional dashboard filters).
    Admin: fixed brand/dealer and every branch (optional branch filter).
    User: own branch only.
    """
    query = {}
    exact = lambda value: value and value != "N/A" and not str(value).startswith("All ")
    if current_user.role == "master":
        if exact(brand): query["brand_name"] = _exact_ci(brand)
        if exact(dealer): query["dealer_name"] = _exact_ci(dealer)
        if exact(branch): query["branch"] = _exact_ci(branch)
    elif current_user.role == "admin":
        if current_user.brand: query["brand_name"] = _exact_ci(current_user.brand)
        if current_user.group: query["dealer_name"] = _exact_ci(current_user.group)
        if exact(branch): query["branch"] = _exact_ci(branch)
    else:
        if current_user.brand: query["brand_name"] = _exact_ci(current_user.brand)
        if current_user.group: query["dealer_name"] = _exact_ci(current_user.group)
        if current_user.location: query["branch"] = _exact_ci(current_user.location)

    if date_from or date_to:
        created = {}
        if date_from:
            created["$gte"] = datetime.fromisoformat(date_from).replace(tzinfo=ZoneInfo("Asia/Kolkata")).astimezone(timezone.utc)
        if date_to:
            created["$lt"] = (datetime.fromisoformat(date_to).replace(tzinfo=ZoneInfo("Asia/Kolkata")) + timedelta(days=1)).astimezone(timezone.utc)
        query["created_at"] = created

    rows = await db.stock_verifications.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=200000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Perpetual Stock"
    headers = ["Date","Session ID","Brand","Dealer / Group","Branch","Verified By","Mobile User ID","Part Number","Part Name","MAV","System Qty","Physical Qty","Difference","Shortage Qty","Excess Qty","Shortage Value","Excess Value","System / PIN Location","Physical Location","Verification Status","Correction Status","Correction Method","Correction Remarks","Corrected By","Corrected At","New / Unlisted Part","Entry Method","Source","Remarks"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    for item in rows:
        created = item.get("created_at") or item.get("verified_at")
        if isinstance(created, datetime):
            created = created.astimezone(ZoneInfo("Asia/Kolkata")).strftime("%d-%m-%Y %I:%M %p")
        ws.append([
            created, item.get("session_id"), item.get("brand_name"), item.get("dealer_name"), item.get("branch"),
            item.get("verified_by_name") or item.get("verified_user"), item.get("mobile_user_id"), item.get("part_number"),
            item.get("part_name"), item.get("mav", 0), item.get("system_quantity", 0), item.get("physical_quantity", 0),
            item.get("difference", 0), item.get("shortage_qty", 0), item.get("excess_qty", 0), item.get("shortage_value", 0),
            item.get("excess_value", 0), item.get("pin_location") or item.get("system_location"),
            item.get("physical_location") or item.get("scanned_location") or item.get("location"),
            item.get("overall_status") or item.get("verification_status") or item.get("quantity_status"),
            item.get("correction_status"), item.get("correction_method"), item.get("correction_remarks"),
            item.get("correction_updated_by_name"), item.get("correction_updated_at"),
            "Yes" if item.get("is_new_part") else "No", item.get("entry_method"), item.get("source") or "WEB",
            item.get("remarks") or item.get("remark"),
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 42)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = width
    output = BytesIO(); wb.save(output); output.seek(0)
    filename = "perpetual_stock_master.xlsx" if current_user.role == "master" else ("perpetual_stock_admin.xlsx" if current_user.role == "admin" else "perpetual_stock_branch.xlsx")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@api_router.get("/mobile/perpetual-stock/sessions/{session_id}")
async def get_mobile_stock_verification_session(session_id: str, current_user: UserResponse = Depends(get_current_user)):
    session = await db.stock_verification_sessions.find_one({"session_id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Verification session not found")
    scope = _verification_scope_query(current_user, session.get("brand_name"), session.get("dealer_name"), session.get("branch"))
    if scope and not await db.stock_verification_sessions.find_one({"session_id": session_id, **scope}, {"_id": 0, "session_id": 1}):
        raise HTTPException(status_code=403, detail="Verification session is outside your permitted scope")
    items = await db.stock_verifications.find({"session_id": session_id}, {"_id": 0}).sort("created_at", 1).to_list(length=10000)
    enriched_items = [await _enrich_verification_item_part_name(item) for item in items]
    return {**session, "items": enriched_items}


@api_router.get("/mobile/perpetual-stock/sessions/{session_id}/excel")
async def export_mobile_stock_verification_session(session_id: str, current_user: UserResponse = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    excel_permissions.require_excel_export(current_user)
    data = await get_mobile_stock_verification_session(session_id, current_user)
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary_rows = [
        ("Session ID", data.get("session_id")), ("Date", data.get("created_at").astimezone(ZoneInfo("Asia/Kolkata")).strftime("%d-%m-%Y %I:%M %p") if data.get("created_at") else ""),
        ("Brand", data.get("brand_name")), ("Dealer", data.get("dealer_name")), ("Branch", data.get("branch")), ("User", data.get("verified_by_name")),
        ("Total Items", data.get("total_items", 0)), ("Total Shortage Qty", data.get("total_shortage_qty", 0)),
        ("Total Shortage Value", data.get("total_shortage_value", 0)), ("Total Excess Qty", data.get("total_excess_qty", 0)), ("Total Excess Value", data.get("total_excess_value", 0)),
    ]
    for r, (label, value) in enumerate(summary_rows, 1):
        summary.cell(r, 1, label).font = openpyxl.styles.Font(bold=True)
        summary.cell(r, 2, value)
    details = wb.create_sheet("Details")
    headers = ["Part Number","Part Name","MAV","System Qty","Physical Qty","Difference","Shortage Qty","Excess Qty","Shortage Value","Excess Value","PIN Location","Physical Location","Status","Correction Status","Correction Method","Correction Remarks","Corrected By","Corrected At","New / Unlisted Part","Entry Method","Source","Remarks"]
    details.append(headers)
    for cell in details[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    for item in data.get("items", []):
        details.append([item.get("part_number"), item.get("part_name"), item.get("mav",0), item.get("system_quantity",0), item.get("physical_quantity",0), item.get("difference",0), item.get("shortage_qty",0), item.get("excess_qty",0), item.get("shortage_value",0), item.get("excess_value",0), item.get("pin_location") or item.get("system_location"), item.get("physical_location") or item.get("scanned_location"), item.get("overall_status"), item.get("correction_status"), item.get("correction_method"), item.get("correction_remarks"), item.get("correction_updated_by_name"), item.get("correction_updated_at"), "Yes" if item.get("is_new_part") else "No", item.get("entry_method"), item.get("source") or "WEB", item.get("remarks") or item.get("remark")])
    for ws in (summary, details):
        for col in ws.columns:
            width = min(max(len(str(c.value or "")) for c in col) + 2, 40)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = width
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{session_id}.xlsx"'})


@api_router.post("/mobile/perpetual-stock")
async def create_mobile_stock_verification(payload: StockVerificationCreate, current_user: UserResponse = Depends(get_current_user)):
    query = _mobile_dashboard_scope_query(
        current_user,
        payload.brand,
        payload.dealer,
        payload.branch,
    )
    clean_part = payload.part_number.strip()
    query["part_number"] = {"$regex": f"^{re.escape(clean_part)}$", "$options": "i"}
    product = await db.products.find_one(query, {"_id": 0}, sort=[("updated_at", -1)])
    if not product:
        raise HTTPException(status_code=404, detail="Part not found in your permitted stock scope")

    # IMPORTANT: This is an immutable information-only snapshot. No product,
    # ledger, order, reservation, quantity or location record is updated here.
    system_qty = _stock_snapshot_quantity(product)
    physical_qty = float(payload.physical_quantity)
    difference = physical_qty - system_qty
    system_location = str(
        product.get("loc")
        or product.get("LOC")
        or product.get("bin_location")
        or product.get("location")
        or product.get("rack_location")
        or ""
    ).strip()
    physical_location = (payload.scanned_location or "").strip()
    quantity_status = "matched" if difference == 0 else ("shortage" if difference < 0 else "excess")
    location_status = "matched" if system_location.casefold() == physical_location.casefold() else "mismatch"
    if quantity_status == "matched" and location_status == "matched":
        overall_status = "matched"
        correction_status = "not_required"
    elif quantity_status != "matched" and location_status != "matched":
        overall_status = "quantity_and_location_mismatch"
        correction_status = "pending"
    elif quantity_status != "matched":
        overall_status = "quantity_mismatch"
        correction_status = "pending"
    else:
        overall_status = "location_mismatch"
        correction_status = "pending"

    now = datetime.now(timezone.utc)
    record = {
        "id": str(uuid.uuid4()),
        "part_number": product.get("part_number") or clean_part,
        "part_name": _resolve_verification_part_name(product),
        "product_id": product.get("id"),
        "system_quantity": system_qty,
        "physical_quantity": physical_qty,
        "difference": difference,
        "quantity_status": quantity_status,
        "system_location": system_location,
        "physical_location": physical_location,
        "scanned_location": physical_location,
        "location_status": location_status,
        "overall_status": overall_status,
        "remarks": (payload.remarks or "").strip(),
        "brand_name": product.get("brand_name") or payload.brand or current_user.brand,
        "dealer_name": product.get("dealer_name") or payload.dealer or current_user.group,
        "branch": product.get("branch") or payload.branch or current_user.location,
        "verified_by": current_user.id,
        "verified_by_name": current_user.username,
        "snapshot_at": now,
        "created_at": now,
        "status": "submitted",
        "correction_status": correction_status,
        "correction_method": "",
        "correction_remarks": "",
        "information_only": True,
        "affects_stock": False,
    }
    await db.stock_verifications.insert_one(record.copy())
    record.pop("_id", None)
    return record


@api_router.get("/mobile/perpetual-stock")
async def list_mobile_stock_verifications(
    brand: str = None, dealer: str = None, branch: str = None,
    status_filter: str = None, part_number: str = None,
    current_user: UserResponse = Depends(get_current_user),
):
    query = _verification_scope_query(current_user, brand, dealer, branch)
    if status_filter and status_filter != "all":
        query["overall_status"] = status_filter
    if part_number:
        query["part_number"] = {"$regex": re.escape(part_number.strip()), "$options": "i"}
    return await db.stock_verifications.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=1000)


@api_router.put("/mobile/perpetual-stock/{verification_id}/correction")
async def update_stock_verification_correction(
    verification_id: str,
    payload: StockVerificationCorrection,
    current_user: UserResponse = Depends(get_current_user),
):
    if current_user.role not in {"master", "admin"}:
        raise HTTPException(status_code=403, detail="Only Master/Admin can update correction information")
    record = await db.stock_verifications.find_one({"id": verification_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Verification record not found")
    scope_query = _verification_scope_query(current_user)
    if scope_query and not await db.stock_verifications.find_one({"id": verification_id, **scope_query}, {"_id": 0, "id": 1}):
        raise HTTPException(status_code=403, detail="Verification is outside your permitted scope")
    correction_status = (payload.correction_status or "pending").strip().lower()
    correction_method = (payload.correction_method or "").strip().lower()
    if correction_status not in {"pending", "corrected", "not_required"}:
        raise HTTPException(status_code=400, detail="Invalid correction status")
    if correction_method not in {"", "system_corrected", "physical_relocated", "both", "no_action"}:
        raise HTTPException(status_code=400, detail="Invalid correction method")
    update = {
        "correction_status": correction_status,
        "correction_method": correction_method,
        "correction_remarks": (payload.correction_remarks or "").strip(),
        "correction_updated_by": current_user.id,
        "correction_updated_by_name": current_user.username,
        "correction_updated_at": datetime.now(timezone.utc),
    }
    await db.stock_verifications.update_one({"id": verification_id}, {"$set": update})
    return await db.stock_verifications.find_one({"id": verification_id}, {"_id": 0})



# Unified Reports Center
try:
    from . import reports_center
except ImportError:
    import reports_center
reports_center.init_reports_center(db, get_current_user, UserResponse)
api_router.include_router(reports_center.router)

try:
    from . import query_desk
except ImportError:
    import query_desk
query_desk.init_query_desk(db, get_current_user, UserResponse)
api_router.include_router(query_desk.router)

try:
    from . import notice_board
except ImportError:
    import notice_board
notice_board.init_notice_board(db, get_current_user, UserResponse)
api_router.include_router(notice_board.router)

try:
    from . import analytics_center
except ImportError:
    import analytics_center
analytics_center.init_analytics_center(db, get_current_user, UserResponse, _nmts_date_key, _nmts_now)
api_router.include_router(analytics_center.router)

# Sleeping Stock Mobile (Mobile User + device-session model)
try:
    from . import mobile_api
except ImportError:
    import mobile_api
mobile_api.init_mobile_api(
    db,
    get_current_user,
    UserResponse,
    pwd_context,
    _request_center_transition,
    _notify_request_status_change,
)
api_router.include_router(mobile_api.router)


# ==================== HYBRID STORAGE ADMIN / OPS (MASTER ONLY MONITOR) ====================

@api_router.get("/storage/status")
async def storage_status(current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    return s3_storage.get_storage().status()


@api_router.get("/storage/monitor")
async def storage_cost_monitor(month: str = None, brand: str = None, dealer: str = None, current_user: UserResponse = Depends(get_current_user)):
    """Master-only Storage & Cost Monitor dashboard payload."""
    await _ensure_master(current_user)
    import storage_monitor as sm
    import storage_usage as su

    payload = await sm.monitor_dashboard(db, month=month)
    if brand or dealer:
        payload["dealer_ranking"] = await su.dealer_usage_ranking(db, month=month, brand=brand, dealer=dealer)
    return payload


@api_router.get("/storage/monitor/dealers")
async def storage_dealer_ranking(month: str = None, brand: str = None, dealer: str = None, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    import storage_usage as su
    return await su.dealer_usage_ranking(db, month=month, brand=brand, dealer=dealer)


@api_router.get("/storage/monitor/migration-report")
async def storage_migration_report(current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    import storage_monitor as sm
    return await sm.migration_space_report(db)


@api_router.post("/storage/migration/archive-dates")
async def storage_migration_archive_dates(
    dry_run: bool = True,
    prune_after: bool = False,
    limit: int = None,
    dates: list = None,
    current_user: UserResponse = Depends(get_current_user),
):
    """Master-only date-by-date historical archive. prune_after still gated by REAL S3 + flag."""
    await _ensure_master(current_user)
    return await history_archive.archive_historical_dates(
        db,
        dates=dates,
        dry_run=dry_run,
        prune_after=prune_after,
        limit=limit,
    )


@api_router.post("/storage/migration/cleanup-published-upload-items")
async def cleanup_published_upload_items_api(dry_run: bool = True, current_user: UserResponse = Depends(get_current_user)):
    """Master-only: remove obsolete Published staging rows (keeps Waiting/pending)."""
    await _ensure_master(current_user)
    return await history_archive.cleanup_published_upload_items(db, dry_run=dry_run)


@api_router.post("/storage/archives/product-history/run")
async def run_product_history_archive(archive_date: str = None, force: bool = False, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    if archive_date:
        return await history_archive.archive_product_history_for_date(db, archive_date, force=force)
    return await archive_scheduler.run_daily_product_archive(db, archive_date)


@api_router.post("/storage/archives/product-history/prune")
async def prune_product_history_archive(archive_date: str, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    return await history_archive.prune_product_history_date(db, archive_date)


@api_router.post("/storage/archives/orders-requests/run")
async def run_orders_requests_archive(archive_month: str = None, force: bool = False, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    if force and archive_month:
        orders = await history_archive.archive_completed_orders_month(db, archive_month, force=True)
        requests = await history_archive.archive_completed_requests_month(db, archive_month, force=True)
        return {"orders": orders, "requests": requests}
    return await archive_scheduler.run_monthly_completed_archives(db, archive_month)


@api_router.post("/storage/archives/verifications/run")
async def run_verification_archive(archive_date: str, force: bool = False, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    return await history_archive.archive_verifications_for_date(db, archive_date, force=force)


@api_router.get("/storage/archives")
async def list_archives(module: str = None, current_user: UserResponse = Depends(get_current_user)):
    await _ensure_master(current_user)
    q = {}
    if module:
        q["module"] = module
    return await db.archive_manifests.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.post("/storage/archives/{archive_id}/retry")
async def retry_failed_archive(archive_id: str, current_user: UserResponse = Depends(get_current_user)):
    """Idempotent retry for FAILED archive jobs (Master only)."""
    await _ensure_master(current_user)
    row = await db.archive_manifests.find_one({"archive_id": archive_id}, {"_id": 0})
    if not row:
        raise HTTPException(status_code=404, detail="Archive job not found")
    if row.get("status") not in {archive_manifest.STATUS_FAILED, archive_manifest.STATUS_CREATING, archive_manifest.STATUS_UPLOADED}:
        return {"status": "skipped", "reason": f"status={row.get('status')} is not retryable", "manifest": row}
    module = row.get("module")
    if module == history_archive.MODULE_PRODUCT_HISTORY:
        return await history_archive.archive_product_history_for_date(db, row.get("archive_date"), force=True)
    if module == history_archive.MODULE_ORDERS:
        return await history_archive.archive_completed_orders_month(db, row.get("archive_month"), force=True)
    if module == history_archive.MODULE_REQUESTS:
        return await history_archive.archive_completed_requests_month(db, row.get("archive_month"), force=True)
    if module == history_archive.MODULE_VERIFICATIONS:
        return await history_archive.archive_verifications_for_date(db, row.get("archive_date"), force=True)
    raise HTTPException(status_code=400, detail=f"Unsupported module for retry: {module}")

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', 'http://localhost:3000,http://127.0.0.1:3000').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

async def ensure_product_hub_indexes():
    """Indexes required to keep Product Hub fast at lakhs-of-records scale.
    Safe to call every startup; create_index is a no-op if the index already exists."""
    await db.products.create_index([("brand_name", 1), ("dealer_name", 1), ("branch", 1)])
    await db.order_requests.create_index([("order_id", 1)])
    await db.order_requests.create_index([("status", 1)])
    await db.order_requests.create_index([("requesting_dealer", 1)])
    await db.order_requests.create_index([("supplying_dealer", 1), ("supplying_branch", 1)])
    await db.order_requests.create_index([("requested_at", -1)])
    await db.notification_logs.create_index([("request_id", 1), ("created_at", -1)])
    await db.notification_logs.create_index([("request_number", 1)])
    await db.notification_logs.create_index([("receiver_email", 1)])
    await db.notification_logs.create_index([("status", 1)])
    await db.notification_logs.create_index([("created_at", -1)])
    # Parts Transfer Request grouping.
    # Historical retries (Rejected / Cancelled / Approved / Dispatched) may
    # legitimately create multiple request_headers for the same
    # (order, supplying dealer, branch). A full unique index on that key
    # fails against shared Atlas data and blocks Branch→Dealer retry flow.
    # Keep uniqueness ONLY for active "Requested" groups so concurrent
    # duplicate sends are still prevented at the DB layer.
    try:
        await db.request_headers.drop_index('order_id_1_supplying_dealer_1_supplying_branch_1')
    except Exception:
        pass
    try:
        # Older name if previously created under a custom name.
        await db.request_headers.drop_index('uniq_request_destination')
    except Exception:
        pass
    await db.request_headers.create_index(
        [("order_id", 1), ("supplying_dealer", 1), ("supplying_branch", 1)],
        unique=True,
        name="uniq_active_request_destination",
        partialFilterExpression={"status": "Requested"},
    )
    await db.request_headers.create_index(
        [("order_id", 1), ("supplying_dealer", 1), ("supplying_branch", 1), ("created_at", -1)],
        name="idx_request_headers_destination_history",
    )
    await db.request_headers.create_index([("request_number", 1)], unique=True)
    await db.request_headers.create_index([("created_at", -1)])
    await db.order_requests.create_index([("request_number", 1)])
    await db.order_requests.create_index([("request_group_id", 1)])
    await db.products.create_index([("is_active_today", 1), ("active_date_key", 1)])
    await db.products.create_index([("publish_status", 1), ("active_date_key", 1)])
    await db.products.create_index([("part_number", 1)])
    await db.products.create_index([("upload_id", 1)])
    await db.products.create_index([("upload_no", 1)])
    await db.products.create_index([("active_date_key", 1)])
    await db.products.create_index([("location", 1)])
    await db.stock_verifications.create_index([("created_at", -1)])
    await db.stock_verifications.create_index([("brand_name", 1), ("dealer_name", 1), ("branch", 1), ("created_at", -1)])
    await db.stock_verifications.create_index([("part_number", 1), ("created_at", -1)])
    await db.stock_verifications.create_index([("session_id", 1), ("created_at", 1)])
    await db.stock_verification_sessions.create_index([("session_id", 1)], unique=True)
    await db.stock_verification_sessions.create_index([("brand_name", 1), ("dealer_name", 1), ("branch", 1), ("created_at", -1)])
    await db.upload_items.create_index([("upload_id", 1)])
    await db.uploads.create_index([("date_key", 1)])
    await db.uploads.create_index([("brand_name", 1), ("dealer_name", 1), ("branch", 1)])
    await db.uploads.create_index([("upload_type", 1), ("created_at", -1)])
    await db.uploads.create_index([("created_at", -1)], name="idx_uploads_created_at")
    await db.batch_summaries.create_index(
        [("brand_name", 1), ("dealer_name", 1), ("branch", 1), ("active_date_key", 1)],
        unique=True,
    )
    # Smart Auto Suggest Allocation Engine: reservation-aware availability.
    await db.stock_reservations.create_index([("stock_id", 1), ("status", 1)])
    await db.stock_reservations.create_index([("order_request_id", 1)])
    await db.stock_reservations.create_index([("order_id", 1)])
    await odw.ensure_order_desk_indexes(db)


@app.on_event("startup")
async def seed_master_user_on_startup():
    try:
        result = await ensure_master_user_exists()
        logger.info(f"Startup init check: {result['message']}")
    except Exception as e:
        logger.error(f"Startup init check failed: {e}")
    try:
        await ensure_product_hub_indexes()
        logger.info("Product Hub indexes verified")
        await reports_center.ensure_indexes()
        logger.info("Reports Center indexes verified")
        await query_desk.ensure_indexes()
        logger.info("Query Desk indexes verified")
        await notice_board.ensure_indexes()
        logger.info("Notice Board indexes verified")
        await analytics_center.ensure_analytics_indexes()
        logger.info("Analytics indexes verified")
        await mobile_api.ensure_mobile_indexes()
        await archive_manifest.ensure_archive_indexes(db)
        logger.info("Archive manifest indexes verified")
        try:
            import storage_usage as su
            await su.ensure_usage_indexes(db)
            logger.info("Storage usage indexes verified")
        except Exception as exc:
            logger.warning("Storage usage index creation failed: %s", exc)
        archive_scheduler.start_archive_scheduler(db)
        logger.info("Archive scheduler started (ARCHIVE_PRUNE_ENABLED=%s)", s3_storage.archive_prune_enabled())
    except Exception as e:
        logger.error(f"Product Hub index creation failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        await archive_scheduler.stop_archive_scheduler()
    except Exception:
        pass
    client.close()
